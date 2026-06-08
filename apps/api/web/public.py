from __future__ import annotations

import json
import logging
import math
import os
import pathlib
import hashlib
import hmac
import time
import uuid
import random
import secrets
from datetime import datetime, timedelta, timezone
from typing import Any, Iterable, Mapping, Sequence

from fastapi import APIRouter, Request, UploadFile, BackgroundTasks, HTTPException, Query, File
from fastapi.responses import (
    JSONResponse,
    Response,
    HTMLResponse,
)
import httpx


from libs.core import catalog as catalog_module
from libs.core import catalog_index
from libs.core.catalog.pdf_catalog_miniprog import CatalogMiniPipeline

# NOTE: reference helpers locally to keep call sites compact
write_catalog_csv = catalog_module.write_catalog_csv
CatalogIndexError = catalog_index.CatalogIndexError

try:  # pragma: no cover - optional dependency during import time
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - openpyxl is optional in some environments
    load_workbook = None  # type: ignore[assignment]

from libs.core import sales_core as core_module  # type: ignore[attr-defined]
from libs.core.sales_core import _normalize_catalog_items, settings  # type: ignore[attr-defined]

from urllib.parse import quote, quote_plus, urlencode

from redis import exceptions as redis_ex

from . import client as C
from . import auth_utils
from .services import avito_oauth_runtime
from .services import avito_oauth_routes
from .services import avito_public_runtime
from .services import avito_webhook_runtime
from .services import amocrm_avatar_runtime
from .services import amocrm_public_runtime
from .services import catalog_file_parse_runtime
from .services import catalog_public_runtime
from .services import max_public_runtime
from .services import public_avatar_runtime
from .services import public_auth_runtime
from .services import client_assets_runtime
from .services import public_photos_runtime
from .services import public_request_runtime
from .services import settings_public_runtime
from .services import tg_public_runtime
from .services import tg_proxy_runtime
from .services import tg_slots_runtime
from .services import wa_public_runtime
from .services import wa_qr_runtime
from .services import avito_callback_html
from libs.core.message_envelope import content_fingerprint, text_or_placeholder
from libs.core.db import (
    find_lead_by_peer,
    get_lead_dialog_metadata,
    get_lead_peer,
    insert_message_out,
    list_messages_for_lead,
)
from libs.core.integrations import amocrm as amocrm_integration
from libs.core.integrations import avito
from libs.core.integrations import max as max_integration
from libs.core.transport import max_personal as max_personal_transport
from libs.core.repo import amocrm_tokens
from libs.core.repo import crm_links
from libs.core.services import amocrm as amocrm_service
from libs.core.services import amocrm_chat as amocrm_chat_service
from libs.core.services.avito_oauth_tokens import (
    AvitoTokenPayloadError,
    build_token_update_payload,
)
from libs.core.services.avito_oauth_state import resolve_tenant_from_state
from libs.core.services import avito_webhook_events
from libs.core.services import max_personal_service
from libs.core.services.tenant_config_merge import (
    build_public_settings_get_config,
    build_public_settings_save_config,
)
from libs.core.learning.service import capture_intervention_episode
from libs.core.repo import crm_chat_links
from libs.core import db as db_module
from libs.core.transport import telegram as telegram_transport
from libs.core.lib.numbers import coerce_int as _coerce_int_shared
from libs.core.lib.tg_slots import (
    TG_SLOT_MAX,
    TG_SLOT_MIN,
    normalize_tg_slot as _normalize_tg_slot_shared,
    virtual_tenant_id as _virtual_tenant_id_shared,
)
from libs.core.common import (
    AVITO_BOT_ECHO_TTL_SECONDS,
    HANDOFF_SILENCE_TTL_SECONDS,
    avito_bot_echo_key,
    handoff_silence_key,
    handoff_silence_meta_key,
    normalize_echo_text,
)
from . import common as common
from .client import read_csv_table, write_csv_table
from . import webhooks as webhook_module  # type: ignore
from .ui import render_template, templates as _templates
from .webhooks import process_incoming


class TgWorkerCallError(RuntimeError):
    """Raised when a call to the tgworker proxy fails."""

    def __init__(self, url: str, detail: str) -> None:
        self.url = url
        self.detail = detail
        message = f"{url}: {detail}" if detail else url
        super().__init__(message)


templates = _templates

logger = logging.getLogger(__name__)
wa_logger = logging.getLogger("wa")
# Unified incoming transport log channel
message_in_logger = logging.getLogger("app.web.message_in")


async def _capture_manager_intervention(
    *,
    tenant_id: int,
    lead_id: int,
    channel: str,
    manager_message_id: int | None,
    source_event: str,
) -> None:
    if not manager_message_id:
        return
    try:
        await capture_intervention_episode(
            tenant_id=int(tenant_id),
            lead_id=int(lead_id),
            channel=str(channel or ""),
            source_event=source_event,
            manager_message_id=int(manager_message_id),
            log_fn=lambda msg: logger.info(msg),
        )
    except Exception:
        logger.exception(
            "learning_v2_capture_failed tenant=%s lead_id=%s channel=%s source_event=%s",
            tenant_id,
            lead_id,
            channel,
            source_event,
        )
_deprecated_hits: dict[str, float] = {}
# Avoid duplicate logging of WA messages via root logger handlers
wa_logger.propagate = False

TG_WORKER_BASE = None  # will be resolved lazily via _tg_base_url()
if not hasattr(C, "valid_key"):
    setattr(C, "valid_key", common.valid_key)

# Redis queue handle (may be None in offline/demo mode)
_redis_queue = getattr(settings, "r", None)

NO_STORE_CACHE_VALUE = "no-store, must-revalidate"

PASSWORD_ATTEMPT_LIMIT = 2
PASSWORD_ATTEMPT_WINDOW = 60.0
_LOCAL_PASSWORD_ATTEMPTS: dict[tuple[int, str], list[float]] = {}

WA_QR_CACHE_TTL_MIN = 30  # seconds
WA_QR_CACHE_TTL_MAX = 60  # seconds

AVITO_STATE_PREFIX = "oauth:avito:state:"
AVITO_STATE_TTL = 4 * 3600  # seconds

router = APIRouter()
oauth_router = APIRouter(prefix="/v1/oauth/avito", tags=["avito_oauth"])
max_router = APIRouter(prefix="/v1/max", tags=["max"])
max_personal_router = APIRouter(prefix="/v1/max-personal", tags=["max_personal"])

CATALOG_VIEW_TEMPLATE = "catalog/view.html"
PHOTO_MAX_BYTES = 24 * 1024 * 1024  # Avito limit (24 MB) sets global cap
PHOTO_ALLOWED_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".heic"}
PHOTO_ALLOWED_MIMES = {
    "image/jpeg",
    "image/png",
    "image/gif",
    "image/bmp",
    "image/heic",
    "image/heif",
}


def _qr_cache_ttl() -> int:
    return random.randint(WA_QR_CACHE_TTL_MIN, WA_QR_CACHE_TTL_MAX)


INCOMING_QUEUE_KEY = getattr(webhook_module, "INCOMING_QUEUE_KEY", "inbox:message_in")


def _normalize_tg_slot(value: Any) -> int:
    return _normalize_tg_slot_shared(value)


def _tg_slot_tenant(tenant_id: int, slot: int) -> int:
    return _virtual_tenant_id_shared(tenant_id, slot)


def _tg_slots_deps() -> tg_slots_runtime.TgSlotsDeps:
    return tg_slots_runtime.TgSlotsDeps(
        slot_min=TG_SLOT_MIN,
        slot_max=TG_SLOT_MAX,
        normalize_slot_fn=_normalize_tg_slot,
    )


def _tg_slots_config(cfg: Mapping[str, Any] | None) -> dict[str, Any]:
    return tg_slots_runtime.tg_slots_config(cfg, _tg_slots_deps())


async def _is_recent_bot_echo(
    tenant: int,
    lead_id: int,
    text: str,
    *,
    window_seconds: int = 120,
) -> bool:
    if not text or tenant <= 0 or lead_id <= 0:
        return False
    try:
        messages = await list_messages_for_lead(tenant, lead_id, limit=10)
    except Exception:
        return False
    if not messages:
        return False
    now = datetime.now(timezone.utc)
    needle = normalize_echo_text(text)
    for msg in messages:
        if not msg or not msg.get("is_bot"):
            continue
        if int(msg.get("direction") or 0) != 1:
            continue
        msg_text = normalize_echo_text(str(msg.get("text") or ""))
        if not msg_text or msg_text != needle:
            continue
        ts = msg.get("created_at")
        if isinstance(ts, datetime):
            if ts.tzinfo is None:
                ts = ts.replace(tzinfo=timezone.utc)
            if (now - ts).total_seconds() <= window_seconds:
                return True
    return False


def _no_store_headers(extra: Mapping[str, str] | None = None) -> dict[str, str]:
    headers = {
        "Cache-Control": NO_STORE_CACHE_VALUE,
        "Pragma": "no-cache",
        "Expires": "0",
    }
    if extra:
        headers.update(extra)
    return headers


def _resolve_client_key(request: Request | None) -> str:
    candidates: list[str | None] = []
    if request is not None:
        query_params = getattr(request, "query_params", None)
        if query_params is not None:
            candidates.append(query_params.get("k"))
            candidates.append(query_params.get("key"))
        headers = getattr(request, "headers", {}) or {}
        for header_name in ("X-Access-Key", "X-Client-Key", "X-Auth-Key"):
            candidates.append(headers.get(header_name))
        auth_header = headers.get("Authorization")
        if auth_header:
            token = auth_header.strip()
            if token.lower().startswith("bearer "):
                token = token[7:]
            candidates.append(token)
        cookies = getattr(request, "cookies", None) or {}
        if cookies:
            candidates.append(cookies.get("client_key"))
    for candidate in candidates:
        if not candidate:
            continue
        value = str(candidate).strip()
        if value:
            return value
    return ""


def _avito_state_key(state: str) -> str:
    return avito_public_runtime.state_key(state, prefix=AVITO_STATE_PREFIX)


AVITO_STATE_COOKIE = "avito_oauth_state"


def _avito_state_secret() -> str:
    return avito_public_runtime.state_secret(settings)


def _b64url_encode(raw: bytes) -> str:
    return avito_public_runtime.b64url_encode(raw)


def _b64url_decode(raw: str) -> bytes:
    return avito_public_runtime.b64url_decode(raw)


def _build_avito_oauth_state(tenant_id: int) -> str:
    return avito_public_runtime.build_oauth_state(
        tenant_id,
        settings_module=settings,
        time_module=time,
        secrets_module=secrets,
    )


def _avito_state_cookie_domain() -> str | None:
    return avito_public_runtime.state_cookie_domain(settings)


def _avito_oauth_public_origin(request: Request) -> str:
    return avito_public_runtime.oauth_public_origin(
        request,
        settings_module=settings,
        public_base_url_fn=common.public_base_url,
    )


def _avito_oauth_redirect_entry_url(request: Request, tenant_id: int, key: str | None) -> str:
    return avito_public_runtime.oauth_redirect_entry_url(
        request,
        tenant_id,
        key,
        settings_module=settings,
        public_base_url_fn=common.public_base_url,
    )


def _set_avito_state_cookie(response: Response, request: Request, state: str) -> None:
    avito_public_runtime.set_state_cookie(
        response,
        request,
        state,
        settings_module=settings,
        cookie_name=AVITO_STATE_COOKIE,
        ttl_seconds=AVITO_STATE_TTL,
    )


def _clear_avito_state_cookie(response: Response) -> None:
    avito_public_runtime.clear_state_cookie(
        response,
        settings_module=settings,
        cookie_name=AVITO_STATE_COOKIE,
    )


def _verify_avito_oauth_state(state: str) -> dict[str, Any] | None:
    return avito_public_runtime.verify_oauth_state(
        state,
        settings_module=settings,
        ttl_seconds=AVITO_STATE_TTL,
        coerce_int_fn=_coerce_int,
        time_module=time,
    )


def _delete_avito_states_for_tenant(client: Any, tenant_id: int) -> int:
    """Keep one active OAuth state per tenant to avoid stale-tab callbacks."""
    return avito_public_runtime.delete_states_for_tenant(
        client,
        tenant_id,
        prefix=AVITO_STATE_PREFIX,
        coerce_int_fn=_coerce_int,
        json_module=json,
    )


def _amocrm_state_secret() -> str:
    return (settings.WEBHOOK_SECRET or "").strip() or (settings.ADMIN_TOKEN or "").strip()


async def _read_amocrm_webhook_payload(request: Request) -> dict[str, Any]:
    try:
        payload = await request.json()
        if isinstance(payload, dict):
            return payload
    except Exception:
        payload = {}
    try:
        form = await request.form()
        if form:
            return dict(form)
    except Exception:
        pass
    return {}


def _extract_amocrm_uninstall_info(payload: Mapping[str, Any]) -> tuple[int | None, str | None]:
    account_id = payload.get("account_id")
    subdomain = payload.get("subdomain") or payload.get("domain") or payload.get("base_domain")
    if isinstance(payload.get("account"), Mapping):
        account_id = payload["account"].get("id") or account_id
        subdomain = payload["account"].get("subdomain") or subdomain
    try:
        account_id_val = int(account_id) if account_id is not None else None
    except Exception:
        account_id_val = None
    subdomain_val = str(subdomain or "").strip() or None
    return account_id_val, subdomain_val


def _amocrm_public_runtime_deps() -> amocrm_public_runtime.AmoCRMPublicDeps:
    from apps.worker.main import send_avito  # local import to avoid circular startup edge

    return amocrm_public_runtime.AmoCRMPublicDeps(
        authorize_public_settings_request_fn=_authorize_public_settings_request,
        read_tenant_config_fn=common.read_tenant_config,
        write_tenant_config_fn=common.write_tenant_config,
        amocrm_service_module=amocrm_service,
        amocrm_integration_module=amocrm_integration,
        amocrm_tokens_module=amocrm_tokens,
        amocrm_chat_service_module=amocrm_chat_service,
        common_module=common,
        logger=logger,
        uuid_module=uuid,
        time_module=time,
        urlencode_fn=urlencode,
        state_secret_fn=_amocrm_state_secret,
        httpx_module=httpx,
        os_module=os,
        json_module=json,
        datetime_cls=datetime,
        timezone_utc=timezone.utc,
        timedelta_cls=timedelta,
        quote_plus_fn=quote_plus,
        no_store_headers_fn=_no_store_headers,
        read_amocrm_webhook_payload_fn=_read_amocrm_webhook_payload,
        extract_amocrm_uninstall_info_fn=_extract_amocrm_uninstall_info,
        crm_chat_links_module=crm_chat_links,
        crm_links_module=crm_links,
        db_module=db_module,
        get_lead_dialog_metadata_fn=get_lead_dialog_metadata,
        get_lead_peer_fn=get_lead_peer,
        content_fingerprint_fn=content_fingerprint,
        text_or_placeholder_fn=text_or_placeholder,
        redis_queue=_redis_queue,
        settings_module=settings,
        avito_bot_echo_key_fn=avito_bot_echo_key,
        avito_bot_echo_ttl_seconds=AVITO_BOT_ECHO_TTL_SECONDS,
        normalize_echo_text_fn=normalize_echo_text,
        telegram_transport_module=telegram_transport,
        insert_message_out_fn=insert_message_out,
        capture_manager_intervention_fn=_capture_manager_intervention,
        handoff_silence_key_fn=handoff_silence_key,
        handoff_silence_meta_key_fn=handoff_silence_meta_key,
        handoff_silence_ttl_seconds=HANDOFF_SILENCE_TTL_SECONDS,
        redis_error_type=redis_ex.RedisError,
        send_avito_fn=send_avito,
    )


def _amocrm_avatar_deps() -> amocrm_avatar_runtime.AmoCRMAvatarDeps:
    return amocrm_avatar_runtime.AmoCRMAvatarDeps(
        read_tenant_config_fn=common.read_tenant_config,
        amocrm_chat_service_module=amocrm_chat_service,
        hmac_module=hmac,
        tg_call_fn=_tg_call,
        tg_worker_call_error_type=TgWorkerCallError,
        no_store_headers_fn=_no_store_headers,
        chat_avatar_fn=chat_avatar,
        get_tenant_pubkey_fn=common.get_tenant_pubkey,
    )


def _public_settings_runtime_deps() -> settings_public_runtime.PublicSettingsDeps:
    return settings_public_runtime.PublicSettingsDeps(
        authorize_public_settings_request_fn=_authorize_public_settings_request,
        common_module=common,
        build_get_config_fn=build_public_settings_get_config,
        build_save_config_fn=build_public_settings_save_config,
        amocrm_service_module=amocrm_service,
        amocrm_tokens_module=amocrm_tokens,
        datetime_cls=datetime,
        timezone_utc=timezone.utc,
        logger=logger,
        no_store_headers_fn=_no_store_headers,
    )


def _avito_public_payload(raw: Mapping[str, Any] | None) -> dict[str, Any]:
    return avito_public_runtime.public_payload(raw)


def _avito_callback_html(ok: bool, message: str, payload: Mapping[str, Any]) -> str:
    return avito_callback_html.render_avito_callback_html(ok, message, payload)


@router.post("/webhook/avito")
async def avito_webhook(request: Request) -> JSONResponse:
    try:
        raw_payload = await request.json()
    except json.JSONDecodeError as exc:  # pragma: no cover - defensive
        raise HTTPException(status_code=422, detail="invalid_json") from exc
    except Exception as exc:
        raise HTTPException(status_code=422, detail="invalid_payload") from exc

    events = raw_payload if isinstance(raw_payload, list) else [raw_payload]
    processed = 0
    for entry in events:
        if not isinstance(entry, Mapping):
            continue
        try:
            handled = await _handle_avito_webhook_event(entry, request)
        except HTTPException:
            raise
        except Exception:  # pragma: no cover - defensive logging
            logger.exception("avito_webhook_processing_failed")
            continue
        if handled:
            processed += 1

    return JSONResponse({"ok": True, "processed": processed})


async def _handle_avito_webhook_event(event: Mapping[str, Any], request: Request) -> bool:
    return await avito_webhook_runtime.handle_avito_webhook_event(
        event,
        request,
        deps=avito_webhook_runtime.AvitoWebhookDeps(
            avito_webhook_events_module=avito_webhook_events,
            logger=logger,
            json_module=json,
            avito_module=avito,
            coerce_int_fn=_coerce_int,
            find_lead_by_peer_fn=find_lead_by_peer,
            redis_queue=_redis_queue,
            content_fingerprint_fn=content_fingerprint,
            avito_bot_echo_key_fn=avito_bot_echo_key,
            normalize_echo_text_fn=normalize_echo_text,
            is_recent_bot_echo_fn=_is_recent_bot_echo,
            time_module=time,
            handoff_silence_key_fn=handoff_silence_key,
            handoff_silence_meta_key_fn=handoff_silence_meta_key,
            handoff_silence_ttl_seconds=HANDOFF_SILENCE_TTL_SECONDS,
            db_module=db_module,
            insert_message_out_fn=insert_message_out,
            capture_manager_intervention_fn=_capture_manager_intervention,
            amocrm_service_module=amocrm_service,
            process_incoming_fn=process_incoming,
        ),
    )


async def _ensure_avito_webhook(tenant: int, request: Request) -> None:
    target_url = common.public_url(request, f"/webhook/avito?tenant={int(tenant)}")
    try:
        success = await avito.ensure_webhook(int(tenant), target_url)
    except avito.AvitoOAuthError as exc:
        logger.warning("avito_webhook_register_failed tenant=%s error=%s", tenant, exc)
    except Exception:
        logger.exception("avito_webhook_register_failed tenant=%s", tenant)
    else:
        if not success:
            logger.warning(
                "avito_webhook_register_failed tenant=%s error=unexpected_response", tenant
            )


@router.get("/connect/avito")
def connect_avito(tenant: int, request: Request, k: str | None = None, key: str | None = None):
    return avito_public_runtime.connect_avito(
        tenant,
        request,
        k=k,
        key=key,
        deps=avito_public_runtime.AvitoConnectDeps(
            common_module=common,
            avito_module=avito,
            logger=logger,
            render_template_fn=render_template,
            quote_plus_fn=quote_plus,
        ),
    )


def _tg_base_url() -> str:
    return tg_proxy_runtime.base_url(os, settings)


def _resolve_tg_base() -> str:
    global TG_WORKER_BASE
    TG_WORKER_BASE = tg_proxy_runtime.resolve_base(TG_WORKER_BASE, _tg_base_url())
    return TG_WORKER_BASE


def _tg_make_url(path: str) -> str:
    return tg_proxy_runtime.make_url(path, base=_resolve_tg_base())


_TG_HTTP_CLIENT: httpx.AsyncClient | None = None


def _tg_admin_headers() -> dict[str, str]:
    return tg_proxy_runtime.admin_headers(os, settings)


def _tg_client() -> httpx.AsyncClient:
    global _TG_HTTP_CLIENT
    _TG_HTTP_CLIENT = tg_proxy_runtime.client(_TG_HTTP_CLIENT, httpx)
    return _TG_HTTP_CLIENT


async def _tg_call(
    method: str,
    path: str,
    *,
    params: Mapping[str, Any] | None = None,
    json: Mapping[str, Any] | None = None,
    timeout: float = 5,
    route: str | None = None,
    peer: Any | None = None,
) -> tuple[int, httpx.Response]:
    return await tg_proxy_runtime.call(
        method,
        path,
        params=params,
        json_payload=json,
        timeout=timeout,
        route=route,
        peer=peer,
        deps=tg_proxy_runtime.TgProxyCallDeps(
            make_url_fn=_tg_make_url,
            admin_headers_fn=_tg_admin_headers,
            client_fn=_tg_client,
            httpx_module=httpx,
            logger=logger,
            worker_call_error_type=TgWorkerCallError,
        ),
    )


def _log_deprecated(route: str) -> None:
    now = time.time()
    last = _deprecated_hits.get(route)
    if last is None or now - last >= 3600:
        _deprecated_hits[route] = now
        logger.warning("deprecated_endpoint route=%s", route)


def _stringify_detail(value: bytes | bytearray | str | None) -> str:
    return tg_proxy_runtime.stringify_detail(value)


def _mask_sensitive_detail(detail: str | None) -> str:
    return tg_proxy_runtime.mask_sensitive_detail(detail)


def _extract_json_detail(body: bytes | bytearray | str | None) -> str | None:
    return tg_proxy_runtime.extract_json_detail(body, json_module=json)


def _log_tg_proxy(
    route: str,
    tenant: int | str | None,
    status: int,
    body: bytes | bytearray | str | None,
    *,
    error: str | None = None,
    force: bool | None = None,
) -> None:
    tg_proxy_runtime.log_tg_proxy(
        logger,
        route,
        tenant,
        status,
        body,
        error=error,
        force=force,
    )


def _fingerprint_public_key(raw: str | None) -> str:
    if not raw:
        return "-"
    try:
        digest = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    except Exception:
        return "-"
    return digest[:10]


def _log_public_tg_request(route: str, tenant_id: int, key: str | None) -> None:
    fingerprint = _fingerprint_public_key(_normalize_public_token(key))
    logger.info(
        "tg_public_request route=%s tenant=%s key=%s",
        route,
        tenant_id,
        fingerprint,
    )


def _parse_force_flag(raw_value: str | None) -> bool:
    if raw_value is None:
        return False
    value = raw_value.strip().lower()
    return value in {"1", "true", "yes", "on"}


def _password_attempt_key(tenant_id: int, token: str) -> str:
    digest = hashlib.sha256(token.encode("utf-8")).hexdigest()[:32]
    return f"tenant:{int(tenant_id)}:twofa_attempts:{digest}"


def _build_public_tg_qr_url(tenant_id: int, key: str | None, qr_id: str | None = None) -> str:
    parts: list[tuple[str, str]] = [("tenant", str(tenant_id))]
    normalized_key = _normalize_public_token(key)
    if normalized_key:
        parts.append(("k", normalized_key))
    if qr_id:
        parts.append(("qr_id", qr_id))
    return f"/pub/tg/qr.png?{urlencode(parts)}"


def _register_password_attempt(tenant_id: int, client_token: str) -> tuple[bool, int | None]:
    token = (client_token or "-").strip() or "-"
    key = _password_attempt_key(tenant_id, token)
    try:
        client = common.redis_client()
    except Exception:
        client = None

    if client is not None:
        try:
            pipe = client.pipeline()
            pipe.incr(key)
            pipe.ttl(key)
            attempts, ttl = pipe.execute()
            if attempts == 1 or ttl is None or ttl < 0:
                client.expire(key, int(PASSWORD_ATTEMPT_WINDOW))
                ttl = client.ttl(key)
            if attempts > PASSWORD_ATTEMPT_LIMIT:
                retry_after = int(ttl) if ttl and ttl > 0 else int(PASSWORD_ATTEMPT_WINDOW)
                return False, retry_after
            return True, None
        except redis_ex.RedisError:
            client = None

    now = time.monotonic()
    local_key = (int(tenant_id), token)
    entries = _LOCAL_PASSWORD_ATTEMPTS.setdefault(local_key, [])
    cutoff = now - PASSWORD_ATTEMPT_WINDOW
    filtered = [stamp for stamp in entries if stamp > cutoff]
    allowed = len(filtered) < PASSWORD_ATTEMPT_LIMIT
    retry_after: int | None = None
    if allowed:
        filtered.append(now)
    else:
        if filtered:
            remaining = PASSWORD_ATTEMPT_WINDOW - (now - filtered[0])
            retry_after = max(1, int(math.ceil(remaining))) if remaining > 0 else 1
        else:
            retry_after = int(PASSWORD_ATTEMPT_WINDOW)
    _LOCAL_PASSWORD_ATTEMPTS[local_key] = filtered
    return allowed, retry_after


def _client_identifier(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for", "")
    if forwarded:
        first = forwarded.split(",")[0].strip()
        if first:
            return first
    client = getattr(request, "client", None)
    host = getattr(client, "host", None) if client else None
    if host:
        return str(host)
    return "-"


MAX_UPLOAD_SIZE_BYTES = 25 * 1024 * 1024
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls", ".pdf"}
CSV_ENCODING_CANDIDATES = ["utf-8", "utf-8-sig", "cp1251", "windows-1251", "koi8-r"]


def _catalog_parse_deps() -> catalog_file_parse_runtime.CatalogParseDeps:
    return catalog_file_parse_runtime.CatalogParseDeps(
        encoding_candidates=CSV_ENCODING_CANDIDATES,
        load_workbook_fn=load_workbook,
        normalize_catalog_items_fn=_normalize_catalog_items,
        settings=settings,
        pipeline_cls=CatalogMiniPipeline,
        catalog_index_module=catalog_index,
        catalog_index_error=CatalogIndexError,
        logger=logger,
    )


def _coerce_tenant(raw: int | str | None) -> int:
    if raw is None:
        raise ValueError("missing_tenant")
    if isinstance(raw, str):
        raw = raw.strip()
        if not raw:
            raise ValueError("missing_tenant")
    try:
        return int(raw)
    except (TypeError, ValueError) as exc:
        raise ValueError("invalid_tenant") from exc


def _normalize_headers(raw: Iterable[Any]) -> list[str]:
    return catalog_file_parse_runtime.normalize_headers(raw)


def _relative_to(path: pathlib.Path, root: pathlib.Path) -> str:
    return catalog_file_parse_runtime.relative_to(path, root)


def _make_safe_filename(filename: str, ext: str, *, fallback: str) -> str:
    return catalog_file_parse_runtime.make_safe_filename(filename, ext, fallback=fallback)


def _stringify(value: Any) -> str:
    return catalog_file_parse_runtime.stringify(value)


def _strip_bom(text: str) -> str:
    return catalog_file_parse_runtime.strip_bom(text)


def _detect_csv_delimiter(text: str) -> str:
    return catalog_file_parse_runtime.detect_csv_delimiter(text)


def _read_csv_bytes(raw: bytes) -> tuple[list[dict[str, str]], dict[str, Any]]:
    return catalog_file_parse_runtime.read_csv_bytes(raw, _catalog_parse_deps())


def _read_excel_bytes(raw: bytes) -> tuple[list[dict[str, str]], dict[str, Any]]:
    return catalog_file_parse_runtime.read_excel_bytes(raw, _catalog_parse_deps())


def _calc_price_coverage(rows: Sequence[Mapping[str, Any]]) -> float:
    return catalog_file_parse_runtime.calc_price_coverage(rows)


def _resolve_job_metrics(
    meta: Mapping[str, Any] | None, rows: Sequence[Mapping[str, Any]]
) -> dict[str, Any]:
    return catalog_file_parse_runtime.resolve_job_metrics(meta, rows)


def _process_pdf(
    *,
    tenant: int,
    saved_path: pathlib.Path,
    tenant_root: pathlib.Path,
    saved_rel_path: pathlib.Path,
    original_name: str,
) -> tuple[list[dict[str, Any]], dict[str, Any], str | None]:
    return catalog_file_parse_runtime.process_pdf(
        tenant=tenant,
        saved_path=saved_path,
        tenant_root=tenant_root,
        saved_rel_path=saved_rel_path,
        original_name=original_name,
        deps=_catalog_parse_deps(),
    )


def _coerce_int(value: Any) -> int | None:
    return _coerce_int_shared(value)


def _find_telegram_user_id(value: Any) -> int | None:
    candidate_keys = (
        "telegram_user_id",
        "telegramUserId",
        "user_id",
        "userId",
        "from_id",
        "fromId",
    )
    if isinstance(value, dict):
        for key in candidate_keys:
            if key in value:
                candidate = _coerce_int(value.get(key))
                if candidate and candidate > 0:
                    return candidate
        for nested in value.values():
            result = _find_telegram_user_id(nested)
            if result is not None:
                return result
    elif isinstance(value, list):
        for entry in value:
            result = _find_telegram_user_id(entry)
            if result is not None:
                return result
    return None


def _find_username(value: Any) -> str | None:
    if isinstance(value, dict):
        if isinstance(value.get("username"), str) and value["username"].strip():
            return value["username"].strip()
        for nested in value.values():
            result = _find_username(nested)
            if result:
                return result
    elif isinstance(value, list):
        for entry in value:
            result = _find_username(entry)
            if result:
                return result
    return None


@router.get("/connect/wa")
def connect_wa(tenant: int, request: Request, k: str | None = None):
    return wa_public_runtime.connect_wa(
        tenant,
        request,
        k=k,
        deps=wa_public_runtime.WaConnectDeps(
            ensure_valid_qr_request_fn=_ensure_valid_qr_request,
            invalid_key_response_fn=_invalid_key_response,
            common_module=common,
            render_template_fn=render_template,
            quote_plus_fn=quote_plus,
            time_module=time,
        ),
    )


@router.get("/connect/tg")
def connect_tg(tenant: int, request: Request, k: str | None = None, key: str | None = None):
    return tg_public_runtime.connect_tg(
        tenant,
        request,
        k=k,
        key=key,
        deps=tg_public_runtime.TgConnectDeps(
            common_module=common,
            settings_module=settings,
            render_template_fn=render_template,
            quote_plus_fn=quote_plus,
        ),
    )


@router.get("/pub/wa/status")
async def wa_status(
    request: Request,
    tenant: int = Query(..., description="Tenant identifier"),
    k: str = Query(..., description="PUBLIC_KEY access token"),
):
    return await wa_public_runtime.wa_status(
        request,
        tenant=tenant,
        key=k,
        deps=wa_public_runtime.WaStatusDeps(
            ensure_valid_qr_request_fn=_ensure_valid_qr_request,
            invalid_key_response_fn=_invalid_key_response,
            as_head_response_fn=_as_head_response,
            common_module=common,
            get_last_qr_id_fn=_get_last_qr_id,
            normalize_qr_id_fn=_normalize_qr_id,
            status_fn=_wa_status_impl,
            baileys_status_fn=_wabaileys_status_impl,
            compose_response_fn=_compose_public_wa_response,
            build_qr_url_fn=_build_public_wa_qr_url,
            no_store_headers_fn=_no_store_headers,
            wa_logger=wa_logger,
        ),
    )


@router.get("/pub/wa/start")
async def wa_start(
    request: Request,
    tenant: int = Query(..., description="Tenant identifier"),
    k: str = Query(..., description="PUBLIC_KEY access token"),
):
    return await wa_public_runtime.wa_start(
        request,
        tenant=tenant,
        key=k,
        deps=wa_public_runtime.WaStartDeps(
            ensure_valid_qr_request_fn=_ensure_valid_qr_request,
            invalid_key_response_fn=_invalid_key_response,
            common_module=common,
            get_last_qr_id_fn=_get_last_qr_id,
            normalize_qr_id_fn=_normalize_qr_id,
            derive_state_fn=_derive_wa_state,
            status_fn=_wa_status_impl,
            baileys_status_fn=_wabaileys_status_impl,
            build_qr_url_fn=_build_public_wa_qr_url,
            no_store_headers_fn=_no_store_headers,
            wa_logger=wa_logger,
        ),
    )


async def _wa_status_impl(tenant: int) -> dict:
    return await wa_public_runtime.legacy_status_impl(tenant, deps=_wa_status_impl_deps())


async def _wabaileys_status_impl(tenant: int) -> dict:
    return await wa_public_runtime.baileys_status_impl(tenant, deps=_wa_status_impl_deps())


def _wa_status_impl_deps() -> wa_public_runtime.WaStatusImplDeps:
    return wa_public_runtime.WaStatusImplDeps(
        common_module=common,
        json_module=json,
        get_last_qr_id_fn=_get_last_qr_id,
        normalize_qr_id_fn=_normalize_qr_id,
        derive_state_fn=_derive_wa_state,
        truthy_flag_fn=_truthy_flag,
    )


def _build_public_wa_qr_url(tenant: int, key: str, qr_id: str | None = None) -> str:
    params: dict[str, Any] = {"tenant": int(tenant), "k": str(key or "")}
    if qr_id:
        params["qr_id"] = str(qr_id)
    return f"/pub/wa/qr.svg?{urlencode(params, doseq=False)}"


def _wa_qr_deps() -> wa_qr_runtime.WaQrDeps:
    return wa_qr_runtime.WaQrDeps(
        common_module=common,
        settings=settings,
        client_config_module=C,
        redis_error_type=redis_ex.RedisError,
        logger=wa_logger,
        no_store_headers_fn=_no_store_headers,
        qr_cache_ttl_fn=_qr_cache_ttl,
    )


def _normalize_qr_id(value: Any) -> str | None:
    return wa_qr_runtime.normalize_qr_id(value)


def _derive_wa_state(data: Mapping[str, Any] | None) -> tuple[str | None, bool]:
    return wa_qr_runtime.derive_wa_state(data)


def _compose_public_wa_response(
    tenant: int,
    key: str | None,
    *,
    status_snapshot: Mapping[str, Any] | None = None,
    qr_id_override: str | None = None,
) -> dict[str, Any]:
    return wa_public_runtime.compose_public_wa_response(
        tenant,
        key,
        status_snapshot=status_snapshot,
        qr_id_override=qr_id_override,
        deps=wa_public_runtime.WaResponseDeps(
            normalize_qr_id_fn=_normalize_qr_id,
            derive_state_fn=_derive_wa_state,
            build_qr_url_fn=_build_public_wa_qr_url,
        ),
    )


def _fetch_qr_bytes(url: str, timeout: float = 6.0):
    return wa_qr_runtime.fetch_qr_bytes(url, _wa_qr_deps(), timeout=timeout)


def _build_qr_candidates(tenant: int, cache_bust: int) -> list[tuple[str, str]]:
    return wa_qr_runtime.build_qr_candidates(tenant, cache_bust, _wa_qr_deps())


def _proxy_qr_with_fallbacks(tenant: int) -> Response:
    return wa_qr_runtime.proxy_qr_with_fallbacks(tenant, _wa_qr_deps())


def _prefetch_qr_session_start(tenant: int) -> None:
    wa_qr_runtime.prefetch_qr_session_start(tenant, _wa_qr_deps())


def _qr_fetch_retry_settings() -> tuple[int, float]:
    return wa_qr_runtime.qr_fetch_retry_settings(_wa_qr_deps())


def _try_fetch_qr_candidate(tenant: int, attempt: int) -> dict[str, Any]:
    return wa_qr_runtime.try_fetch_qr_candidate(tenant, attempt, _wa_qr_deps())


def _qr_fetch_error_response(
    last_status: int,
    last_stage: str,
    last_body_present: bool,
    last_content_type: str,
) -> Response:
    return wa_qr_runtime.qr_fetch_error_response(
        last_status,
        last_stage,
        last_body_present,
        last_content_type,
        _wa_qr_deps(),
    )


def _ensure_valid_qr_request(
    raw_tenant: int | str | None,
    raw_key: str | None,
    request: Request | None = None,
    *,
    query_param_only: bool = False,
) -> tuple[int, str] | None:
    return public_auth_runtime.ensure_valid_public_access(
        raw_tenant,
        raw_key,
        request,
        query_param_only=query_param_only,
        deps=public_auth_runtime.PublicAccessDeps(
            coerce_tenant_fn=_coerce_tenant,
            admin_token_valid_fn=_admin_token_valid,
            list_keys_fn=common.list_keys,
            get_tenant_pubkey_fn=common.get_tenant_pubkey,
            resolve_public_key_candidate_fn=_resolve_public_key_candidate,
            expected_public_key_value_fn=_expected_public_key_value,
            valid_key_fn=common.valid_key,
        ),
    )


def _get_last_qr_id(tenant: int) -> tuple[str | None, bool]:
    return wa_qr_runtime.get_last_qr_id(tenant, _wa_qr_deps())


def _load_cached_qr_entry(tenant: int, qr_id: str) -> tuple[dict[str, Any] | None, bool]:
    return wa_qr_runtime.load_cached_qr_entry(tenant, qr_id, _wa_qr_deps())


def _resolve_cached_qr(tenant: int) -> tuple[str | None, dict[str, Any] | None, bool]:
    return wa_qr_runtime.resolve_cached_qr(tenant, _wa_qr_deps())


def _load_cached_svg(tenant: int, qr_id: str) -> tuple[str | None, bool]:
    return wa_qr_runtime.load_cached_svg(tenant, qr_id, _wa_qr_deps())


def _qr_expired_response(qr_id: str | None = None) -> JSONResponse:
    return wa_qr_runtime.qr_expired_response(_no_store_headers, qr_id)


def _as_head_response(response: Response, request: Request) -> Response:
    return wa_qr_runtime.as_head_response(response, request)


def _render_qr_svg_from_text(qr_text: str) -> str | None:
    return wa_qr_runtime.render_qr_svg_from_text(qr_text)


def _render_qr_png_bytes(qr_text: str) -> bytes | None:
    return wa_qr_runtime.render_qr_png_bytes(qr_text)


def _cache_qr_payload(
    tenant: int,
    qr_id: str,
    entry: Mapping[str, Any],
    *,
    include_last: bool = True,
) -> None:
    wa_qr_runtime.cache_qr_payload(
        tenant,
        qr_id,
        entry,
        _wa_qr_deps(),
        include_last=include_last,
    )


def _normalize_qr_cache_values(
    tenant: int,
    qr_id: str,
    entry: Mapping[str, Any],
) -> tuple[dict[str, Any], str | None, str | None, str | None]:
    return wa_qr_runtime.normalize_qr_cache_values(tenant, qr_id, entry)


def _write_qr_cache_values(
    tenant: int,
    qr_id: str,
    *,
    json_payload: str | None,
    svg_value: str | None,
    png_value: str | None,
    txt_value: str | None,
    include_last: bool,
) -> None:
    wa_qr_runtime.write_qr_cache_values(
        tenant,
        qr_id,
        json_payload=json_payload,
        svg_value=svg_value,
        png_value=png_value,
        txt_value=txt_value,
        include_last=include_last,
        deps=_wa_qr_deps(),
    )


def _persist_qr_entry(tenant: int, qr_id: str, entry: Mapping[str, Any]) -> None:
    wa_qr_runtime.persist_qr_entry(tenant, qr_id, entry, _wa_qr_deps())


async def _resolve_tenant_and_key(
    request: Request | None,
    raw_tenant: int | str | None,
    raw_key: str | None,
    *,
    query_keys: tuple[str, ...] = ("key", "k"),
    allow_body: bool = True,
) -> tuple[int | str | None, str | None]:
    return await public_request_runtime.resolve_tenant_and_key(
        request,
        raw_tenant,
        raw_key,
        query_keys=query_keys,
        allow_body=allow_body,
        json_module=json,
    )


def require_client_key(
    raw_tenant: int | str | None,
    raw_key: str | None,
) -> tuple[int, str] | Response:
    try:
        tenant_id = _coerce_tenant(raw_tenant)
    except ValueError:
        return JSONResponse({"error": "invalid_key"}, status_code=401, headers=_no_store_headers())

    key = "" if raw_key is None else str(raw_key).strip()
    if not key or not common.valid_key(tenant_id, key):
        return JSONResponse({"error": "invalid_key"}, status_code=401, headers=_no_store_headers())

    return tenant_id, key


def _normalize_public_token(value: str | None) -> str:
    return public_auth_runtime.normalize_public_token(value)


def _expected_public_key_value() -> str:
    return public_auth_runtime.expected_public_key_value(settings)


def _resolve_public_key_candidate(
    key_candidate: str | None,
    request: Request | None = None,
    *,
    query_param_only: bool = False,
) -> str:
    return public_auth_runtime.resolve_public_key_candidate(
        key_candidate,
        request,
        query_param_only=query_param_only,
    )


def _ensure_public_key(
    key_candidate: str | None,
    request: Request | None = None,
    *,
    query_param_only: bool = False,
) -> str | None:
    return public_auth_runtime.ensure_public_key(
        key_candidate,
        request,
        query_param_only=query_param_only,
        expected_key_fn=_expected_public_key_value,
    )


def _invalid_key_response() -> JSONResponse:
    return JSONResponse(
        {"error": "invalid_key"},
        status_code=401,
        headers=_no_store_headers(),
    )


def _resolve_qr_identifier(primary: str | None, legacy: str | None = None) -> str:
    candidate = primary if primary is not None else legacy
    if candidate is None:
        return ""
    return str(candidate).strip()


def _admin_token_valid(request: Request) -> bool:
    token = request.headers.get("X-Admin-Token")
    return bool(token) and token == settings.ADMIN_TOKEN


def _has_public_tg_access(
    request: Request,
    key_candidate: str | None,
    *,
    allow_admin: bool = True,
    query_param_only: bool = False,
) -> tuple[bool, str | None]:
    if allow_admin and _admin_token_valid(request):
        return True, "admin"

    resolved = _ensure_public_key(
        key_candidate,
        request,
        query_param_only=query_param_only,
    )
    return (resolved is not None, resolved)


def _has_tg_access_for_tenant(
    tenant_id: int,
    request: Request,
    key_candidate: str | None,
    *,
    allow_admin: bool = True,
    query_param_only: bool = False,
) -> tuple[bool, str | None]:
    key = (key_candidate or "").strip()
    if key and common.valid_key(tenant_id, key):
        return True, key
    return _has_public_tg_access(
        request,
        key_candidate,
        allow_admin=allow_admin,
        query_param_only=query_param_only,
    )


def _invalid_tenant_response(
    route: str,
    tenant_candidate: int | str | None,
    *,
    force: bool | None = None,
) -> JSONResponse:
    _log_tg_proxy(route, tenant_candidate, 400, None, error="invalid_tenant", force=force)
    return JSONResponse({"error": "invalid_tenant"}, status_code=400, headers=_no_store_headers())


def _unauthorized_response(
    route: str,
    tenant_id: int | str | None,
    *,
    force: bool | None = None,
) -> JSONResponse:
    _log_tg_proxy(route, tenant_id, 401, None, error="unauthorized", force=force)
    return _invalid_key_response()


def _tg_unavailable_response(
    route: str,
    tenant_id: int | str | None,
    detail: str | Exception | None,
    *,
    force: bool | None = None,
) -> JSONResponse:
    detail_text = _stringify_detail(str(detail)) if detail not in (None, "") else "tg_unavailable"
    if not detail_text:
        detail_text = "tg_unavailable"
    _log_tg_proxy(route, tenant_id, 0, None, error=detail_text, force=force)
    headers = _no_store_headers({"X-Telegram-Upstream-Status": "-"})
    body: dict[str, Any] = {"error": "tg_unavailable"}
    if detail_text and detail_text != "tg_unavailable":
        body["detail"] = detail_text
    return JSONResponse(body, status_code=502, headers=headers)


def _passthrough_upstream_response(
    route: str,
    tenant_id: int | str | None,
    upstream: Any,
    *,
    success_content_type: str | None = "application/json",
    error_content_type: str | None = "application/json",
    include_no_store: bool = True,
    force: bool | None = None,
) -> Response:
    return tg_proxy_runtime.passthrough_upstream_response(
        route,
        tenant_id,
        upstream,
        no_store_headers_fn=_no_store_headers,
        log_tg_proxy_fn=_log_tg_proxy,
        success_content_type=success_content_type,
        error_content_type=error_content_type,
        include_no_store=include_no_store,
        force=force,
    )


def _proxy_headers(headers: Mapping[str, str] | None, status_code: int) -> dict[str, str]:
    return tg_proxy_runtime.proxy_headers(
        headers,
        status_code,
        no_store_value=NO_STORE_CACHE_VALUE,
    )


def _truthy_flag(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        normalized = value.strip().lower()
        return normalized in {"1", "true", "yes", "on"}
    return False


def _coerce_body_bytes(body: Any) -> bytes:
    if isinstance(body, (bytes, bytearray)):
        return bytes(body)
    if isinstance(body, str):
        return body.encode("utf-8")
    if body is None:
        return b""
    try:
        return json.dumps(body, ensure_ascii=False).encode("utf-8")
    except Exception:
        return b""


@router.api_route("/pub/wa/qr.svg", methods=["GET", "HEAD"])
async def wa_qr_svg(
    request: Request,
    tenant: int = Query(..., description="Tenant identifier"),
    k: str = Query(..., description="PUBLIC_KEY access token"),
    qr_id: str | None = Query(None, description="Explicit QR identifier from status"),
):
    return await wa_public_runtime.wa_qr_svg(
        request,
        tenant=tenant,
        key=k,
        qr_id=qr_id,
        deps=wa_public_runtime.WaPublicDeps(
            ensure_valid_qr_request_fn=_ensure_valid_qr_request,
            invalid_key_response_fn=_invalid_key_response,
            as_head_response_fn=_as_head_response,
            common_module=common,
            proxy_baileys_qr_fn=_proxy_baileys_qr,
            normalize_qr_id_fn=_normalize_qr_id,
            get_last_qr_id_fn=_get_last_qr_id,
            no_store_headers_fn=_no_store_headers,
            load_cached_svg_fn=_load_cached_svg,
            httpx_module=httpx,
            wa_logger=wa_logger,
            qr_expired_response_fn=_qr_expired_response,
            cache_qr_payload_fn=_cache_qr_payload,
        ),
    )


@router.get("/pub/tg/slots")
async def tg_slots_get(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
):
    route = "/pub/tg/slots"
    tenant_candidate, key_candidate = await _resolve_tenant_and_key(
        request,
        tenant,
        k,
        query_keys=("k",),
        allow_body=False,
    )
    auth = await _authorize_public_settings_request(request, tenant_candidate, key_candidate)
    if isinstance(auth, Response):
        return auth
    tenant_id, validated_key = auth
    _log_public_tg_request(route, tenant_id, validated_key or "session")
    cfg = common.read_tenant_config(int(tenant_id)) or {}
    return JSONResponse({"ok": True, **_tg_slots_config(cfg)}, headers=_no_store_headers())


@router.post("/pub/tg/slots")
async def tg_slots_save(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
):
    route = "/pub/tg/slots"
    tenant_candidate, key_candidate = await _resolve_tenant_and_key(
        request,
        tenant,
        k,
        query_keys=("k",),
        allow_body=True,
    )
    auth = await _authorize_public_settings_request(request, tenant_candidate, key_candidate)
    if isinstance(auth, Response):
        return auth
    tenant_id, validated_key = auth
    _log_public_tg_request(route, tenant_id, validated_key or "session")
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    cfg = common.read_tenant_config(int(tenant_id)) or {}
    tg_slots_runtime.apply_tg_slots_payload(cfg, payload, _tg_slots_deps())
    common.write_tenant_config(int(tenant_id), cfg)
    return JSONResponse({"ok": True, **_tg_slots_config(cfg)}, headers=_no_store_headers())


def _tg_public_deps() -> tg_public_runtime.TgPublicDeps:
    return tg_public_runtime.TgPublicDeps(
        log_deprecated_fn=_log_deprecated,
        resolve_tenant_and_key_fn=_resolve_tenant_and_key,
        authorize_public_settings_request_fn=_authorize_public_settings_request,
        tg_slot_tenant_fn=_tg_slot_tenant,
        log_public_tg_request_fn=_log_public_tg_request,
        client_identifier_fn=_client_identifier,
        log_tg_proxy_fn=_log_tg_proxy,
        no_store_headers_fn=_no_store_headers,
        register_password_attempt_fn=_register_password_attempt,
        tg_call_fn=_tg_call,
        tg_worker_call_error_type=TgWorkerCallError,
        resolve_qr_identifier_fn=_resolve_qr_identifier,
        quote_fn=quote,
        common_module=common,
        resolve_tg_base_fn=_resolve_tg_base,
        extract_json_detail_fn=_extract_json_detail,
        stringify_detail_fn=_stringify_detail,
        proxy_headers_fn=_proxy_headers,
        json_module=json,
    )


@router.api_route("/pub/tg/start", methods=["GET", "POST"])
async def tg_start(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    slot: int = Query(1, ge=1, le=TG_SLOT_MAX),
):
    return await tg_public_runtime.start(
        "/pub/tg/start",
        request,
        tenant,
        k,
        allow_body=request.method.upper() == "POST",
        slot=slot,
        deps=_tg_public_deps(),
    )


async def _handle_tg_twofa(
    route: str,
    request: Request,
    tenant: int | str | None,
    key: str | None,
    slot: int = 1,
) -> Response:
    return await tg_public_runtime.handle_twofa(
        route,
        request,
        tenant,
        key,
        slot=slot,
        deps=_tg_public_deps(),
    )


@router.post("/pub/tg/2fa")
async def tg_twofa(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    key: str | None = None,
    slot: int = Query(1, ge=1, le=TG_SLOT_MAX),
):
    return await _handle_tg_twofa("/pub/tg/2fa", request, tenant, k or key, slot=slot)


@router.post("/pub/tg/twofa.submit")
async def tg_twofa_submit(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    key: str | None = None,
    slot: int = Query(1, ge=1, le=TG_SLOT_MAX),
):
    return await _handle_tg_twofa("/pub/tg/twofa.submit", request, tenant, k or key, slot=slot)


@router.post("/pub/tg/password")
async def tg_password(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    key: str | None = None,
    slot: int = Query(1, ge=1, le=TG_SLOT_MAX),
):
    return await _handle_tg_twofa("/pub/tg/password", request, tenant, k or key, slot=slot)


@router.post("/pub/tg/restart")
async def tg_restart(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    key: str | None = None,
    slot: int = Query(1, ge=1, le=TG_SLOT_MAX),
):
    route = "/pub/tg/restart"
    _log_deprecated(route)
    tenant_candidate, key_candidate = await _resolve_tenant_and_key(request, tenant, k or key)
    auth = await _authorize_public_settings_request(request, tenant_candidate, key_candidate)
    if isinstance(auth, Response):
        return auth
    tenant_id, _ = auth
    tg_tenant_id = _tg_slot_tenant(tenant_id, slot)

    try:
        upstream = await common.tg_post(
            "/session/restart",
            {"tenant_id": tg_tenant_id},
            timeout=5.0,
        )
    except httpx.HTTPError as exc:
        return _tg_unavailable_response(route, tenant_id, exc, force=True)
    except Exception as exc:
        return _tg_unavailable_response(route, tenant_id, exc, force=True)

    return _passthrough_upstream_response(route, tenant_id, upstream, force=True)


@router.get("/pub/tg/status")
async def tg_status(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    slot: int = Query(1, ge=1, le=TG_SLOT_MAX),
):
    return await tg_public_runtime.status(
        "/pub/tg/status",
        request,
        tenant,
        k,
        slot=slot,
        deps=_tg_public_deps(),
    )


@router.get("/pub/tg/qr.png")
async def tg_qr_png(
    request: Request,
    tenant: int | str | None = None,
    qr_id: str | None = None,
    k: str | None = None,
    slot: int = Query(1, ge=1, le=TG_SLOT_MAX),
):
    return await tg_public_runtime.qr_png(
        "/pub/tg/qr.png",
        request,
        tenant,
        qr_id,
        k,
        slot=slot,
        deps=_tg_public_deps(),
    )


@router.get("/pub/tg/media/{peer_id}/{message_id}")
async def tg_media(
    request: Request,
    peer_id: str,
    message_id: str,
    tenant: int | str | None = None,
    k: str | None = None,
):
    return await tg_public_runtime.proxy_tg_resource(
        "/pub/tg/media",
        request,
        tenant,
        k,
        resource_path_fn=lambda tenant_id: f"/media/{tenant_id}/{int(str(peer_id))}/{int(str(message_id))}",
        deps=_tg_public_deps(),
        timeout=15.0,
    )


@router.get("/pub/tg/avatar/{peer_id}")
async def tg_avatar(
    request: Request,
    peer_id: str,
    tenant: int | str | None = None,
    k: str | None = None,
):
    return await tg_public_runtime.proxy_tg_resource(
        "/pub/tg/avatar",
        request,
        tenant,
        k,
        resource_path_fn=lambda tenant_id: f"/avatar/{tenant_id}/{int(str(peer_id))}",
        deps=_tg_public_deps(),
        timeout=15.0,
    )


@router.get("/pub/chat/avatar/{lead_id}")
async def chat_avatar(
    request: Request,
    lead_id: str,
    tenant: int | str | None = None,
    k: str | None = None,
):
    route = "/pub/chat/avatar"
    tenant_candidate, key_candidate = await _resolve_tenant_and_key(
        request,
        tenant,
        k,
        query_keys=("k",),
        allow_body=False,
    )
    auth = await _authorize_public_settings_request(request, tenant_candidate, key_candidate)
    if isinstance(auth, Response):
        return auth
    tenant_id, validated_key = auth
    _log_public_tg_request(route, tenant_id, validated_key or "session")
    try:
        lead_ref = int(str(lead_id))
    except Exception:
        return JSONResponse({"error": "bad_params"}, status_code=400, headers=_no_store_headers())
    return await public_avatar_runtime.chat_avatar_response(
        tenant_id=int(tenant_id),
        lead_id=lead_ref,
        deps=public_avatar_runtime.PublicAvatarDeps(
            get_lead_dialog_metadata_fn=get_lead_dialog_metadata,
            resolve_avito_profile_fn=avito.resolve_chat_participant_profile,
            no_store_headers_fn=_no_store_headers,
            http_client_cls=httpx.AsyncClient,
        ),
    )


@router.get("/pub/tg/qr.txt")
async def tg_qr_txt(
    request: Request,
    tenant: int | str | None = None,
    qr_id: str | None = None,
    k: str | None = None,
    key: str | None = None,
    slot: int = Query(1, ge=1, le=TG_SLOT_MAX),
):
    return await tg_public_runtime.qr_txt(
        "/pub/tg/qr.txt",
        request,
        tenant,
        qr_id,
        k or key,
        slot=slot,
        deps=_tg_public_deps(),
    )


@router.api_route("/pub/tg/logout", methods=["GET", "POST"])
async def tg_logout(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    key: str | None = None,
    slot: int = Query(1, ge=1, le=TG_SLOT_MAX),
):
    route = "/pub/tg/logout"
    tenant_candidate, key_candidate = await _resolve_tenant_and_key(request, tenant, k or key)
    auth = await _authorize_public_settings_request(request, tenant_candidate, key_candidate)
    if isinstance(auth, Response):
        return auth
    tenant_id, _ = auth
    tg_tenant_id = _tg_slot_tenant(tenant_id, slot)

    try:
        upstream = await common.tg_post(
            "/session/logout",
            {"tenant_id": tg_tenant_id, "force": True},
            timeout=5.0,
        )
    except httpx.HTTPError as exc:
        return _tg_unavailable_response(route, tenant_id, exc)
    except Exception as exc:
        return _tg_unavailable_response(route, tenant_id, exc)

    return _passthrough_upstream_response(route, tenant_id, upstream)


@router.get("/pub/wa/qr.png")
def wa_qr_png(
    request: Request,
    tenant: int = Query(..., description="Tenant identifier"),
    k: str = Query(..., description="PUBLIC_KEY access token"),
    qr_id: str | None = Query(None, description="Explicit QR identifier from status"),
):
    return wa_qr_runtime.wa_qr_png_response(
        request=request,
        tenant=tenant,
        key=k,
        qr_id=qr_id,
        ensure_valid_qr_request_fn=_ensure_valid_qr_request,
        invalid_key_response_fn=_invalid_key_response,
        deps=_wa_qr_deps(),
    )


def _proxy_baileys_qr(tenant: int) -> Response:
    return wa_qr_runtime.proxy_baileys_qr(tenant, _wa_qr_deps())


@router.get("/pub/wa/restart")
async def wa_restart(
    request: Request,
    tenant: int = Query(..., description="Tenant identifier"),
    k: str = Query(..., description="PUBLIC_KEY access token"),
):
    """Force-restart waweb session to issue a fresh QR.

    Security: requires a valid public access key `k` for the tenant.
    """
    return await wa_public_runtime.wa_restart(
        request,
        tenant=tenant,
        key=k,
        deps=wa_public_runtime.WaRestartDeps(
            ensure_valid_qr_request_fn=_ensure_valid_qr_request,
            invalid_key_response_fn=_invalid_key_response,
            common_module=common,
            json_module=json,
            wa_logger=wa_logger,
        ),
    )


def _resolve_public_settings_key(request: Request, key_candidate: str | None) -> str:
    return public_auth_runtime.resolve_public_settings_key(request, key_candidate)


async def _authorize_public_settings_request(
    request: Request,
    tenant: int | str | None,
    key_candidate: str | None,
) -> tuple[int, str] | Response:
    return await public_auth_runtime.authorize_public_settings_request(
        request,
        tenant,
        key_candidate,
        public_auth_runtime.PublicAuthDeps(
            get_current_user_fn=auth_utils.get_current_user,
            coerce_tenant_fn=_coerce_tenant,
            resolve_public_settings_key_fn=_resolve_public_settings_key,
            get_tenant_pubkey_fn=getattr(C, "get_tenant_pubkey", common.get_tenant_pubkey),
            list_keys_fn=getattr(C, "list_keys", common.list_keys),
            magic_link_enabled_fn=auth_utils.magic_link_enabled,
            valid_key_fn=common.valid_key,
            settings=settings,
        ),
    )


def _avito_oauth_deps() -> avito_oauth_runtime.AvitoOAuthDeps:
    return avito_oauth_runtime.AvitoOAuthDeps(
        authorize_public_settings_request_fn=_authorize_public_settings_request,
        coerce_int_fn=_coerce_int,
        avito_module=avito,
        logger=logger,
        common_module=common,
        json_module=json,
        redis_error_type=redis_ex.RedisError,
        avito_state_ttl=AVITO_STATE_TTL,
        avito_state_cookie=AVITO_STATE_COOKIE,
        avito_state_key_fn=_avito_state_key,
        build_avito_oauth_state_fn=_build_avito_oauth_state,
        avito_oauth_redirect_entry_url_fn=_avito_oauth_redirect_entry_url,
        set_avito_state_cookie_fn=_set_avito_state_cookie,
        avito_callback_html_fn=_avito_callback_html,
        clear_avito_state_cookie_fn=_clear_avito_state_cookie,
        verify_avito_oauth_state_fn=_verify_avito_oauth_state,
        resolve_tenant_from_state_fn=resolve_tenant_from_state,
        build_token_update_payload_fn=build_token_update_payload,
        avito_token_payload_error=AvitoTokenPayloadError,
    )


avito_oauth_routes.register_routes(oauth_router, _avito_oauth_deps)


async def avito_oauth_authorize(
    request: Request,
    tenant: int | None = None,
    k: str | None = None,
    redirect: bool = False,
):
    return await avito_oauth_runtime.oauth_authorize(
        request, tenant=tenant, key=k, redirect=redirect, deps=_avito_oauth_deps()
    )


async def avito_oauth_callback(
    request: Request,
    code: str | None = None,
    state: str | None = None,
    error: str | None = None,
):
    return await avito_oauth_runtime.oauth_callback(
        request, code=code, state=state, error=error, deps=_avito_oauth_deps()
    )


def _max_webhook_url(request: Request, tenant_id: int, secret: str) -> str:
    return max_public_runtime.max_webhook_url(
        request,
        tenant_id,
        secret,
        public_url_fn=common.public_url,
    )


def _max_personal_callback_url(request: Request, tenant_id: int, secret: str) -> str:
    return max_public_runtime.max_personal_callback_url(
        request,
        tenant_id,
        secret,
        public_url_fn=common.public_url,
    )


def _max_public_deps() -> max_public_runtime.MaxPublicDeps:
    return max_public_runtime.MaxPublicDeps(
        authorize_fn=_authorize_public_settings_request,
        max_integration=max_integration,
        logger=logger,
        public_url_fn=common.public_url,
        secrets_module=secrets,
        time_module=time,
    )


def _max_personal_deps() -> max_public_runtime.MaxPersonalDeps:
    return max_public_runtime.MaxPersonalDeps(
        authorize_fn=_authorize_public_settings_request,
        service=max_personal_service,
        transport=max_personal_transport,
        refresh_status_fn=_max_personal_refresh_status,
        callback_url_fn=_max_personal_callback_url,
    )


@max_router.get("/status")
async def max_status(request: Request, tenant: int | None = None, k: str | None = None):
    return await max_public_runtime.max_status(request, tenant, k, _max_public_deps())


@max_router.post("/connect")
async def max_connect(request: Request, tenant: int | None = None, k: str | None = None):
    return await max_public_runtime.max_connect(request, tenant, k, _max_public_deps())


@max_router.post("/disconnect")
async def max_disconnect(request: Request, tenant: int | None = None, k: str | None = None):
    return await max_public_runtime.max_disconnect(request, tenant, k, _max_public_deps())


async def _max_personal_refresh_status(tenant_id: int) -> dict[str, Any] | None:
    return await max_public_runtime.refresh_max_personal_status(
        tenant_id,
        max_public_runtime.MaxPersonalDeps(
            authorize_fn=_authorize_public_settings_request,
            service=max_personal_service,
            transport=max_personal_transport,
            refresh_status_fn=_max_personal_refresh_status,
            callback_url_fn=_max_personal_callback_url,
        ),
    )


@max_personal_router.get("/status")
async def max_personal_status(request: Request, tenant: int | None = None, k: str | None = None):
    return await max_public_runtime.max_personal_status(request, tenant, k, _max_personal_deps())


@max_personal_router.post("/connect")
async def max_personal_connect(request: Request, tenant: int | None = None, k: str | None = None):
    return await max_public_runtime.max_personal_connect(request, tenant, k, _max_personal_deps())


@max_personal_router.post("/session/start")
async def max_personal_session_start(
    request: Request, tenant: int | None = None, k: str | None = None
):
    return await max_personal_connect(request, tenant=tenant, k=k)


@max_personal_router.get("/session/qr")
async def max_personal_session_qr(request: Request, tenant: int | None = None, k: str | None = None):
    return await max_public_runtime.max_personal_session_qr(request, tenant, k, _max_personal_deps())


@max_personal_router.post("/session/logout")
async def max_personal_session_logout(
    request: Request, tenant: int | None = None, k: str | None = None
):
    return await max_public_runtime.max_personal_session_logout(
        request,
        tenant,
        k,
        _max_personal_deps(),
    )


@max_personal_router.post("/disconnect")
async def max_personal_disconnect(
    request: Request, tenant: int | None = None, k: str | None = None
):
    return await max_public_runtime.max_personal_disconnect(request, tenant, k, _max_personal_deps())


@max_personal_router.post("/send")
async def max_personal_send(request: Request, tenant: int | None = None, k: str | None = None):
    return await max_public_runtime.max_personal_send(request, tenant, k, _max_personal_deps())


@router.get("/pub/integrations/amocrm/oauth/start")
async def amocrm_oauth_start(
    request: Request,
    tenant_id: int | None = None,
    tenant: int | None = None,
    k: str | None = None,
):
    return await amocrm_public_runtime.oauth_start(
        request,
        tenant_id=tenant_id,
        tenant=tenant,
        key=k,
        deps=_amocrm_public_runtime_deps(),
    )


@router.get("/pub/integrations/amocrm/oauth/callback", name="amocrm_oauth_callback")
async def amocrm_oauth_callback(
    request: Request,
    code: str | None = None,
    state: str | None = None,
):
    return await amocrm_public_runtime.oauth_callback(
        request,
        code=code,
        state=state,
        deps=_amocrm_public_runtime_deps(),
    )


@router.get("/pub/integrations/amocrm/status")
async def amocrm_status(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
):
    return await amocrm_public_runtime.oauth_status(
        request,
        tenant=tenant,
        key=k,
        deps=_amocrm_public_runtime_deps(),
    )


async def _amocrm_chat_webhook_impl(
    request: Request,
    *,
    token: str | None = None,
    scope_id: str | None = None,
):
    return await amocrm_public_runtime.chat_webhook(
        request,
        token=token,
        scope_id=scope_id,
        deps=_amocrm_public_runtime_deps(),
    )


@router.post("/pub/integrations/amocrm/chat/webhook")
async def amocrm_chat_webhook(
    request: Request,
    token: str | None = None,
):
    return await _amocrm_chat_webhook_impl(request, token=token)


@router.post("/pub/integrations/amocrm/chat/webhook/{scope_id}")
async def amocrm_chat_webhook_scoped(
    request: Request,
    scope_id: str,
    token: str | None = None,
):
    return await _amocrm_chat_webhook_impl(request, token=token, scope_id=scope_id)


@router.get("/pub/integrations/amocrm/chat/avatar/{tenant_id}/{peer_id}/{token}")
async def amocrm_chat_avatar_proxy(
    request: Request,
    tenant_id: int,
    peer_id: str,
    token: str,
):
    return await amocrm_avatar_runtime.chat_avatar_proxy(
        request,
        tenant_id,
        peer_id,
        token,
        deps=_amocrm_avatar_deps(),
    )


@router.get("/pub/integrations/amocrm/chat/lead-avatar/{tenant_id}/{lead_id}/{token}")
async def amocrm_chat_lead_avatar_proxy(
    request: Request,
    tenant_id: int,
    lead_id: int,
    token: str,
):
    return await amocrm_avatar_runtime.lead_avatar_proxy(
        request,
        tenant_id,
        lead_id,
        token,
        deps=_amocrm_avatar_deps(),
    )


@router.get("/pub/integrations/amocrm/pipeline")
async def amocrm_pipeline(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    apply: int | None = None,
    pipeline_id: int | None = None,
):
    return await amocrm_public_runtime.pipeline(
        request,
        tenant=tenant,
        key=k,
        apply=apply,
        pipeline_id=pipeline_id,
        deps=_amocrm_public_runtime_deps(),
    )


@router.post("/pub/integrations/amocrm/test")
async def amocrm_test(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
):
    return await amocrm_public_runtime.test_connection(
        request,
        tenant=tenant,
        key=k,
        deps=_amocrm_public_runtime_deps(),
    )


@router.post("/pub/integrations/amocrm/disconnect")
@router.post("/pub/integrations/amocrm/uninstall")
async def amocrm_disconnect(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
):
    return await amocrm_public_runtime.disconnect(
        request,
        tenant=tenant,
        key=k,
        deps=_amocrm_public_runtime_deps(),
    )


@router.get("/pub/settings/get")
async def settings_get(request: Request, tenant: int | str | None = None, k: str | None = None):
    return await settings_public_runtime.settings_get(
        request,
        tenant=tenant,
        key=k,
        deps=_public_settings_runtime_deps(),
    )


@router.post("/pub/settings/save")
async def settings_save(request: Request, tenant: int | str | None = None, k: str | None = None):
    return await settings_public_runtime.settings_save(
        request,
        tenant=tenant,
        key=k,
        deps=_public_settings_runtime_deps(),
    )


def _photo_root(tenant_id: int) -> pathlib.Path:
    core_module.ensure_tenant_files(tenant_id)
    root = core_module.tenant_dir(tenant_id) / "uploads" / "photos"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _photo_manifest_path(tenant_id: int) -> pathlib.Path:
    return _photo_root(tenant_id) / "manifest.json"


def _read_photo_manifest(tenant_id: int) -> list[dict[str, Any]]:
    path = _photo_manifest_path(tenant_id)
    if not path.exists() or not path.is_file():
        return []
    try:
        with open(path, "r", encoding="utf-8") as fh:
            raw = json.load(fh)
    except Exception:
        return []
    if isinstance(raw, list):
        return [entry for entry in raw if isinstance(entry, dict)]
    return []


def _write_photo_manifest(tenant_id: int, entries: list[dict[str, Any]]) -> None:
    path = _photo_manifest_path(tenant_id)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(entries, fh, ensure_ascii=False, indent=2)


def _photo_public_url(request: Request, tenant_id: int, key: str, photo_id: str) -> str:
    base = common.public_url(request, f"/pub/files/photos/{photo_id}")
    if not base:
        return ""
    joiner = "&" if "?" in base else "?"
    return f"{base}{joiner}tenant={tenant_id}&k={quote_plus(key)}"


def _validate_photo_upload(filename: str, content_type: str | None) -> tuple[bool, str]:
    if not filename:
        return False, "empty_file"
    ext = pathlib.Path(filename).suffix.lower()
    if ext not in PHOTO_ALLOWED_EXTS:
        return False, "unsupported_extension"
    if content_type:
        mime = content_type.strip().lower()
        if mime and mime not in PHOTO_ALLOWED_MIMES and mime != "application/octet-stream":
            return False, "unsupported_mime"
    return True, ""


# Move public catalog helpers closer to settings endpoints for discoverability.
@router.get("/pub/catalog/csv")
async def public_catalog_csv_get(
    request: Request, tenant: int | str | None = None, k: str | None = None
):
    auth = await _authorize_public_settings_request(request, tenant, k)
    if isinstance(auth, Response):
        return auth

    tenant_id, _ = auth

    cfg = common.read_tenant_config(tenant_id)
    try:
        table = read_csv_table(tenant_id, cfg)
    except FileNotFoundError:
        return JSONResponse({"detail": "csv_not_ready"}, status_code=404)

    return {"ok": True, **table}


@router.post("/pub/catalog/csv")
async def public_catalog_csv_save(
    request: Request, tenant: int | str | None = None, k: str | None = None
):
    auth = await _authorize_public_settings_request(request, tenant, k)
    if isinstance(auth, Response):
        return auth

    tenant_id, _ = auth

    try:
        payload = await request.json()
    except Exception:
        payload = {}

    columns = payload.get("columns") if isinstance(payload, dict) else None
    rows = payload.get("rows") if isinstance(payload, dict) else None

    cfg = common.read_tenant_config(tenant_id)
    try:
        written = write_csv_table(tenant_id, columns, rows, cfg)
    except FileNotFoundError:
        return JSONResponse({"detail": "csv_not_ready"}, status_code=404)
    except ValueError as exc:
        detail = str(exc) or "invalid_rows"
        return JSONResponse({"detail": detail}, status_code=400)

    return {"ok": True, "rows": written}


# Move public catalog upload off the client namespace to avoid route collisions
# with the client router. The tenant is accepted as a query parameter.


def _catalog_public_deps() -> catalog_public_runtime.CatalogPublicDeps:
    from . import client as client_module

    return catalog_public_runtime.CatalogPublicDeps(
        logger=logger,
        resolve_key_fn=client_module._resolve_key,
        auth_fn=client_module._auth,
        common_module=common,
        allowed_extensions=ALLOWED_EXTENSIONS,
        max_upload_size_bytes=MAX_UPLOAD_SIZE_BYTES,
        make_safe_filename_fn=_make_safe_filename,
        relative_to_fn=_relative_to,
        read_csv_bytes_fn=_read_csv_bytes,
        read_excel_bytes_fn=_read_excel_bytes,
        process_pdf_fn=_process_pdf,
        resolve_job_metrics_fn=_resolve_job_metrics,
        catalog_index_error=CatalogIndexError,
        write_catalog_csv_fn=write_catalog_csv,
        stringify_fn=_stringify,
        amocrm_service_module=amocrm_service,
        write_tenant_config_fn=common.write_tenant_config,
        read_tenant_config_fn=common.read_tenant_config,
        quote_plus_fn=quote_plus,
    )


def _public_photos_deps() -> public_photos_runtime.PublicPhotosDeps:
    return public_photos_runtime.PublicPhotosDeps(
        authorize_fn=_authorize_public_settings_request,
        read_manifest_fn=_read_photo_manifest,
        write_manifest_fn=_write_photo_manifest,
        photo_url_fn=_photo_public_url,
        validate_upload_fn=_validate_photo_upload,
        photo_root_fn=_photo_root,
        tenant_dir_fn=core_module.tenant_dir,
        max_bytes=PHOTO_MAX_BYTES,
        logger=logger,
        sync_asset_fn=client_assets_runtime.sync_public_photo_asset,
        compile_asset_fn=client_assets_runtime.compile_public_photo_asset_rule,
    )


def _catalog_view_deps() -> catalog_public_runtime.CatalogViewDeps:
    return catalog_public_runtime.CatalogViewDeps(
        core_module=core_module,
        render_template_fn=render_template,
        template_name=CATALOG_VIEW_TEMPLATE,
        time_module=time,
    )


@router.post("/pub/catalog/upload")
async def catalog_upload(
    request: Request,
    background_tasks: BackgroundTasks,
    tenant: str | None = Query(None),
):
    return await catalog_public_runtime.catalog_upload(
        request,
        background_tasks,
        tenant=tenant,
        deps=_catalog_public_deps(),
    )


@router.get("/pub/files/photos/list")
async def photos_list(request: Request, tenant: int | str | None = None, k: str | None = None):
    return await public_photos_runtime.photos_list(
        request,
        tenant=tenant,
        key=k,
        deps=_public_photos_deps(),
    )


@router.post("/pub/files/photos/upload")
async def photos_upload(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    file: UploadFile = File(...),
):
    return await public_photos_runtime.photos_upload(
        request,
        tenant=tenant,
        key=k,
        file=file,
        deps=_public_photos_deps(),
    )


@router.delete("/pub/files/photos/{photo_id}")
async def photos_delete(
    photo_id: str,
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
):
    return await public_photos_runtime.photos_delete(
        photo_id,
        request,
        tenant=tenant,
        key=k,
        deps=_public_photos_deps(),
    )


@router.get("/pub/files/photos/{photo_id}")
async def photos_file(
    photo_id: str,
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
):
    return await public_photos_runtime.photos_file(
        photo_id,
        request,
        tenant=tenant,
        key=k,
        deps=_public_photos_deps(),
    )


@router.post("/pub/files/photos/{photo_id}/meta")
async def photos_update_meta(
    photo_id: str,
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
):
    return await public_photos_runtime.photos_update_meta(
        photo_id,
        request,
        tenant=tenant,
        key=k,
        deps=_public_photos_deps(),
    )


@router.get("/pub/catalog/view/{tenant}", response_class=HTMLResponse)
def catalog_view_public(tenant: int, request: Request):
    return catalog_public_runtime.catalog_view_public(
        tenant=tenant,
        request=request,
        deps=_catalog_view_deps(),
    )


@router.get("/pub/catalog/file/{tenant}")
def public_catalog_file(tenant: int):
    return catalog_public_runtime.public_catalog_file(
        tenant=tenant,
        deps=_catalog_view_deps(),
    )


# Public job status endpoint aligned with the new public upload path


def _load_catalog_status_payload(
    tenant_id: int, job_id: str
) -> tuple[dict[str, Any] | None, str | None]:
    return catalog_public_runtime.load_catalog_status_payload(
        tenant_id,
        job_id,
        deps=_catalog_public_deps(),
    )


@router.get("/pub/catalog/upload/status/{job_id}")
def catalog_upload_status(tenant: int, job_id: str, request: Request):
    return catalog_public_runtime.catalog_upload_status(
        tenant=tenant,
        job_id=job_id,
        request=request,
        deps=_catalog_public_deps(),
    )


@router.get("/pub/catalog/status/{job_id}")
def public_catalog_status(
    request: Request,
    job_id: str,
    tenant: int = Query(...),
    k: str | None = Query(None),
):
    return catalog_public_runtime.public_catalog_status(
        request=request,
        job_id=job_id,
        tenant=tenant,
        key=k,
        deps=_catalog_public_deps(),
    )


def _sanitize_catalog_status_public(payload: Any) -> Any:
    return catalog_public_runtime.sanitize_catalog_status_public(payload)


@router.get("/pub/catalog/status")
def catalog_status_public(
    request: Request,
    tenant: str = Query(...),
    job: str = Query(...),
    k: str | None = Query(None),
):
    return catalog_public_runtime.catalog_status_public(
        request=request,
        tenant=tenant,
        job=job,
        key=k,
        deps=_catalog_public_deps(),
    )


router.include_router(oauth_router)
router.include_router(max_router)
router.include_router(max_personal_router)
