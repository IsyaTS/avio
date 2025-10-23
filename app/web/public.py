from __future__ import annotations

import csv
import importlib
import io
import json
import logging
import math
import mimetypes
import os
import pathlib
import re
import hashlib
import sys
import time
import uuid
import asyncio
import base64
import random
import secrets
import html
from typing import Any, Iterable, Mapping, Optional

import qrcode
from qrcode.image.svg import SvgImage

from fastapi import APIRouter, File, Request, UploadFile, BackgroundTasks, HTTPException, Query
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse, Response, HTMLResponse
import httpx
import urllib.request
import urllib.error


def _import_alias(module: str):
    """Load module by bare name with ``app.<module>`` fallback."""

    try:
        return importlib.import_module(module)
    except ImportError:
        fallback = importlib.import_module(f"app.{module}")
        sys.modules.setdefault(module, fallback)
        return fallback


catalog_module = _import_alias("catalog")
catalog_index = _import_alias("catalog_index")

# NOTE: reference helpers locally to keep call sites compact
write_catalog_csv = catalog_module.write_catalog_csv
CatalogIndexError = catalog_index.CatalogIndexError
build_pdf_index = catalog_index.build_pdf_index
index_to_catalog_items = catalog_index.index_to_catalog_items

try:  # pragma: no cover - optional dependency during import time
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - openpyxl is optional in some environments
    load_workbook = None  # type: ignore[assignment]

try:
    from app.core import _normalize_catalog_items, settings  # type: ignore[attr-defined]
    import app.core as core_module  # type: ignore[attr-defined]
except ImportError:  # pragma: no cover - fallback for legacy layout
    try:
        from core import _normalize_catalog_items, settings  # type: ignore[attr-defined]
        core_module = _import_alias("core")
    except ImportError:
        core_module = _import_alias("core")
        _normalize_catalog_items = core_module._normalize_catalog_items
        settings = core_module.settings

from urllib.parse import quote, quote_plus, urlencode

from redis import exceptions as redis_ex

from config import tg_worker_url

from app.core import client as C
from app.metrics import MESSAGE_IN_COUNTER, DB_ERRORS_COUNTER
from app.db import insert_message_in, upsert_lead
from app.integrations import avito
from . import common as common
try:  # pragma: no cover - optional webhooks import
    from . import webhooks as webhook_module  # type: ignore
except ImportError:  # pragma: no cover - fallback when module alias missing
    try:
        from app.web import webhooks as webhook_module  # type: ignore
    except ImportError:
        webhook_module = None  # type: ignore[assignment]
from .ui import templates
from .webhooks import router as webhook_router, process_incoming

logger = logging.getLogger(__name__)
wa_logger = logging.getLogger("wa")
# Unified incoming transport log channel
message_in_logger = logging.getLogger("app.web.message_in")
_deprecated_hits: dict[str, float] = {}
# Avoid duplicate logging of WA messages via root logger handlers
wa_logger.propagate = False

TG_WORKER_BASE = tg_worker_url()
if not hasattr(C, "valid_key"):
    setattr(C, "valid_key", common.valid_key)

NO_STORE_CACHE_VALUE = "no-store, no-cache, must-revalidate"

PASSWORD_ATTEMPT_LIMIT = 2
PASSWORD_ATTEMPT_WINDOW = 60.0
_LOCAL_PASSWORD_ATTEMPTS: dict[tuple[int, str], list[float]] = {}

WA_QR_CACHE_TTL_MIN = 30  # seconds
WA_QR_CACHE_TTL_MAX = 60  # seconds

AVITO_STATE_PREFIX = "oauth:avito:state:"
AVITO_STATE_TTL = 600  # seconds

router = APIRouter()


def _qr_cache_ttl() -> int:
    return random.randint(WA_QR_CACHE_TTL_MIN, WA_QR_CACHE_TTL_MAX)


INCOMING_QUEUE_KEY = getattr(webhook_module, "INCOMING_QUEUE_KEY", "inbox:message_in")


def _no_store_headers(extra: Mapping[str, str] | None = None) -> dict[str, str]:
    headers = {
        "Cache-Control": NO_STORE_CACHE_VALUE,
        "Pragma": "no-cache",
        "Expires": "0",
    }
    if extra:
        headers.update(extra)
    return headers


def _resolve_client_key(request: Request | None) -> str:
    candidates: list[str | None] = []
    if request is not None:
        query_params = getattr(request, "query_params", None)
        if query_params is not None:
            candidates.append(query_params.get("k"))
            candidates.append(query_params.get("key"))
        headers = getattr(request, "headers", {}) or {}
        for header_name in ("X-Access-Key", "X-Client-Key", "X-Auth-Key"):
            candidates.append(headers.get(header_name))
        auth_header = headers.get("Authorization")
        if auth_header:
            token = auth_header.strip()
            if token.lower().startswith("bearer "):
                token = token[7:]
            candidates.append(token)
        cookies = getattr(request, "cookies", None) or {}
        if cookies:
            candidates.append(cookies.get("client_key"))
    for candidate in candidates:
        if not candidate:
            continue
        value = str(candidate).strip()
        if value:
            return value
    return ""


def _avito_state_key(state: str) -> str:
    return f"{AVITO_STATE_PREFIX}{state}"


def _avito_public_payload(raw: Mapping[str, Any] | None) -> dict[str, Any]:
    if not isinstance(raw, Mapping):
        return {"connected": False}
    access = str(raw.get("access_token") or "").strip()
    expires_at = raw.get("expires_at")
    try:
        expires_at_int = int(expires_at)
    except Exception:
        expires_at_int = None
    obtained_at = raw.get("obtained_at")
    try:
        obtained_at_int = int(obtained_at)
    except Exception:
        obtained_at_int = None
    info = {
        "connected": bool(access),
        "expires_at": expires_at_int,
        "obtained_at": obtained_at_int,
    }
    scope = raw.get("scope")
    if isinstance(scope, str) and scope.strip():
        info["scope"] = scope.strip()
    account_id = raw.get("account_id")
    if account_id is not None:
        try:
            info["account_id"] = int(account_id)
        except Exception:
            info["account_id"] = str(account_id)
    return info


def _avito_callback_html(ok: bool, message: str, payload: Mapping[str, Any]) -> str:
    safe_message = html.escape(message, quote=False)
    try:
        data_json = json.dumps(dict(payload), ensure_ascii=False)
    except Exception:
        data_json = json.dumps({"source": "avito-oauth", "ok": ok})
    status_class = "success" if ok else "error"
    return f"""<!doctype html>
<html lang="ru">
  <head>
    <meta charset="utf-8">
    <title>Avito OAuth</title>
    <style>
      body {{
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        padding: 32px;
        background: #f9fafb;
        color: #111827;
      }}
      .card {{
        max-width: 460px;
        margin: 0 auto;
        padding: 24px;
        background: #fff;
        border-radius: 12px;
        box-shadow: 0 10px 30px rgba(15, 23, 42, 0.08);
      }}
      .card h1 {{
        margin: 0 0 12px;
        font-size: 20px;
        font-weight: 700;
      }}
      .card p {{
        margin: 0 0 16px;
        line-height: 1.5;
      }}
      .status {{
        display: inline-block;
        padding: 6px 12px;
        border-radius: 999px;
        font-size: 13px;
        font-weight: 600;
      }}
      .status.success {{
        background: #dcfce7;
        color: #166534;
      }}
      .status.error {{
        background: #fee2e2;
        color: #b91c1c;
      }}
      .hint {{
        font-size: 13px;
        color: #6b7280;
      }}
    </style>
  </head>
  <body>
    <div class="card">
      <div class="status {status_class}">{'Успешно' if ok else 'Ошибка'}</div>
      <h1>Avito OAuth</h1>
      <p>{safe_message}</p>
      <p class="hint">Окно закроется автоматически. Если этого не произошло — закройте его вручную.</p>
    </div>
    <script>
      (function() {{
        var payload = {data_json};
        try {{
          if (typeof payload === 'object' && payload) {{
            payload.source = 'avito-oauth';
            payload.ok = { 'true' if ok else 'false' };
          }}
          if (window.opener && window.opener !== window) {{
            window.opener.postMessage(payload, '*');
          }}
        }} catch (err) {{}}
        setTimeout(function() {{
          try {{
            window.close();
          }} catch (err) {{}}
        }}, 2000);
      }})();
    </script>
  </body>
</html>"""


@router.post("/webhook/avito")
async def avito_webhook(request: Request) -> JSONResponse:
    try:
        raw_payload = await request.json()
    except json.JSONDecodeError as exc:  # pragma: no cover - defensive
        raise HTTPException(status_code=422, detail="invalid_json") from exc
    except Exception as exc:
        raise HTTPException(status_code=422, detail="invalid_payload") from exc

    events = raw_payload if isinstance(raw_payload, list) else [raw_payload]
    processed = 0
    for entry in events:
        if not isinstance(entry, Mapping):
            continue
        try:
            handled = await _handle_avito_webhook_event(entry, request)
        except HTTPException:
            raise
        except Exception:  # pragma: no cover - defensive logging
            logger.exception("avito_webhook_processing_failed")
            continue
        if handled:
            processed += 1

    return JSONResponse({"ok": True, "processed": processed})



async def _handle_avito_webhook_event(event: Mapping[str, Any], request: Request) -> bool:
    payload_raw = event.get("payload")
    payload = payload_raw if isinstance(payload_raw, Mapping) else {}

    tenant: Optional[int] = None
    account_id = _coerce_int(
        payload.get("account_id")
        or event.get("account_id")
        or (payload.get("account") or {}).get("id")
        or (event.get("account") or {}).get("id")
    )
    if account_id is not None:
        tenant = avito.find_tenant_by_account(account_id)

    if tenant is None:
        tenant = _coerce_int(payload.get("tenant") or event.get("tenant"))
    if tenant is None:
        tenant = _coerce_int(os.getenv("TENANT", "1"))

    if tenant is None or tenant <= 0:
        logger.warning("avito_webhook_skip reason=unknown_tenant account_id=%s", account_id)
        return False

    tenant = int(tenant)

    if account_id is None:
        integration = avito.get_integration(tenant)
        if integration and integration.get("account_id"):
            account_id = _coerce_int(integration.get("account_id"))
        if account_id is None:
            logger.warning("avito_webhook_skip reason=no_account_id tenant=%s", tenant)
            return False

    value_raw = payload.get("value") or event.get("value") or {}
    value = value_raw if isinstance(value_raw, Mapping) else {}
    if not value:
        logger.warning("avito_webhook_skip reason=no_value tenant=%s account_id=%s", tenant, account_id)
        return False

    content_raw = value.get("content") if isinstance(value.get("content"), Mapping) else {}

    chat_candidate = value.get("chat_id") or value.get("conversation_id")
    if isinstance(chat_candidate, Mapping):
        chat_candidate = chat_candidate.get("id")
    if chat_candidate is None:
        chat_candidate = payload.get("chat_id") or payload.get("conversation_id")
    if isinstance(chat_candidate, Mapping):
        chat_candidate = chat_candidate.get("id")
    chat_id = str(chat_candidate).strip() if chat_candidate else ""
    if not chat_id:
        logger.warning("avito_webhook_skip reason=no_chat account_id=%s tenant=%s", account_id, tenant)
        return False

    message_type = str(value.get("type") or "").strip().lower()
    text_candidate = ""
    if isinstance(content_raw, Mapping):
        text_candidate = content_raw.get("text") or ""
    if not text_candidate:
        text_candidate = value.get("text") or payload.get("text") or ""
    text = str(text_candidate or "").strip()

    attachments: list[dict[str, Any]] = []
    if isinstance(content_raw, Mapping):
        if message_type == "image":
            image = content_raw.get("image") if isinstance(content_raw.get("image"), Mapping) else {}
            sizes = image.get("sizes") if isinstance(image.get("sizes"), list) else []
            url = ""
            for entry in sizes:
                if isinstance(entry, Mapping) and entry.get("url"):
                    url = entry["url"]
            if url:
                attachments.append({"type": "image", "url": url, "name": image.get("name") or "image"})
        elif message_type == "voice":
            voice = content_raw.get("voice") if isinstance(content_raw.get("voice"), Mapping) else {}
            voice_id = voice.get("voice_id") or voice.get("id")
            if voice_id:
                attachments.append({"type": "voice", "url": voice_id})

    avito_user_id = _coerce_int(
        content_raw.get("author_id")
        or value.get("author_id")
        or value.get("sender_id")
        or payload.get("user_id")
    )
    if account_id is not None and avito_user_id is not None and avito_user_id == account_id:
        return False

    avito_login = None
    login_candidate = value.get("author_login") or payload.get("user_login")
    if isinstance(login_candidate, str) and login_candidate.strip():
        avito_login = login_candidate.strip()

    if not text and not attachments:
        logger.info("avito_webhook_skip reason=empty_message tenant=%s account_id=%s chat_id=%s", tenant, account_id, chat_id)
        return False

    message_id = value.get("id") or event.get("event_id") or event.get("id")
    message_id_str = str(message_id) if message_id is not None else None

    lead_id = avito.stable_lead_id(account_id, chat_id)

    incoming_body: dict[str, Any] = {
        "provider": "avito",
        "channel": "avito",
        "tenant": tenant,
        "tenant_id": tenant,
        "account_id": account_id,
        "chat_id": chat_id,
        "lead_id": lead_id,
        "avito_user_id": avito_user_id,
        "avito_login": avito_login,
        "source": {"type": "avito", "tenant": tenant, "account_id": account_id, "chat_id": chat_id},
        "message": {
            "id": message_id_str,
            "message_id": message_id_str,
            "text": text,
            "chat_id": chat_id,
            "direction": message_type,
            "attachments": attachments,
            "author_id": avito_user_id,
        },
        "attachments": attachments,
        "peer": chat_id,
        "auto_reply_handled": False,
    }

    created_at = value.get("created") or content_raw.get("created") or payload.get("created")
    if created_at is not None:
        incoming_body["message"]["created_at"] = created_at
    published_at = value.get("published_at") or payload.get("published_at")
    if published_at is not None:
        incoming_body["message"]["published_at"] = published_at

    lead_contacts = {"avito": {"peer": chat_id}}
    if avito_login:
        lead_contacts["avito"]["contact"] = avito_login
    incoming_body["lead_contacts"] = lead_contacts

    await process_incoming(incoming_body, request)
    return True


async def _ensure_avito_webhook(tenant: int, request: Request) -> None:
    target_url = common.public_url(request, "/webhook/avito")
    try:
        success = await avito.ensure_webhook(int(tenant), target_url)
    except avito.AvitoOAuthError as exc:
        logger.warning("avito_webhook_register_failed tenant=%s error=%s", tenant, exc)
    except Exception:
        logger.exception("avito_webhook_register_failed tenant=%s", tenant)
    else:
        if not success:
            logger.warning("avito_webhook_register_failed tenant=%s error=unexpected_response", tenant)


@router.get("/connect/avito")
def connect_avito(tenant: int, request: Request, k: str | None = None, key: str | None = None):
    tenant_id = int(tenant)
    access_key = (k or key or request.query_params.get("k") or request.query_params.get("key") or "").strip()
    if not common.valid_key(tenant_id, access_key):
        return JSONResponse({"detail": "invalid_key"}, status_code=401)

    common.ensure_tenant_files(tenant_id)
    cfg = common.read_tenant_config(tenant_id) or {}
    passport = cfg.get("passport", {}) if isinstance(cfg, dict) else {}
    brand = ""
    if isinstance(passport, dict):
        brand = str(passport.get("brand") or "").strip()

    avito_integration = avito.get_integration(tenant_id)
    avito_info = _avito_public_payload(avito_integration)

    behavior = cfg.setdefault("behavior", {})
    changed_behavior = False
    if behavior.get("auto_reply") is not True:
        behavior["auto_reply"] = True
        changed_behavior = True
    if behavior.get("auto_reply_enabled") is not True:
        behavior["auto_reply_enabled"] = True
        changed_behavior = True
    if changed_behavior:
        try:
            common.write_tenant_config(tenant_id, cfg)
        except Exception:
            logger.exception("avito_behavior_update_failed tenant=%s", tenant_id)

    primary_key = (common.get_tenant_pubkey(tenant_id) or "").strip()
    resolved_key = primary_key or access_key

    settings_link = ""
    try:
        raw_settings = request.url_for("client_settings", tenant=str(tenant_id))
        if resolved_key:
            settings_link = common.public_url(
                request,
                f"{raw_settings}?k={quote_plus(resolved_key)}",
            )
    except Exception:
        settings_link = ""

    context = {
        "request": request,
        "tenant": tenant_id,
        "key": resolved_key,
        "tenant_key": access_key,
        "subtitle": brand,
        "passport": passport if isinstance(passport, Mapping) else {},
        "avito": avito_info,
        "settings_link": settings_link,
    }
    return templates.TemplateResponse(request, "connect/avito.html", context)


def _tg_base_url() -> str:
    raw = getattr(settings, "TGWORKER_BASE_URL", "") or os.getenv("TGWORKER_BASE_URL") or ""
    base = str(raw).strip() or "http://tgworker:9000"
    return base.rstrip("/") or "http://tgworker:9000"


def _tg_make_url(path: str) -> str:
    if not path:
        return _tg_base_url()
    lowered = path.lower()
    if lowered.startswith("http://") or lowered.startswith("https://"):
        return path
    if not path.startswith("/"):
        path = f"/{path}"
    return f"{_tg_base_url()}{path}"


_TG_HTTP_CLIENT: httpx.AsyncClient | None = None


def _tg_admin_headers() -> dict[str, str]:
    token = getattr(settings, "ADMIN_TOKEN", "") or ""
    return {"X-Admin-Token": token}


def _tg_client() -> httpx.AsyncClient:
    global _TG_HTTP_CLIENT
    if _TG_HTTP_CLIENT is None or _TG_HTTP_CLIENT.is_closed:
        _TG_HTTP_CLIENT = httpx.AsyncClient(timeout=httpx.Timeout(10.0))
    return _TG_HTTP_CLIENT


async def _tg_call(
    method: str,
    path: str,
    *,
    params: Mapping[str, Any] | None = None,
    json: Mapping[str, Any] | None = None,
    timeout: float = 5,
    route: str | None = None,
    peer: Any | None = None,
) -> tuple[int, httpx.Response]:
    url = _tg_make_url(path)
    base_headers = _tg_admin_headers()
    request_kwargs: dict[str, Any] = {
        "params": dict(params or {}),
        "headers": base_headers,
        "follow_redirects": False,
        "timeout": httpx.Timeout(timeout),
    }
    if json is not None:
        request_kwargs["json"] = dict(json)
    try:
        client = _tg_client()
        response = await client.request(method.upper(), url, **request_kwargs)
    except httpx.HTTPError as exc:  # pragma: no cover - network failures
        detail = str(exc)
        logger.warning(
            "event=tg_proxy_error route=%s url=%s status=error detail=%s",
            route or path,
            url,
            detail,
        )
        raise TgWorkerCallError(url, detail) from exc

    status_code = int(getattr(response, "status_code", 0) or 0)
    peer_info = "-" if peer is None else str(peer)
    log_args = (route or path, url, status_code, peer_info)
    if status_code == 401:
        logger.warning(
            "event=tg_proxy_response route=%s url=%s status=%s peer=%s unauthorized",
            *log_args,
        )
    else:
        logger.info(
            "event=tg_proxy_response route=%s url=%s status=%s peer=%s",
            *log_args,
        )
    return status_code, response


def _log_deprecated(route: str) -> None:
    now = time.time()
    last = _deprecated_hits.get(route)
    if last is None or now - last >= 3600:
        _deprecated_hits[route] = now
        logger.warning("deprecated_endpoint route=%s", route)
def _stringify_detail(value: bytes | bytearray | str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, (bytes, bytearray)):
        try:
            return value.decode("utf-8", errors="ignore")
        except Exception:
            return ""
    return str(value)


_JSON_DBL_PASSWORD = re.compile(r'("password"\s*:\s*")([^"\\]*)(")', re.IGNORECASE)
_JSON_SGL_PASSWORD = re.compile(r"('password'\s*:\s*')([^'\\]*)(')", re.IGNORECASE)
_QUERY_PASSWORD = re.compile(r'(password\s*=\s*)([^&\s]+)', re.IGNORECASE)


def _mask_sensitive_detail(detail: str | None) -> str:
    if not detail:
        return ""
    masked = str(detail)
    masked = _JSON_DBL_PASSWORD.sub(lambda m: f"{m.group(1)}******{m.group(3)}", masked)
    masked = _JSON_SGL_PASSWORD.sub(lambda m: f"{m.group(1)}******{m.group(3)}", masked)
    masked = _QUERY_PASSWORD.sub(r"\1******", masked)
    return masked


def _extract_json_detail(body: bytes | bytearray | str | None) -> str | None:
    if body is None:
        return None
    data: Any
    payload = body
    if isinstance(payload, (bytes, bytearray)):
        try:
            payload = payload.decode("utf-8")
        except Exception:
            return None
    if isinstance(payload, str):
        payload = payload.strip()
        if not payload:
            return None
        try:
            data = json.loads(payload)
        except Exception:
            return None
    else:
        data = payload
    if isinstance(data, dict):
        detail = data.get("detail")
        if isinstance(detail, str):
            return detail
    return None


def _log_tg_proxy(
    route: str,
    tenant: int | str | None,
    status: int,
    body: bytes | bytearray | str | None,
    *,
    error: str | None = None,
    force: bool | None = None,
) -> None:
    detail_raw = error if error is not None else _stringify_detail(body)
    detail = _mask_sensitive_detail(detail_raw)
    log_fn = logger.info if 200 <= int(status or 0) < 300 else logger.warning
    tenant_value = "-" if tenant is None else tenant
    if route == "/pub/tg/password":
        log_fn("tg_proxy route=%s tenant=%s tg_code=%s", route, tenant_value, status)
        return
    force_fragment = " force=%s" % ("1" if force else "0") if force is not None else ""
    log_fn(
        "tg_proxy route=%s tenant=%s tg_code=%s%s detail=%s",
        route,
        tenant_value,
        status,
        force_fragment,
        detail or "",
    )


def _fingerprint_public_key(raw: str | None) -> str:
    if not raw:
        return "-"
    try:
        digest = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    except Exception:
        return "-"
    return digest[:10]


def _log_public_tg_request(route: str, tenant_id: int, key: str | None) -> None:
    fingerprint = _fingerprint_public_key(_normalize_public_token(key))
    logger.info(
        "tg_public_request route=%s tenant=%s key=%s",
        route,
        tenant_id,
        fingerprint,
    )


def _parse_force_flag(raw_value: str | None) -> bool:
    if raw_value is None:
        return False
    value = raw_value.strip().lower()
    return value in {"1", "true", "yes", "on"}


def _password_attempt_key(tenant_id: int, token: str) -> str:
    digest = hashlib.sha256(token.encode("utf-8")).hexdigest()[:32]
    return f"tenant:{int(tenant_id)}:twofa_attempts:{digest}"


def _build_public_tg_qr_url(
    tenant_id: int, key: str | None, qr_id: str | None = None
) -> str:
    parts: list[tuple[str, str]] = [("tenant", str(tenant_id))]
    normalized_key = _normalize_public_token(key)
    if normalized_key:
        parts.append(("k", normalized_key))
    if qr_id:
        parts.append(("qr_id", qr_id))
    return f"/pub/tg/qr.png?{urlencode(parts)}"


def _register_password_attempt(tenant_id: int, client_token: str) -> tuple[bool, int | None]:
    token = (client_token or "-").strip() or "-"
    key = _password_attempt_key(tenant_id, token)
    try:
        client = common.redis_client()
    except Exception:
        client = None

    if client is not None:
        try:
            pipe = client.pipeline()
            pipe.incr(key)
            pipe.ttl(key)
            attempts, ttl = pipe.execute()
            if attempts == 1 or ttl is None or ttl < 0:
                client.expire(key, int(PASSWORD_ATTEMPT_WINDOW))
                ttl = client.ttl(key)
            if attempts > PASSWORD_ATTEMPT_LIMIT:
                retry_after = int(ttl) if ttl and ttl > 0 else int(PASSWORD_ATTEMPT_WINDOW)
                return False, retry_after
            return True, None
        except redis_ex.RedisError:
            client = None

    now = time.monotonic()
    local_key = (int(tenant_id), token)
    entries = _LOCAL_PASSWORD_ATTEMPTS.setdefault(local_key, [])
    cutoff = now - PASSWORD_ATTEMPT_WINDOW
    filtered = [stamp for stamp in entries if stamp > cutoff]
    allowed = len(filtered) < PASSWORD_ATTEMPT_LIMIT
    retry_after: int | None = None
    if allowed:
        filtered.append(now)
    else:
        if filtered:
            remaining = PASSWORD_ATTEMPT_WINDOW - (now - filtered[0])
            retry_after = max(1, int(math.ceil(remaining))) if remaining > 0 else 1
        else:
            retry_after = int(PASSWORD_ATTEMPT_WINDOW)
    _LOCAL_PASSWORD_ATTEMPTS[local_key] = filtered
    return allowed, retry_after


def _client_identifier(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for", "")
    if forwarded:
        first = forwarded.split(",")[0].strip()
        if first:
            return first
    client = getattr(request, "client", None)
    host = getattr(client, "host", None) if client else None
    if host:
        return str(host)
    return "-"

MAX_UPLOAD_SIZE_BYTES = 25 * 1024 * 1024
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls", ".pdf"}
CSV_ENCODING_CANDIDATES = ["utf-8", "utf-8-sig", "cp1251", "windows-1251", "koi8-r"]


def _coerce_tenant(raw: int | str | None) -> int:
    if raw is None:
        raise ValueError("missing_tenant")
    if isinstance(raw, str):
        raw = raw.strip()
        if not raw:
            raise ValueError("missing_tenant")
    try:
        return int(raw)
    except (TypeError, ValueError) as exc:
        raise ValueError("invalid_tenant") from exc


def _normalize_headers(raw: Iterable[Any]) -> list[str]:
    normalized: list[str] = []
    seen: dict[str, int] = {}
    for idx, cell in enumerate(raw):
        text = "" if cell is None else str(cell)
        clean = text.strip().lstrip("\ufeff")
        if not clean:
            clean = f"column_{idx + 1}"
        if clean in seen:
            seen[clean] += 1
            clean = f"{clean}_{seen[clean]}"
        else:
            seen[clean] = 0
        normalized.append(clean)
    if not normalized:
        normalized.append("title")
    return normalized


def _relative_to(path: pathlib.Path, root: pathlib.Path) -> str:
    try:
        return str(path.relative_to(root))
    except Exception:
        return str(path)


def _make_safe_filename(filename: str, ext: str, *, fallback: str) -> str:
    base = pathlib.Path(filename).stem or fallback
    base = re.sub(r"[^0-9A-Za-z._-]+", "_", base)
    base = base.strip("._") or fallback
    return f"{base}{ext}"


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value)


def _read_csv_bytes(raw: bytes) -> tuple[list[dict[str, str]], dict[str, Any]]:
    encoding_used: str | None = None
    text: str | None = None
    for encoding in CSV_ENCODING_CANDIDATES:
        try:
            text = raw.decode(encoding)
            encoding_used = encoding
            break
        except UnicodeDecodeError:
            continue
    if text is None or encoding_used is None:
        raise ValueError("encoding_detection_failed")

    stream = io.StringIO(text)
    sample = stream.read(2048)
    stream.seek(0)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except Exception:
        delimiter = ","

    reader = csv.reader(stream, delimiter=delimiter)
    header: list[str] | None = None
    for row in reader:
        if row and any((cell or "").strip() for cell in row):
            header = _normalize_headers(row)
            break
    records: list[dict[str, str]] = []
    if header is None:
        header = ["title"]
    for row in reader:
        if not row or not any((_stringify(cell) for cell in row)):
            continue
        while len(header) < len(row):
            header.append(f"column_{len(header) + 1}")
        record: dict[str, str] = {}
        for idx, value in enumerate(row):
            key = header[idx]
            record[key] = _stringify(value)
        if any(record.values()):
            records.append(record)

    meta = {
        "type": "csv",
        "encoding": encoding_used,
        "delimiter": delimiter,
        "columns": header,
    }
    normalized = _normalize_catalog_items(records, meta)
    return normalized, meta


def _read_excel_bytes(raw: bytes) -> tuple[list[dict[str, str]], dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("excel_support_unavailable")

    workbook = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            header = ["title"]
        else:
            header = _normalize_headers(header_row)
        records: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            record: dict[str, str] = {}
            values = list(row)
            while len(header) < len(values):
                header.append(f"column_{len(header) + 1}")
            for idx, value in enumerate(values):
                key = header[idx]
                record[key] = _stringify(value)
            if any(record.values()):
                records.append(record)
    finally:
        workbook.close()

    meta = {
        "type": "excel",
        "columns": header,
        "sheet": sheet.title if sheet is not None else "Sheet1",
    }
    normalized = _normalize_catalog_items(records, meta)
    return normalized, meta


def _collapse_items_one_per_page(index, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_page: dict[str, dict[str, Any]] = {}
    def score(it: dict[str, Any]) -> tuple[int, int]:
        price = str(it.get("price") or "").strip()
        has_price = 1 if price else 0
        attr_count = len([k for k in it.keys() if k not in {"id", "title", "price"}])
        title_len = len(str(it.get("title") or ""))
        return (has_price, max(attr_count, title_len))

    def _is_attr_like_title(title: str) -> bool:
        t = (title or "").strip().lower()
        if not t:
            return True
        attr_tokens = (
            "толщина", "размер", "ширина", "высота", "диаметр",
            "материал", "цвет", "уплотнен", "замок", "замк",
        )
        if any(tok in t for tok in attr_tokens):
            # Allow if looks like model (letters+digits mixed)
            has_letter = any(ch.isalpha() for ch in t)
            has_digit = any(ch.isdigit() for ch in t)
            if has_letter and has_digit:
                return False
            return True
        return False

    def _strong_enough(it: dict[str, Any]) -> bool:
        price = str(it.get("price") or "").strip()
        has_price = bool(price)
        attr_count = len([k for k in it.keys() if k not in {"id", "title", "price", "page"}])
        if has_price:
            return True
        return attr_count >= 2 and not _is_attr_like_title(it.get("title") or "")

    for it in items:
        page = str(it.get("page") or "")
        if not page:
            continue
        current = by_page.get(page)
        # Prefer items that are strong-enough and not attribute-like titles
        if (current is None) or (score(it) > score(current)):
            by_page[page] = it

    # Optionally fabricate items for pages без распознанных блоков,
    # но не для стоп-разделов.
    try:
        STOP_RE = getattr(catalog_index, '_STOP_KEYWORDS_RE', None)
    except Exception:
        STOP_RE = None
    chunks = list(getattr(index, "chunks", []) or [])
    for ch in chunks:
        pg = str(getattr(ch, "page", "") or "")
        if not pg or pg in by_page:
            continue
        title = str(getattr(ch, "title", "") or "")
        if STOP_RE is not None and STOP_RE.search(title or ""):
            continue
        by_page[pg] = {"title": title, "price": "", "page": pg}

    def page_key(k: str) -> int:
        try:
            return int(k)
        except Exception:
            return 0

    # Build map for quick chunk title lookup
    chunk_title_by_page: dict[str, str] = {}
    for ch in getattr(index, "chunks", []) or []:
        pg = str(getattr(ch, "page", "") or "")
        if pg and pg not in chunk_title_by_page:
            chunk_title_by_page[pg] = str(getattr(ch, "title", "") or "")

    result: list[dict[str, Any]] = []
    for page in sorted(by_page.keys(), key=page_key):
        candidate = dict(by_page[page])
        # If кандидат выглядит как характеристика — заменим заголовок на заголовок чанка
        if _is_attr_like_title(candidate.get("title") or ""):
            chunk_title = chunk_title_by_page.get(page) or (candidate.get("title") or "")
            candidate["title"] = chunk_title
        # Удаляем служебные поля
        candidate.pop("page", None)
        result.append(candidate)
    return result


def _process_pdf(
    *,
    tenant: int,
    saved_path: pathlib.Path,
    tenant_root: pathlib.Path,
    saved_rel_path: pathlib.Path,
    original_name: str,
) -> tuple[list[dict[str, Any]], dict[str, Any], str | None]:
    index_dir = tenant_root / "indexes"
    index = build_pdf_index(
        saved_path,
        output_dir=index_dir,
        source_relpath=str(saved_rel_path),
        original_name=original_name,
    )
    items = index_to_catalog_items(index)
    # Optional: collapse to exactly one item per page if enabled in tenant behavior
    try:
        cfg = common.read_tenant_config(tenant)
        one_per_page = bool((cfg.get("behavior") or {}).get("pdf_one_item_per_page"))
    except Exception:
        one_per_page = False
    if one_per_page:
        items = _collapse_items_one_per_page(index, items)
    manifest_path = index.index_path.with_suffix(".manifest.json")
    manifest_rel = _relative_to(manifest_path, tenant_root) if manifest_path.exists() else None
    try:
        rel_index = _relative_to(index.index_path, tenant_root)
    except Exception:
        rel_index = str(index.index_path)
    meta: dict[str, Any] = {
        "type": "pdf",
        "index_path": rel_index,
        "indexed_at": index.generated_at,
        "chunk_count": index.chunk_count,
        "sha1": index.sha1,
        "page_count": index.page_count,
        "source_path": str(saved_rel_path),
        "original": original_name,
        "encoding": "utf-8-sig",
    }
    normalized = _normalize_catalog_items(items, meta)
    return normalized, meta, manifest_rel


def _coerce_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        candidate = int(str(value).strip())
    except Exception:
        return None
    return candidate


def _find_telegram_user_id(value: Any) -> int | None:
    candidate_keys = (
        "telegram_user_id",
        "telegramUserId",
        "user_id",
        "userId",
        "from_id",
        "fromId",
    )
    if isinstance(value, dict):
        for key in candidate_keys:
            if key in value:
                candidate = _coerce_int(value.get(key))
                if candidate and candidate > 0:
                    return candidate
        for nested in value.values():
            result = _find_telegram_user_id(nested)
            if result is not None:
                return result
    elif isinstance(value, list):
        for entry in value:
            result = _find_telegram_user_id(entry)
            if result is not None:
                return result
    return None


def _find_username(value: Any) -> str | None:
    if isinstance(value, dict):
        if isinstance(value.get("username"), str) and value["username"].strip():
            return value["username"].strip()
        for nested in value.values():
            result = _find_username(nested)
            if result:
                return result
    elif isinstance(value, list):
        for entry in value:
            result = _find_username(entry)
            if result:
                return result
    return None


@router.get("/connect/wa")
def connect_wa(tenant: int, request: Request, k: str | None = None):
    query_candidate = k or request.query_params.get("k") or ""
    guard = _ensure_valid_qr_request(tenant, query_candidate, request, query_param_only=True)
    if guard is None:
        return _invalid_key_response()

    tenant_id, resolved = guard
    tenant_id = int(tenant_id)
    resolved_key = resolved or ""

    if not resolved_key:
        items = common.list_keys(tenant_id)
        if items:
            resolved_key = items[0].get("key", "")

    common.ensure_tenant_files(tenant_id)
    cfg = common.read_tenant_config(tenant_id)
    persona = common.read_persona(tenant_id)
    passport = cfg.get("passport", {})
    subtitle = passport.get("brand") or "Подключение WhatsApp" if passport else "Подключение WhatsApp"
    persona_preview = "\n".join((persona or "").splitlines()[:6])

    settings_link = ""
    if resolved_key:
        raw_settings = request.url_for('client_settings', tenant=str(tenant_id))
        settings_link = common.public_url(request, f"{raw_settings}?k={quote_plus(resolved_key)}")

    context = {
        "request": request,
        "tenant": tenant_id,
        "key": resolved_key,
        "k": resolved_key,
        "timestamp": int(time.time()),
        "passport": passport,
        "persona_preview": persona_preview,
        "title": "Подключение WhatsApp",
        "subtitle": subtitle,
        "settings_link": settings_link,
        "public_base": common.public_base_url(request),
    }
    return templates.TemplateResponse(request, "connect/wa.html", context)


@router.get("/connect/tg")
def connect_tg(tenant: int, request: Request, k: str | None = None, key: str | None = None):
    tenant = int(tenant)
    access_key = (k or key or request.query_params.get("k") or request.query_params.get("key") or "").strip()
    if not common.valid_key(tenant, access_key):
        return JSONResponse({"detail": "invalid_key"}, status_code=401)

    common.ensure_tenant_files(tenant)
    cfg = common.read_tenant_config(tenant)
    passport = cfg.get("passport", {}) if isinstance(cfg, dict) else {}
    brand = ""
    if isinstance(passport, dict):
        brand = str(passport.get("brand") or "").strip()

    persona_text = common.read_persona(tenant)
    persona_preview = ""
    if persona_text:
        lines = str(persona_text).splitlines()
        persona_preview = "\n".join(lines[:6]).strip()

    primary_key = (common.get_tenant_pubkey(tenant) or "").strip()
    resolved_key = primary_key or access_key

    public_key = getattr(settings, "PUBLIC_KEY", "")
    encoded_public_key = quote_plus(public_key)
    tg_qr_url = f"/pub/tg/qr.png?k={encoded_public_key}" if public_key else "/pub/tg/qr.png"
    tg_status_url = f"/pub/tg/status?k={encoded_public_key}" if public_key else "/pub/tg/status"
    tg_start_url = f"/pub/tg/start?k={encoded_public_key}" if public_key else "/pub/tg/start"
    tg_twofa_url = f"/pub/tg/2fa?k={encoded_public_key}" if public_key else "/pub/tg/2fa"

    tg_connect_config = {
        "tenant": tenant,
        "key": public_key or resolved_key,
        "urls": {
            "public_key": public_key,
            "tg_status": "/pub/tg/status",
            "tg_status_url": tg_status_url,
            "tg_start": "/pub/tg/start",
            "tg_start_url": tg_start_url,
            "tg_qr_png": tg_qr_url,
            "tg_2fa": "/pub/tg/2fa",
            "tg_2fa_url": tg_twofa_url,
            "tg_password": "/pub/tg/2fa",
        },
    }

    context = {
        "request": request,
        "tenant": tenant,
        "key": public_key or resolved_key,
        "tenant_key": access_key,
        "subtitle": brand,
        "persona_preview": persona_preview,
        "tg_connect_config": tg_connect_config,
    }
    return templates.TemplateResponse(request, "connect/tg.html", context)


@router.get("/pub/wa/status")
async def wa_status(
    request: Request,
    tenant: int = Query(..., description="Tenant identifier"),
    k: str = Query(..., description="PUBLIC_KEY access token"),
):
    ok = _ensure_valid_qr_request(tenant, k, request, query_param_only=True)
    if ok is None:
        response = _invalid_key_response()
        return _as_head_response(response, request)
    tenant_id, validated_key = ok

    cached_qr_id, redis_failed = _get_last_qr_id(int(tenant_id))
    qr_id_override = None if redis_failed else cached_qr_id
    if redis_failed:
        wa_logger.info("wa_qr_cache_unavailable tenant=%s", tenant_id)

    snapshot = await _wa_status_impl(int(tenant_id))

    result = _compose_public_wa_response(
        int(tenant_id),
        validated_key,
        status_snapshot=snapshot,
        qr_id_override=qr_id_override,
    )

    effective_qr_id = None
    if qr_id_override:
        effective_qr_id = _normalize_qr_id(qr_id_override)
    if not effective_qr_id:
        effective_qr_id = _normalize_qr_id(result.get("qr_id"))

    if effective_qr_id:
        result["qr_id"] = effective_qr_id
        if validated_key:
            result["qr_url"] = _build_public_wa_qr_url(
                int(tenant_id), validated_key, effective_qr_id
            )
    else:
        result.pop("qr_id", None)
        if result.get("need_qr") and validated_key:
            result.setdefault("state", "qr")
            result["qr_url"] = _build_public_wa_qr_url(int(tenant_id), validated_key)
        elif not result.get("need_qr"):
            result.pop("qr_url", None)

    return JSONResponse(result, headers=_no_store_headers())


@router.get("/pub/wa/start")
async def wa_start(
    request: Request,
    tenant: int = Query(..., description="Tenant identifier"),
    k: str = Query(..., description="PUBLIC_KEY access token"),
):
    ok = _ensure_valid_qr_request(tenant, k, request, query_param_only=True)
    if ok is None:
        return _invalid_key_response()
    tenant_id, validated_key = ok

    webhook = common.webhook_url()
    payload = {"tenant_id": int(tenant_id), "webhook_url": webhook}
    try:
        response = await common.wa_post(
            f"/session/{int(tenant_id)}/start",
            payload,
        )
    except Exception:
        return JSONResponse({"error": "wa_unavailable"}, status_code=502)

    status_code = int(getattr(response, "status_code", 0) or 0)
    if status_code < 200 or status_code >= 400:
        return JSONResponse({"error": "wa_unavailable"}, status_code=502)

    try:
        response_data = response.json()
    except Exception:
        response_data = {}
    if not isinstance(response_data, dict):
        response_data = {}

    qr_id_value, redis_failed = _get_last_qr_id(int(tenant_id))
    if redis_failed:
        wa_logger.info("wa_qr_cache_unavailable tenant=%s", tenant_id)
        qr_id_value = None
    elif not qr_id_value:
        qr_id_value = _normalize_qr_id(response_data.get("qr_id") or response_data.get("qrId"))

    status_snapshot = await _wa_status_impl(int(tenant_id))

    result = _compose_public_wa_response(
        int(tenant_id),
        validated_key,
        status_snapshot=status_snapshot,
        qr_id_override=qr_id_value,
    )
    if result.get("need_qr") and not result.get("qr_url"):
        result["qr_url"] = _build_public_wa_qr_url(int(tenant_id), validated_key)
    if result.get("need_qr") and result.get("state") != "qr":
        result["state"] = "qr"

    return JSONResponse(result, headers=_no_store_headers())


async def _wa_status_impl(tenant: int) -> dict:
    cached_qr_id, redis_failed = _get_last_qr_id(int(tenant))

    code, raw = common.http(
        "GET", f"{common.WA_WEB_URL}/session/{int(tenant)}/status", timeout=3.0
    )
    try:
        data = json.loads(raw)
    except Exception:
        data = {}
    if not isinstance(data, dict):
        data = {}

    state_value, need_qr_flag = _derive_wa_state(data)
    qr_id_value = _normalize_qr_id(data.get("qr_id") or data.get("qrId"))
    if qr_id_value is None and cached_qr_id and not redis_failed:
        qr_id_value = cached_qr_id

    ready_flag = _truthy_flag(data.get("ready"))
    connected_flag = _truthy_flag(data.get("connected"))
    qr_flag = _truthy_flag(data.get("qr"))
    last_value = data.get("last")

    payload: dict[str, Any] = {
        "ok": True,
        "state": state_value,
        "status_code": int(code or 0),
        "raw": data,
        "need_qr": need_qr_flag,
    }
    payload["ready"] = bool(data.get("ready")) if "ready" in data else ready_flag
    payload["connected"] = (
        bool(data.get("connected")) if "connected" in data else connected_flag or ready_flag
    )
    payload["qr"] = bool(data.get("qr")) if "qr" in data else qr_flag
    if redis_failed:
        payload["qr_cache_unavailable"] = True
    if last_value is not None:
        payload["last"] = last_value
    if qr_id_value is not None:
        payload["qr_id"] = qr_id_value
    return payload


def _build_public_wa_qr_url(tenant: int, key: str, qr_id: str | None = None) -> str:
    params: dict[str, Any] = {"tenant": int(tenant), "k": str(key or "")}
    if qr_id:
        params["qr_id"] = str(qr_id)
    return f"/pub/wa/qr.svg?{urlencode(params, doseq=False)}"


def _normalize_qr_id(value: Any) -> str | None:
    if value is None:
        return None
    candidate = value
    if isinstance(candidate, (bytes, bytearray)):
        try:
            candidate = candidate.decode("utf-8", errors="ignore")
        except Exception:
            candidate = bytes(candidate).decode("utf-8", errors="ignore")
    if isinstance(candidate, bool):
        candidate = int(candidate)
    if isinstance(candidate, (int,)):
        return str(candidate)
    if isinstance(candidate, float):
        if not math.isfinite(candidate):
            return None
        if candidate.is_integer():
            return str(int(candidate))
        return str(int(candidate))
    text = str(candidate).strip()
    if not text:
        return None
    return text


def _derive_wa_state(data: Mapping[str, Any] | None) -> tuple[str | None, bool]:
    if not isinstance(data, Mapping):
        return None, False
    state_value = data.get("state")
    if state_value is not None:
        state_value = str(state_value)
    ready_flag = _truthy_flag(data.get("ready"))
    need_qr_flag = _truthy_flag(data.get("need_qr"))
    qr_flag = _truthy_flag(data.get("qr"))
    if state_value is None:
        if ready_flag:
            state_value = "ready"
        elif need_qr_flag or qr_flag:
            state_value = "qr"
        elif data.get("last") is not None:
            state_value = str(data.get("last"))
    if not need_qr_flag:
        need_qr_flag = not ready_flag and (qr_flag or state_value == "qr")
    return state_value, need_qr_flag


def _compose_public_wa_response(
    tenant: int,
    key: str | None,
    *,
    status_snapshot: Mapping[str, Any] | None = None,
    qr_id_override: str | None = None,
) -> dict[str, Any]:
    state_value: str | None = None
    need_qr_flag = False
    qr_id_value = qr_id_override
    raw_snapshot: Mapping[str, Any] | None = None

    result: dict[str, Any] = {"ok": True}

    if isinstance(status_snapshot, Mapping):
        raw_snapshot_candidate = status_snapshot.get("raw")
        if isinstance(raw_snapshot_candidate, Mapping):
            raw_snapshot = raw_snapshot_candidate
        else:
            raw_snapshot = status_snapshot
        state_candidate = status_snapshot.get("state")
        if state_candidate is not None:
            state_value = str(state_candidate)
        need_qr_flag = bool(status_snapshot.get("need_qr"))
        if qr_id_value is None:
            qr_id_value = _normalize_qr_id(status_snapshot.get("qr_id"))

        for snapshot_key, value in status_snapshot.items():
            if snapshot_key == "raw":
                continue
            if snapshot_key == "ok":
                result["ok"] = bool(value)
                continue
            result[snapshot_key] = value

    derived_state, derived_need_qr = _derive_wa_state(raw_snapshot)
    if state_value is None:
        state_value = derived_state
    if not need_qr_flag:
        need_qr_flag = derived_need_qr
    if qr_id_value is None and raw_snapshot is not None:
        qr_id_value = _normalize_qr_id(
            raw_snapshot.get("qr_id") or raw_snapshot.get("qrId")
        )

    if need_qr_flag and state_value != "qr":
        state_value = "qr"
    if state_value is not None:
        state_value = str(state_value)

    qr_url_value: str | None = None
    if key:
        if qr_id_value:
            qr_url_value = _build_public_wa_qr_url(int(tenant), key, qr_id_value)
        elif need_qr_flag:
            qr_url_value = _build_public_wa_qr_url(int(tenant), key)

    result.setdefault("tenant", int(tenant))
    if state_value is not None:
        result["state"] = state_value
    elif "state" in result and result["state"] is None:
        result.pop("state", None)

    result["need_qr"] = bool(need_qr_flag)

    if qr_id_value is not None:
        result["qr_id"] = qr_id_value
    else:
        result.pop("qr_id", None)

    if qr_url_value is not None:
        result["qr_url"] = qr_url_value
    else:
        result.pop("qr_url", None)

    return result

def _fetch_qr_bytes(url: str, timeout: float = 6.0):
    req = urllib.request.Request(url, method="GET")
    # Propagate waweb auth token if configured
    try:
        token = getattr(C, "WA_WEB_TOKEN", "") or getattr(C, "WA_INTERNAL_TOKEN", "") or ""
        if token:
            req.add_header("X-Auth-Token", token)
    except Exception:
        pass
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            body = resp.read()
            ctype = resp.headers.get("Content-Type", "")
            try:
                wa_logger.info("qr_upstream ok code=%s ctype=%s len=%s", getattr(resp, 'status', 200), ctype, len(body or b""))
            except Exception:
                pass
            return resp.status, ctype, body
    except urllib.error.HTTPError as e:
        try:
            data = e.read()
        except Exception:
            data = b""
        try:
            wa_logger.info("qr_upstream http_error code=%s len=%s", getattr(e, 'code', 0), len(data or b""))
        except Exception:
            pass
        return e.code, "", data
    except Exception as exc:  # pragma: no cover
        try:
            wa_logger.exception("qr_upstream failed: %s", exc)
        except Exception:
            pass
        return 0, "", b""


def _build_qr_candidates(tenant: int, cache_bust: int) -> list[tuple[str, str]]:
    base = common.WA_WEB_URL.rstrip("/")
    ts_param = f"ts={cache_bust}"
    return [
        (f"{base}/session/{tenant}/qr?format=svg&{ts_param}", "tenant_query_svg"),
        (f"{base}/session/{tenant}/qr.svg?{ts_param}", "tenant_ext_svg"),
        (f"{base}/session/{tenant}/qr.png?{ts_param}", "tenant_ext_png"),
        (f"{base}/session/qr?format=svg&{ts_param}", "global_query_svg"),
        (f"{base}/session/qr.svg?{ts_param}", "global_ext_svg"),
        (f"{base}/session/qr?format=png&{ts_param}", "global_query_png"),
        (f"{base}/session/qr.png?{ts_param}", "global_ext_png"),
    ]


def _proxy_qr_with_fallbacks(tenant: int) -> Response:
    wa_logger.info("qr_fetch start tenant=%s", tenant)
    if getattr(settings, "WA_PREFETCH_START", True):
        try:
            hook = common.webhook_url()
            payload = json.dumps({"tenant_id": int(tenant), "webhook_url": hook}, ensure_ascii=False).encode("utf-8")
            code, _ = common.http("POST", f"{common.WA_WEB_URL}/session/{int(tenant)}/start", body=payload, timeout=4.0)
            wa_logger.info("qr_prefetch_start code=%s", code)
        except Exception:
            wa_logger.info("qr_prefetch_start_failed")

    attempts_raw = getattr(settings, "WA_QR_FETCH_ATTEMPTS", 1) or 1
    try:
        attempts = max(1, int(attempts_raw))
    except (TypeError, ValueError):
        attempts = 1
    delay_raw = getattr(settings, "WA_QR_FETCH_RETRY_DELAY", 0.0) or 0.0
    try:
        retry_delay = max(0.0, float(delay_raw))
    except (TypeError, ValueError):
        retry_delay = 0.0

    last_status = 0
    last_stage = ""
    last_body_present = False
    last_content_type = ""

    for attempt in range(attempts):
        cache_bust = int(time.time() * 1000)
        candidates = _build_qr_candidates(tenant, cache_bust)
        for url, stage in candidates:
            wa_logger.info("qr_fetch url=%s stage=%s attempt=%s", url, stage, attempt + 1)
            status, ctype, body = _fetch_qr_bytes(url)
            last_status, last_stage = status, stage
            last_body_present = bool(body)
            last_content_type = (ctype or "").lower()
            wa_logger.info("upstream status=%s stage=%s attempt=%s", status, stage, attempt + 1)
            if int(status or 0) == 200 and last_content_type.startswith("image/") and body:
                headers = {
                    "Cache-Control": "no-store",
                    "X-Debug-Stage": f"served_qr:{stage}",
                }
                wa_logger.info("return=200 len=%s ctype=%s stage=%s attempt=%s", len(body or b""), ctype, stage, attempt + 1)
                return StreamingResponse(io.BytesIO(body), media_type=ctype, headers=headers)
        if attempt + 1 < attempts and retry_delay:
            wa_logger.info("qr_fetch_retry sleep=%s attempt=%s", retry_delay, attempt + 1)
            try:
                time.sleep(retry_delay)
            except Exception:
                wa_logger.info("qr_fetch_retry_sleep_failed attempt=%s", attempt + 1)

    headers = _no_store_headers()
    headers["Cache-Control"] = "no-store"
    if int(last_status or 0) in (204, 404) or (
        int(last_status or 0) == 200 and (not last_body_present or not last_content_type.startswith("image/"))
    ):
        stage_label = last_stage or "unknown"
        headers["X-Debug-Stage"] = f"no_content:{stage_label}"
        wa_logger.info("return=204 stage=%s status=%s", last_stage, last_status)
        return Response(status_code=204, headers=headers)

    headers["X-Debug-Stage"] = f"bad_gateway:{last_stage}" if last_stage else "bad_gateway"
    wa_logger.info("return=502 stage=%s status=%s", last_stage, last_status)
    return JSONResponse({"error": "wa_unavailable"}, status_code=502, headers=headers)


def _ensure_valid_qr_request(
    raw_tenant: int | str | None,
    raw_key: str | None,
    request: Request | None = None,
    *,
    query_param_only: bool = False,
) -> tuple[int, str] | None:
    try:
        tenant_id = _coerce_tenant(raw_tenant)
    except ValueError:
        return None

    if request is not None and _admin_token_valid(request):
        items = common.list_keys(tenant_id)
        if items:
            return tenant_id, items[0].get("key", "") or ""
        primary_key = (common.get_tenant_pubkey(tenant_id) or "").strip()
        return tenant_id, primary_key

    candidate = _resolve_public_key_candidate(
        raw_key,
        request,
        query_param_only=query_param_only,
    )
    if not candidate:
        return None

    expected = _expected_public_key_value()
    if expected and candidate == expected:
        items = common.list_keys(tenant_id)
        if items:
            return tenant_id, items[0].get("key", candidate)
        primary_key = (common.get_tenant_pubkey(tenant_id) or "").strip()
        return tenant_id, primary_key or candidate

    if common.valid_key(tenant_id, candidate):
        items = common.list_keys(tenant_id)
        if items:
            return tenant_id, items[0].get("key", candidate)
        return tenant_id, candidate

    return None


def _get_last_qr_id(tenant: int) -> tuple[str | None, bool]:
    key = f"wa:qr:last:{tenant}"
    try:
        client = common.redis_client()
        value = client.get(key)
    except redis_ex.RedisError:
        return None, True
    if not value:
        return None, False
    normalized = _normalize_qr_id(value)
    return normalized, False


def _load_cached_qr_entry(tenant: int, qr_id: str) -> tuple[dict[str, Any] | None, bool]:
    key = f"wa:qr:{tenant}:{qr_id}"
    try:
        client = common.redis_client()
        raw = client.get(key)
    except redis_ex.RedisError:
        return None, True
    entry: dict[str, Any] | None = None
    if raw is None:
        try:
            svg_value, png_value, txt_value = client.mget(
                f"{key}:svg",
                f"{key}:png",
                f"{key}:txt",
            )
        except redis_ex.RedisError:
            return None, True
        interim: dict[str, Any] = {}
        if svg_value:
            interim["qr_svg"] = svg_value
        if png_value:
            interim["qr_png"] = png_value
        if txt_value:
            interim["qr_text"] = txt_value
        if interim:
            entry = interim
        else:
            return None, False
    if isinstance(raw, bytes):
        raw = raw.decode('utf-8', errors='ignore')
    if entry is None:
        if isinstance(raw, str):
            stripped = raw.strip()
            if not stripped:
                return None, False
            try:
                parsed = json.loads(stripped)
            except Exception:
                parsed = None
            if isinstance(parsed, dict):
                entry = parsed
            else:
                if '<svg' in stripped.lower():
                    entry = {'qr_svg': stripped}
                else:
                    entry = {'qr_text': stripped}
        elif isinstance(raw, dict):
            entry = raw
    if entry is None:
        return None, False
    result: dict[str, Any] = dict(entry)
    result.setdefault('tenant', tenant)
    result.setdefault('ts', qr_id)
    if isinstance(result.get('qr_svg'), str) and not result['qr_svg'].strip():
        result.pop('qr_svg', None)
    if isinstance(result.get('qr_png'), str) and not result['qr_png'].strip():
        result.pop('qr_png', None)
    if isinstance(result.get('qr_text'), str) and not result['qr_text'].strip():
        result.pop('qr_text', None)
    return result, False


def _resolve_cached_qr(tenant: int) -> tuple[str | None, dict[str, Any] | None, bool]:
    qr_id, redis_failed = _get_last_qr_id(tenant)
    if redis_failed:
        return None, None, True
    if not qr_id:
        return None, None, False
    entry, entry_failed = _load_cached_qr_entry(tenant, qr_id)
    if entry_failed:
        return None, None, True
    if entry is None:
        return None, None, False
    return qr_id, entry, False


def _load_cached_svg(tenant: int, qr_id: str) -> tuple[str | None, bool]:
    key = f"wa:qr:{tenant}:{qr_id}:svg"
    try:
        client = common.redis_client()
        cached = client.get(key)
    except redis_ex.RedisError:
        return None, True
    if cached:
        candidate = cached
        if isinstance(candidate, (bytes, bytearray)):
            candidate = bytes(candidate).decode("utf-8", errors="ignore")
        candidate_str = str(candidate).strip()
        if candidate_str:
            return candidate_str, False
    entry, failed = _load_cached_qr_entry(tenant, qr_id)
    if failed:
        return None, True
    if entry is None:
        return None, False
    svg_value = entry.get("qr_svg") if isinstance(entry, dict) else None
    if isinstance(svg_value, str) and svg_value.strip():
        return svg_value, False
    qr_text = entry.get("qr_text") if isinstance(entry, dict) else None
    if isinstance(qr_text, str) and qr_text.strip():
        rendered = _render_qr_svg_from_text(qr_text.strip())
        if rendered:
            try:
                _cache_qr_payload(
                    tenant,
                    qr_id,
                    {"qr_svg": rendered, "qr_text": qr_text.strip()},
                    include_last=False,
                )
            except Exception:
                wa_logger.info("wa_qr_cache_update_failed tenant=%s qr_id=%s", tenant, qr_id)
            return rendered, False
    return None, False


def _qr_expired_response(qr_id: str | None = None) -> JSONResponse:
    headers = _no_store_headers()
    if qr_id:
        headers["X-WA-QR-ID"] = str(qr_id)
    return JSONResponse({"error": "qr_expired"}, status_code=410, headers=headers)


def _as_head_response(response: Response, request: Request) -> Response:
    if request.method.upper() != "HEAD":
        return response

    headers = dict(response.headers.items())
    media_type = response.media_type or headers.get("content-type") or headers.get("Content-Type")
    return Response(status_code=response.status_code, headers=headers, media_type=media_type)


def _render_qr_svg_from_text(qr_text: str) -> str | None:
    if not qr_text:
        return None
    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_Q,
        box_size=8,
        border=2,
    )
    qr.add_data(qr_text)
    qr.make(fit=True)
    img = qr.make_image(image_factory=SvgImage)
    buf = io.BytesIO()
    img.save(buf)
    return buf.getvalue().decode("utf-8")


def _render_qr_png_bytes(qr_text: str) -> bytes | None:
    if not qr_text:
        return None
    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_Q,
        box_size=10,
        border=2,
    )
    qr.add_data(qr_text)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#000000", back_color="#FFFFFF").convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _cache_qr_payload(
    tenant: int,
    qr_id: str,
    entry: Mapping[str, Any],
    *,
    include_last: bool = True,
) -> None:
    if not qr_id:
        return
    data = dict(entry or {})
    data.setdefault("tenant", tenant)
    data.setdefault("qr_id", qr_id)
    data.setdefault("ts", data.get("ts") or data.get("timestamp") or qr_id)

    svg_value = data.get("qr_svg")
    if isinstance(svg_value, str):
        svg_value = svg_value.strip() or None
    else:
        svg_value = None

    png_value = data.get("qr_png") or data.get("qr_png_base64")
    raw_png_bytes = data.get("qr_png_bytes")
    if png_value is None and isinstance(raw_png_bytes, (bytes, bytearray)):
        png_value = base64.b64encode(raw_png_bytes).decode("utf-8")
    if isinstance(png_value, (bytes, bytearray)):
        png_value = base64.b64encode(png_value).decode("utf-8")
    elif isinstance(png_value, str):
        png_value = png_value.strip() or None
    else:
        png_value = None

    txt_value = data.get("qr_text") or data.get("txt")
    if isinstance(txt_value, str):
        txt_value = txt_value.strip() or None
    else:
        txt_value = None

    serialisable = dict(data)
    if svg_value is None:
        serialisable.pop("qr_svg", None)
    else:
        serialisable["qr_svg"] = svg_value
    serialisable.pop("qr_png_bytes", None)
    if png_value is None:
        serialisable.pop("qr_png", None)
    else:
        serialisable["qr_png"] = png_value
    if txt_value is None:
        serialisable.pop("qr_text", None)
    else:
        serialisable["qr_text"] = txt_value

    try:
        json_payload = json.dumps(serialisable, ensure_ascii=False)
    except (TypeError, ValueError):
        json_payload = None

    try:
        client = common.redis_client()
        pipe = client.pipeline()
        wrote = False
        ttl = _qr_cache_ttl()
        if json_payload is not None:
            pipe.setex(f"wa:qr:{tenant}:{qr_id}", ttl, json_payload)
            wrote = True
        if svg_value is not None:
            pipe.setex(f"wa:qr:{tenant}:{qr_id}:svg", ttl, svg_value)
            wrote = True
        if png_value is not None:
            pipe.setex(f"wa:qr:{tenant}:{qr_id}:png", ttl, png_value)
            wrote = True
        if txt_value is not None:
            pipe.setex(f"wa:qr:{tenant}:{qr_id}:txt", ttl, txt_value)
            wrote = True
        if include_last:
            pipe.setex(f"wa:qr:last:{tenant}", ttl, qr_id)
            wrote = True
        if wrote:
            pipe.execute()
    except redis_ex.RedisError:
        wa_logger.info("wa_qr_cache_write_skip tenant=%s qr_id=%s", tenant, qr_id)


def _persist_qr_entry(tenant: int, qr_id: str, entry: Mapping[str, Any]) -> None:
    _cache_qr_payload(tenant, qr_id, entry, include_last=True)


async def _resolve_tenant_and_key(
    request: Request | None,
    raw_tenant: int | str | None,
    raw_key: str | None,
    *,
    query_keys: tuple[str, ...] = ("key", "k"),
    allow_body: bool = True,
) -> tuple[int | str | None, str | None]:
    tenant_candidate: int | str | None = raw_tenant
    key_candidate: str | None = raw_key

    if request is not None:
        if tenant_candidate is None:
            tenant_candidate = request.query_params.get("tenant")
        if not key_candidate:
            for query_key in query_keys:
                value = request.query_params.get(query_key)
                if value:
                    key_candidate = value
                    break

        needs_body = (
            allow_body and request.method.upper() in {"POST", "PUT", "PATCH"}
        )
        if needs_body and (tenant_candidate is None or not key_candidate):
            try:
                raw_body = await request.body()
            except Exception:
                raw_body = b""

            payload: dict[str, Any] = {}
            if raw_body:
                try:
                    decoded = raw_body.decode("utf-8")
                except UnicodeDecodeError:
                    decoded = ""
                if decoded:
                    try:
                        data = json.loads(decoded)
                    except json.JSONDecodeError:
                        data = {}
                    if isinstance(data, dict):
                        payload.update(data)

            if not payload:
                try:
                    form = await request.form()
                except Exception:
                    form = None
                if form is not None:
                    payload = {}
                    for form_key, value in form.multi_items():
                        if form_key not in payload:
                            payload[form_key] = value

            if tenant_candidate is None:
                tenant_candidate = payload.get("tenant")
            if not key_candidate:
                for query_key in query_keys:
                    value = payload.get(query_key)
                    if value:
                        key_candidate = value
                        break

    return tenant_candidate, key_candidate


def require_client_key(
    raw_tenant: int | str | None,
    raw_key: str | None,
) -> tuple[int, str] | Response:
    try:
        tenant_id = _coerce_tenant(raw_tenant)
    except ValueError:
        return JSONResponse({"error": "invalid_key"}, status_code=401, headers=_no_store_headers())

    key = "" if raw_key is None else str(raw_key).strip()
    if not key or not common.valid_key(tenant_id, key):
        return JSONResponse({"error": "invalid_key"}, status_code=401, headers=_no_store_headers())

    return tenant_id, key


def _normalize_public_token(value: str | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _expected_public_key_value() -> str:
    env_public = _normalize_public_token(os.getenv("PUBLIC_KEY"))
    if env_public:
        return env_public

    env_admin = _normalize_public_token(os.getenv("ADMIN_TOKEN"))
    if env_admin:
        return env_admin

    return _normalize_public_token(getattr(settings, "ADMIN_TOKEN", ""))


def _resolve_public_key_candidate(
    key_candidate: str | None,
    request: Request | None = None,
    *,
    query_param_only: bool = False,
) -> str:
    candidate = _normalize_public_token(key_candidate)
    if request is None:
        return candidate

    if query_param_only:
        return _normalize_public_token(request.query_params.get("k"))

    if not candidate:
        return _normalize_public_token(request.query_params.get("k"))

    return candidate


def _ensure_public_key(
    key_candidate: str | None,
    request: Request | None = None,
    *,
    query_param_only: bool = False,
) -> str | None:
    candidate = _resolve_public_key_candidate(
        key_candidate,
        request,
        query_param_only=query_param_only,
    )
    expected = _expected_public_key_value()
    if expected and candidate and candidate == expected:
        return candidate
    return None


def _invalid_key_response() -> JSONResponse:
    return JSONResponse(
        {"error": "invalid_key"},
        status_code=401,
        headers=_no_store_headers(),
    )


def _resolve_qr_identifier(primary: str | None, legacy: str | None = None) -> str:
    candidate = primary if primary is not None else legacy
    if candidate is None:
        return ""
    return str(candidate).strip()


def _admin_token_valid(request: Request) -> bool:
    token = request.headers.get("X-Admin-Token")
    return bool(token) and token == settings.ADMIN_TOKEN


def _has_public_tg_access(
    request: Request,
    key_candidate: str | None,
    *,
    allow_admin: bool = True,
    query_param_only: bool = False,
) -> tuple[bool, str | None]:
    if allow_admin and _admin_token_valid(request):
        return True, "admin"

    resolved = _ensure_public_key(
        key_candidate,
        request,
        query_param_only=query_param_only,
    )
    return (resolved is not None, resolved)


def _invalid_tenant_response(
    route: str,
    tenant_candidate: int | str | None,
    *,
    force: bool | None = None,
) -> JSONResponse:
    _log_tg_proxy(route, tenant_candidate, 400, None, error="invalid_tenant", force=force)
    return JSONResponse({"error": "invalid_tenant"}, status_code=400, headers=_no_store_headers())


def _unauthorized_response(
    route: str,
    tenant_id: int | str | None,
    *,
    force: bool | None = None,
) -> JSONResponse:
    _log_tg_proxy(route, tenant_id, 401, None, error="unauthorized", force=force)
    return _invalid_key_response()


def _tg_unavailable_response(
    route: str,
    tenant_id: int | str | None,
    detail: str | Exception | None,
    *,
    force: bool | None = None,
) -> JSONResponse:
    detail_text = _stringify_detail(str(detail)) if detail not in (None, "") else "tg_unavailable"
    if not detail_text:
        detail_text = "tg_unavailable"
    _log_tg_proxy(route, tenant_id, 0, None, error=detail_text, force=force)
    headers = _no_store_headers({"X-Telegram-Upstream-Status": "-"})
    body: dict[str, Any] = {"error": "tg_unavailable"}
    if detail_text and detail_text != "tg_unavailable":
        body["detail"] = detail_text
    return JSONResponse(body, status_code=502, headers=headers)


_UPSTREAM_HEADER_MAP = {"content-type": "Content-Type", "retry-after": "Retry-After"}


def _passthrough_upstream_response(
    route: str,
    tenant_id: int | str | None,
    upstream: Any,
    *,
    success_content_type: str | None = "application/json",
    error_content_type: str | None = "application/json",
    include_no_store: bool = True,
    force: bool | None = None,
) -> Response:
    status_code = int(getattr(upstream, "status_code", 0) or 0)
    body_bytes = bytes(getattr(upstream, "content", b"") or b"")
    if 200 <= status_code < 300:
        detail = None
    else:
        detail = (
            _stringify_detail(body_bytes)
            or _stringify_detail(getattr(upstream, "text", ""))
            or f"status_{status_code}"
        )
    _log_tg_proxy(route, tenant_id, status_code, body_bytes, error=detail, force=force)

    if status_code <= 0:
        headers = _no_store_headers({"X-Telegram-Upstream-Status": "-"})
        return JSONResponse({"error": "tg_unavailable"}, status_code=502, headers=headers)

    headers: dict[str, str] = {}
    if include_no_store:
        headers.update(_no_store_headers())
    headers["X-Telegram-Upstream-Status"] = str(status_code)

    upstream_headers = getattr(upstream, "headers", {}) or {}
    for name, value in upstream_headers.items():
        if not value:
            continue
        lowered = name.lower()
        mapped = _UPSTREAM_HEADER_MAP.get(lowered)
        if mapped:
            headers[mapped] = value

    default_content_type = success_content_type if 200 <= status_code < 300 else error_content_type
    if default_content_type and "Content-Type" not in headers:
        headers["Content-Type"] = default_content_type

    return Response(content=body_bytes, status_code=status_code, headers=headers)


def _proxy_headers(headers: Mapping[str, str] | None, status_code: int) -> dict[str, str]:
    allowed = {"content-type", "cache-control"}
    result: dict[str, str] = {}
    for name, value in (headers or {}).items():
        if not value:
            continue
        if name.lower() in allowed:
            result[name] = value
    result["Cache-Control"] = NO_STORE_CACHE_VALUE
    result["Pragma"] = "no-cache"
    result["Expires"] = "0"
    result["X-Telegram-Upstream-Status"] = str(status_code)
    return result


def _truthy_flag(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        normalized = value.strip().lower()
        return normalized in {"1", "true", "yes", "on"}
    return False


def _coerce_body_bytes(body: Any) -> bytes:
    if isinstance(body, (bytes, bytearray)):
        return bytes(body)
    if isinstance(body, str):
        return body.encode("utf-8")
    if body is None:
        return b""
    try:
        return json.dumps(body, ensure_ascii=False).encode("utf-8")
    except Exception:
        return b""


@router.api_route("/pub/wa/qr.svg", methods=["GET", "HEAD"])
async def wa_qr_svg(
    request: Request,
    tenant: int = Query(..., description="Tenant identifier"),
    k: str = Query(..., description="PUBLIC_KEY access token"),
    qr_id: str | None = Query(None, description="Explicit QR identifier from status"),
):
    ok = _ensure_valid_qr_request(tenant, k, request, query_param_only=True)
    if ok is None:
        response = _invalid_key_response()
        return _as_head_response(response, request)
    tenant_id, _ = ok

    requested_id = _normalize_qr_id(qr_id) if qr_id is not None else None
    query_params = request.query_params
    force_value = query_params.get("force")
    bypass_cache = False
    if "t" in query_params:
        bypass_cache = True
    elif force_value is not None:
        force_normalized = str(force_value).strip().lower()
        bypass_cache = force_normalized not in ("", "0", "false")

    redis_failed = False
    if not requested_id:
        requested_id, redis_failed = _get_last_qr_id(tenant_id)
    if redis_failed:
        headers = _no_store_headers()
        if requested_id:
            headers["X-WA-QR-ID"] = str(requested_id)
        response = JSONResponse({"error": "wa_cache_error"}, status_code=500, headers=headers)
        return _as_head_response(response, request)

    cached_svg: str | None = None
    if requested_id and not bypass_cache:
        cached_svg, redis_failed = _load_cached_svg(tenant_id, requested_id)
        if redis_failed:
            headers = _no_store_headers({"X-WA-QR-ID": str(requested_id)})
            response = JSONResponse({"error": "wa_cache_error"}, status_code=500, headers=headers)
            return _as_head_response(response, request)

    svg_value: str | None = cached_svg if not bypass_cache else None

    if bypass_cache or not svg_value:
        fallback_headers: dict[str, str] = {}
        if getattr(common, "WA_INTERNAL_TOKEN", ""):
            fallback_headers["X-Auth-Token"] = common.WA_INTERNAL_TOKEN
        timeout = httpx.Timeout(3.0, connect=2.0)
        try:
            async with httpx.AsyncClient(timeout=timeout) as client:
                response = await client.get(
                    f"{common.WA_WEB_URL}/session/{int(tenant_id)}/qr.svg",
                    headers=fallback_headers,
                )
        except httpx.HTTPError as exc:
            wa_logger.info(
                "wa_qr_upstream_error tenant=%s reason=%s",
                tenant_id,
                getattr(exc, "__class__", type(exc)).__name__,
            )
            response = JSONResponse({"error": "wa_unavailable"}, status_code=502)
            return _as_head_response(response, request)

        status_code = int(response.status_code or 0)
        if status_code == 404:
            response = _qr_expired_response(requested_id)
            return _as_head_response(response, request)
        if 400 <= status_code < 500:
            wa_logger.info(
                "wa_qr_upstream_status tenant=%s status=%s",
                tenant_id,
                status_code,
            )
            response = _qr_expired_response(requested_id)
            return _as_head_response(response, request)
        if status_code < 200 or status_code >= 300:
            wa_logger.info(
                "wa_qr_upstream_status tenant=%s status=%s",
                tenant_id,
                status_code,
            )
            response = JSONResponse({"error": "wa_unavailable"}, status_code=502)
            return _as_head_response(response, request)

        svg_candidate = response.text.strip()
        if not svg_candidate or not svg_candidate.lstrip().startswith("<svg"):
            wa_logger.info("wa_qr_upstream_invalid tenant=%s", tenant_id)
            response = JSONResponse({"error": "wa_unavailable"}, status_code=502)
            return _as_head_response(response, request)

        svg_value = svg_candidate
        upstream_qr_id = _normalize_qr_id(
            response.headers.get("X-WA-QR-ID")
            or response.headers.get("X-Wa-Qr-Id")
            or requested_id
        )
        if upstream_qr_id:
            requested_id = upstream_qr_id

        if requested_id:
            try:
                _cache_qr_payload(
                    tenant_id,
                    requested_id,
                    {"qr_svg": svg_value},
                    include_last=True,
                )
            except Exception:
                wa_logger.info(
                    "wa_qr_cache_store_failed tenant=%s qr_id=%s",
                    tenant_id,
                    requested_id,
                )

    if not svg_value:
        response = _qr_expired_response(requested_id)
        return _as_head_response(response, request)

    headers = {"Content-Type": "image/svg+xml"}
    headers.update(_no_store_headers())
    if requested_id:
        headers["X-WA-QR-ID"] = str(requested_id)
    response = Response(content=svg_value, media_type="image/svg+xml", headers=headers)
    return _as_head_response(response, request)


@router.api_route("/pub/tg/start", methods=["GET", "POST"])
async def tg_start(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
):
    route = "/pub/tg/start"
    tenant_candidate, key_candidate = await _resolve_tenant_and_key(
        request,
        tenant,
        k,
        query_keys=("k",),
        allow_body=request.method.upper() == "POST",
    )
    try:
        tenant_id = _coerce_tenant(tenant_candidate)
    except ValueError:
        return _invalid_tenant_response(route, tenant_candidate)

    allowed, validated_key = _has_public_tg_access(
        request,
        key_candidate,
        allow_admin=False,
        query_param_only=True,
    )
    if not allowed:
        return _unauthorized_response(route, tenant_id)

    _log_public_tg_request(route, tenant_id, validated_key)

    fallback_paths = ["/qr/start", "/rpc/start", "/session/start"]
    payload = {"tenant": tenant_id}
    last_error: str | None = None
    upstream: httpx.Response | None = None
    last_status: int | None = None

    for candidate in fallback_paths:
        try:
            status_code, response = await _tg_call("POST", candidate, json=payload, timeout=5.0)
        except TgWorkerCallError as exc:
            last_error = exc.detail
            continue
        upstream = response
        last_status = status_code
        break

    if upstream is None:
        reason = last_error or "tg_unavailable"
        headers = _no_store_headers({"X-Telegram-Upstream-Status": "-"})
        _log_tg_proxy(route, tenant_id, 502, None, error=reason)
        return JSONResponse(
            {"error": "tg_unavailable", "detail": reason},
            status_code=502,
            headers=headers,
        )

    status_code = int(last_status or upstream.status_code)
    body_bytes = bytes(upstream.content or b"")
    headers = _no_store_headers({"X-Telegram-Upstream-Status": str(status_code)})

    if 200 <= status_code < 300:
        content_type = upstream.headers.get("content-type")
        if content_type:
            headers["Content-Type"] = content_type
        _log_tg_proxy(route, tenant_id, status_code, body_bytes, error=None)
        return Response(content=body_bytes, status_code=status_code, headers=headers)

    detail_text = upstream.text if hasattr(upstream, "text") else ""
    reason = detail_text.strip() or f"status_{status_code}"
    _log_tg_proxy(route, tenant_id, status_code, body_bytes, error=reason)
    headers["Content-Type"] = "application/json"
    return JSONResponse(
        {"error": "tg_upstream", "detail": reason},
        status_code=status_code,
        headers=headers,
    )


async def _handle_tg_twofa(
    route: str,
    request: Request,
    tenant: int | str | None,
    key: str | None,
) -> Response:
    _log_deprecated(route)
    tenant_candidate, key_candidate = await _resolve_tenant_and_key(request, tenant, key)
    try:
        tenant_id = _coerce_tenant(tenant_candidate)
    except ValueError:
        return _invalid_tenant_response(route, tenant_candidate)

    allowed, validated_key = _has_public_tg_access(
        request,
        key_candidate,
        allow_admin=False,
    )
    if not allowed:
        return _unauthorized_response(route, tenant_id)

    _log_public_tg_request(route, tenant_id, validated_key)

    client_token = _client_identifier(request)

    password_value: str | None = None
    try:
        data = await request.json()
    except Exception:
        data = None
    if isinstance(data, dict) and "password" in data:
        candidate = data.get("password")
        if isinstance(candidate, str):
            password_value = candidate
    if password_value is None:
        try:
            form = await request.form()
        except Exception:
            form = None
        if form is not None:
            candidate = form.get("password")
            if isinstance(candidate, str):
                password_value = candidate

    password_text = (password_value or "").strip()
    if not password_text:
        _log_tg_proxy(route, tenant_id, 400, None, error="password_required")
        return JSONResponse({"error": "password_required"}, status_code=400, headers=_no_store_headers())

    allowed, retry_after = _register_password_attempt(tenant_id, client_token)
    if not allowed:
        headers = _no_store_headers()
        if retry_after and retry_after > 0:
            headers["Retry-After"] = str(int(retry_after))
        _log_tg_proxy(route, tenant_id, 429, None, error="flood_wait")
        body = {"error": "flood_wait"}
        if retry_after and retry_after > 0:
            body["retry_after"] = int(retry_after)
        return JSONResponse(body, status_code=429, headers=headers)

    fallback_paths = ["/2fa", "/rpc/twofa.submit"]
    upstream: httpx.Response | None = None
    last_error: str | None = None
    last_status: int | None = None
    payload_body = {"tenant": tenant_id, "password": password_text}

    for candidate in fallback_paths:
        try:
            status_code, response = await _tg_call("POST", candidate, json=payload_body, timeout=5.0)
        except TgWorkerCallError as exc:
            last_error = exc.detail
            continue
        upstream = response
        last_status = status_code
        break

    if upstream is None:
        reason = last_error or "tg_unavailable"
        headers = _no_store_headers({"X-Telegram-Upstream-Status": "-"})
        _log_tg_proxy(route, tenant_id, 502, None, error=reason)
        return JSONResponse({"error": "tg_unavailable", "detail": reason}, status_code=502, headers=headers)

    status_code = int(last_status or upstream.status_code)
    body_bytes = bytes(getattr(upstream, "content", b"") or b"")

    try:
        payload = upstream.json()
    except ValueError:
        payload = {}

    error_code = str(payload.get("error") or "").strip()
    headers = _no_store_headers({"X-Telegram-Upstream-Status": str(status_code or "-")})

    if status_code <= 0:
        _log_tg_proxy(route, tenant_id, status_code, body_bytes, error="tg_unavailable")
        headers["Content-Type"] = "application/json"
        return JSONResponse({"error": "tg_unavailable"}, status_code=502, headers=headers)

    if status_code == 401 or error_code == "bad_password":
        response = {"error": "bad_password"}
        detail = payload.get("detail")
        if detail:
            response["detail"] = detail
        _log_tg_proxy(route, tenant_id, 401, body_bytes, error="bad_password")
        headers["Content-Type"] = "application/json"
        return JSONResponse(response, status_code=401, headers=headers)

    if status_code == 409 and error_code:
        _log_tg_proxy(route, tenant_id, 409, body_bytes, error=error_code)
        headers["Content-Type"] = "application/json"
        return JSONResponse({"error": error_code}, status_code=409, headers=headers)

    if not (200 <= status_code < 300):
        failure = error_code or payload.get("detail") or f"status_{status_code}"
        _log_tg_proxy(route, tenant_id, status_code, body_bytes, error=failure)
        headers["Content-Type"] = "application/json"
        return JSONResponse({"error": failure}, status_code=502, headers=headers)

    state_value = str(payload.get("state") or payload.get("status") or "").strip()
    needs_twofa = bool(state_value == "need_2fa" or payload.get("needs_2fa"))
    response_payload = {
        "authorized": bool(payload.get("authorized")),
        "state": state_value,
        "needs_2fa": needs_twofa,
        "last_error": payload.get("last_error"),
        "expires_at": payload.get("expires_at"),
        "ok": bool(payload.get("ok", True)),
    }
    _log_tg_proxy(route, tenant_id, status_code, body_bytes, error=None)
    headers["Content-Type"] = "application/json"
    return JSONResponse(response_payload, headers=headers)


@router.post("/pub/tg/2fa")
async def tg_twofa(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    key: str | None = None,
):
    return await _handle_tg_twofa("/pub/tg/2fa", request, tenant, k or key)


@router.post("/pub/tg/twofa.submit")
async def tg_twofa_submit(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    key: str | None = None,
):
    return await _handle_tg_twofa("/pub/tg/twofa.submit", request, tenant, k or key)


@router.post("/pub/tg/password")
async def tg_password(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    key: str | None = None,
):
    return await _handle_tg_twofa("/pub/tg/password", request, tenant, k or key)


@router.post("/pub/tg/restart")
async def tg_restart(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    key: str | None = None,
):
    route = "/pub/tg/restart"
    _log_deprecated(route)
    tenant_candidate, key_candidate = await _resolve_tenant_and_key(request, tenant, k or key)
    try:
        tenant_id = _coerce_tenant(tenant_candidate)
    except ValueError:
        return _invalid_tenant_response(route, tenant_candidate, force=True)

    allowed, _ = _has_public_tg_access(request, key_candidate)
    if not allowed:
        return _unauthorized_response(route, tenant_id, force=True)

    try:
        upstream = await C.tg_post(
            "/session/restart",
            {"tenant_id": tenant_id},
            timeout=5.0,
        )
    except httpx.HTTPError as exc:
        return _tg_unavailable_response(route, tenant_id, exc, force=True)
    except Exception as exc:
        return _tg_unavailable_response(route, tenant_id, exc, force=True)

    return _passthrough_upstream_response(route, tenant_id, upstream, force=True)

@router.get("/pub/tg/status")
async def tg_status(request: Request, tenant: int | str | None = None, k: str | None = None):
    route = "/pub/tg/status"
    tenant_candidate, key_candidate = await _resolve_tenant_and_key(
        request,
        tenant,
        k,
        query_keys=("k",),
        allow_body=False,
    )
    try:
        tenant_id = _coerce_tenant(tenant_candidate)
    except ValueError:
        return _invalid_tenant_response(route, tenant_candidate)

    allowed, validated_key = _has_public_tg_access(
        request,
        key_candidate,
        allow_admin=False,
        query_param_only=True,
    )
    if not allowed:
        return _unauthorized_response(route, tenant_id)

    _log_public_tg_request(route, tenant_id, validated_key)

    fallback_paths = ["/status", "/rpc/status", "/session/status"]
    params = {"tenant": tenant_id}
    last_error: str | None = None
    upstream: httpx.Response | None = None
    last_status: int | None = None

    for candidate in fallback_paths:
        try:
            status_code, response = await _tg_call("GET", candidate, params=params, timeout=5.0)
        except TgWorkerCallError as exc:
            last_error = exc.detail
            continue
        if not (200 <= status_code < 300):
            detail_text = response.text if hasattr(response, "text") else ""
            last_error = detail_text.strip() or f"status_{status_code}"
            continue
        upstream = response
        last_status = status_code
        break

    if upstream is None:
        reason = last_error or "tg_unavailable"
        headers = _no_store_headers({"X-Telegram-Upstream-Status": "-"})
        _log_tg_proxy(route, tenant_id, 502, None, error=reason)
        return JSONResponse({"error": "tg_unavailable", "detail": reason}, status_code=502, headers=headers)

    status_code = int(last_status or upstream.status_code)
    body_bytes = bytes(upstream.content or b"")
    headers = _no_store_headers({"X-Telegram-Upstream-Status": str(status_code)})
    content_type = upstream.headers.get("content-type")
    if content_type:
        headers["Content-Type"] = content_type
    _log_tg_proxy(route, tenant_id, status_code, body_bytes, error=None)
    return Response(content=body_bytes, status_code=status_code, headers=headers)


@router.get("/pub/tg/qr.png")
async def tg_qr_png(
    request: Request,
    tenant: int | str | None = None,
    qr_id: str | None = None,
    k: str | None = None,
):
    route = "/pub/tg/qr.png"
    tenant_candidate, key_candidate = await _resolve_tenant_and_key(
        request,
        tenant,
        k,
        query_keys=("k",),
        allow_body=False,
    )
    try:
        tenant_id = _coerce_tenant(tenant_candidate)
    except ValueError:
        return _invalid_tenant_response(route, tenant_candidate)

    allowed, validated_key = _has_public_tg_access(
        request,
        key_candidate,
        allow_admin=False,
        query_param_only=True,
    )
    if not allowed:
        return _unauthorized_response(route, tenant_id)

    qr_identifier = _resolve_qr_identifier(qr_id, request.query_params.get("id"))
    if not qr_identifier:
        _log_tg_proxy(route, tenant_id, 400, None, error="missing_qr_id")
        return JSONResponse({"error": "missing_qr_id"}, status_code=400, headers=_no_store_headers())

    safe_qr = quote(qr_identifier, safe="")
    base_params = {"tenant": tenant_id}
    fallback_paths: list[tuple[str, dict[str, Any]]] = [
        ("/qr/png", {**base_params, "qr_id": qr_identifier}),
        (f"/session/qr/{safe_qr}.png", dict(base_params)),
    ]

    upstream: httpx.Response | None = None
    last_status: int | None = None
    last_error: str | None = None

    for candidate, params in fallback_paths:
        try:
            status_code, response = await _tg_call("GET", candidate, params=params, timeout=5.0)
        except TgWorkerCallError as exc:
            last_error = exc.detail
            continue
        if not (200 <= status_code < 300):
            detail_text = response.text if hasattr(response, "text") else ""
            last_error = detail_text.strip() or f"status_{status_code}"
            continue
        upstream = response
        last_status = status_code
        break

    if upstream is None:
        reason = last_error or "tg_unavailable"
        headers = _no_store_headers({"X-Telegram-Upstream-Status": "-"})
        _log_tg_proxy(route, tenant_id, 502, None, error=reason)
        return JSONResponse({"error": "tg_unavailable", "detail": reason}, status_code=502, headers=headers)

    status_code = int(last_status or upstream.status_code)
    body_bytes = bytes(upstream.content or b"")
    headers = _no_store_headers({"X-Telegram-Upstream-Status": str(status_code)})
    headers["Cache-Control"] = "no-store"
    content_type = upstream.headers.get("content-type") or "image/png"
    headers["Content-Type"] = content_type
    _log_tg_proxy(route, tenant_id, status_code, body_bytes, error=None)
    return Response(content=body_bytes, status_code=status_code, headers=headers)


@router.get("/pub/tg/qr.txt")
def tg_qr_txt(request: Request, qr_id: str | None = None, k: str | None = None, key: str | None = None):
    key_candidate = k or key or request.query_params.get("k") or request.query_params.get("key")
    allowed, _ = _has_public_tg_access(request, key_candidate)
    if not allowed:
        return _unauthorized_response("/pub/tg/qr.txt", None)
    qr_value = _resolve_qr_identifier(qr_id, request.query_params.get("id"))
    if not qr_value:
        _log_tg_proxy("/pub/tg/qr.txt", None, 400, None, error="missing_qr_id")
        return JSONResponse(
            {"error": "missing_qr_id"},
            status_code=400,
            headers=_no_store_headers(),
        )

    safe_qr = quote(qr_value, safe="")
    status_code, body, headers = common.tg_http(
        "GET",
        f"{TG_WORKER_BASE}/session/qr/{safe_qr}.txt",
        timeout=15.0,
    )
    body_bytes = body if isinstance(body, (bytes, bytearray)) else ("" if body is None else str(body)).encode("utf-8")
    detail_from_json = _extract_json_detail(body_bytes)
    if status_code == 200:
        detail = None
    elif detail_from_json:
        detail = detail_from_json
    else:
        detail = _stringify_detail(body_bytes) or f"status_{status_code}"

    _log_tg_proxy("/pub/tg/qr.txt", None, status_code, body_bytes, error=detail)

    if status_code <= 0:
        return JSONResponse(
            {"error": "tg_unavailable"},
            status_code=502,
            headers=_no_store_headers({"X-Telegram-Upstream-Status": str(status_code)}),
        )

    if status_code == 404:
        detail_value = detail_from_json or "qr_not_found"
        headers_out = _proxy_headers(headers or {}, status_code)
        headers_out.update(_no_store_headers())
        if not body_bytes:
            body_bytes = json.dumps({"detail": detail_value}).encode("utf-8")
        media_type = headers_out.get("Content-Type") or "application/json"
        return Response(
            content=body_bytes,
            status_code=status_code,
            headers=headers_out,
            media_type=media_type,
        )

    if status_code != 200:
        headers_out = _proxy_headers(headers or {}, status_code)
        headers_out.update(_no_store_headers())
        return JSONResponse({"error": "tg_unavailable"}, status_code=502, headers=headers_out)

    response_headers = _proxy_headers(headers or {}, status_code)
    response_headers.update(_no_store_headers())
    response_headers.setdefault("Content-Type", "text/plain; charset=utf-8")
    return Response(content=body_bytes, status_code=status_code, headers=response_headers)


@router.api_route("/pub/tg/logout", methods=["GET", "POST"])
async def tg_logout(
    request: Request,
    tenant: int | str | None = None,
    k: str | None = None,
    key: str | None = None,
):
    route = "/pub/tg/logout"
    tenant_candidate, key_candidate = await _resolve_tenant_and_key(request, tenant, k or key)
    try:
        tenant_id = _coerce_tenant(tenant_candidate)
    except ValueError:
        return _invalid_tenant_response(route, tenant_candidate)

    allowed, _ = _has_public_tg_access(request, key_candidate)
    if not allowed:
        return _unauthorized_response(route, tenant_id)

    try:
        upstream = await C.tg_post(
            "/session/logout",
            {"tenant_id": tenant_id},
            timeout=5.0,
        )
    except httpx.HTTPError as exc:
        return _tg_unavailable_response(route, tenant_id, exc)
    except Exception as exc:
        return _tg_unavailable_response(route, tenant_id, exc)

    return _passthrough_upstream_response(route, tenant_id, upstream)


@router.get("/pub/wa/qr.png")
def wa_qr_png(
    request: Request,
    tenant: int = Query(..., description="Tenant identifier"),
    k: str = Query(..., description="PUBLIC_KEY access token"),
    qr_id: str | None = Query(None, description="Explicit QR identifier from status"),
):
    ok = _ensure_valid_qr_request(tenant, k, request, query_param_only=True)
    if ok is None:
        return _invalid_key_response()
    tenant_id, _ = ok

    requested_id = (qr_id or "").strip()
    redis_failed = False
    if not requested_id:
        requested_id, redis_failed = _get_last_qr_id(tenant_id)
    if redis_failed:
        return JSONResponse({"error": "wa_cache_unavailable"}, status_code=503)
    if not requested_id:
        return _qr_expired_response()

    entry, redis_failed = _load_cached_qr_entry(tenant_id, requested_id)
    if redis_failed:
        return JSONResponse({"error": "wa_cache_unavailable"}, status_code=503)

    png_value = entry.get("qr_png") if isinstance(entry, dict) else None
    binary: bytes | None = None
    mutated = False
    if isinstance(png_value, str) and png_value.strip():
        normalized = png_value.split(",")[-1].strip()
        try:
            binary = base64.b64decode(normalized, validate=False)
        except Exception:
            binary = None
            wa_logger.warning("wa_qr_cache_invalid_png tenant=%s qr_id=%s", tenant_id, requested_id)

    if binary is None:
        qr_text = entry.get("qr_text") if isinstance(entry, dict) else None
        if isinstance(qr_text, str) and qr_text.strip():
            binary = _render_qr_png_bytes(qr_text.strip())
            if binary:
                entry["qr_png"] = base64.b64encode(binary).decode("ascii")
                mutated = True

    if binary is None:
        return _qr_expired_response(requested_id)

    if mutated:
        try:
            entry_to_store = dict(entry)
            _persist_qr_entry(tenant_id, requested_id, entry_to_store)
        except Exception:
            wa_logger.info("wa_qr_cache_update_failed tenant=%s qr_id=%s format=png", tenant_id, requested_id)

    headers = _no_store_headers()
    headers["X-WA-QR-ID"] = requested_id
    return Response(content=binary, media_type="image/png", headers=headers)


@router.get("/pub/wa/restart")
async def wa_restart(
    request: Request,
    tenant: int = Query(..., description="Tenant identifier"),
    k: str = Query(..., description="PUBLIC_KEY access token"),
):
    """Force-restart waweb session to issue a fresh QR.

    Security: requires a valid public access key `k` for the tenant.
    """

    ok = _ensure_valid_qr_request(tenant, k, request)
    if ok is None:
        return _invalid_key_response()
    tenant_id, _ = ok

    wa_logger.info("wa_restart click tenant=%s", tenant_id)

    try:
        webhook = common.webhook_url()
        start_payload = json.dumps({"tenant_id": tenant_id, "webhook_url": webhook}, ensure_ascii=False).encode("utf-8")
        empty_payload = json.dumps({}, ensure_ascii=False).encode("utf-8")

        code_restart, _ = common.http(
            "POST",
            f"{common.WA_WEB_URL}/session/{tenant_id}/restart",
            body=start_payload,
        )
        if 200 <= int(code_restart or 0) < 300:
            wa_logger.info("wa_restart success tenant=%s stage=tenant_restart code=%s", tenant_id, code_restart)
            return JSONResponse({"ok": True})

        code_logout, _ = common.http(
            "POST",
            f"{common.WA_WEB_URL}/session/{tenant_id}/logout",
            body=empty_payload,
        )
        code_start, _ = common.http(
            "POST",
            f"{common.WA_WEB_URL}/session/{tenant_id}/start",
            body=start_payload,
        )
        if 200 <= int(code_start or 0) < 300:
            wa_logger.info(
                "wa_restart success tenant=%s stage=tenant_logout_start logout=%s start=%s",
                tenant_id,
                code_logout,
                code_start,
            )
            return JSONResponse({"ok": True})

        code_global_restart, _ = common.http("POST", f"{common.WA_WEB_URL}/session/restart", body=start_payload)
        if 200 <= int(code_global_restart or 0) < 300:
            wa_logger.info(
                "wa_restart success tenant=%s stage=global_restart code=%s",
                tenant_id,
                code_global_restart,
            )
            return JSONResponse({"ok": True})

        code_global_start, _ = common.http("POST", f"{common.WA_WEB_URL}/session/start", body=start_payload)
        if 200 <= int(code_global_start or 0) < 300:
            wa_logger.info(
                "wa_restart success tenant=%s stage=global_start code=%s",
                tenant_id,
                code_global_start,
            )
            return JSONResponse({"ok": True})

        wa_logger.info(
            "wa_restart failed tenant=%s codes=%s",
            tenant_id,
            {
                "tenant_restart": code_restart,
                "tenant_logout": code_logout,
                "tenant_start": code_start,
                "global_restart": code_global_restart,
                "global_start": code_global_start,
            },
        )
        return JSONResponse({"error": "wa_unavailable"}, status_code=502)
    except Exception as exc:  # pragma: no cover
        try:
            wa_logger.exception("wa_restart_failed: %s", exc)
        except Exception:
            pass
        return JSONResponse({"error": "wa_unavailable"}, status_code=502)


def _resolve_public_settings_key(request: Request, key_candidate: str | None) -> str:
    candidate = (key_candidate or "").strip()
    if candidate:
        return candidate

    query_value = (request.query_params.get("k") or "").strip()
    if query_value:
        return query_value

    cookies = getattr(request, "cookies", None) or {}
    return (cookies.get("client_key") or "").strip()


def _authorize_public_settings_request(
    request: Request,
    tenant: int | str | None,
    key_candidate: str | None,
) -> tuple[int, str] | Response:
    try:
        tenant_id = _coerce_tenant(tenant)
    except ValueError as exc:
        return JSONResponse({"detail": str(exc)}, status_code=400)

    resolved_key = _resolve_public_settings_key(request, key_candidate)
    if not resolved_key or not common.valid_key(tenant_id, resolved_key):
        return JSONResponse({"detail": "invalid_key"}, status_code=401)

    return tenant_id, resolved_key


@router.get("/pub/settings/get")
def settings_get(request: Request, tenant: int | str | None = None, k: str | None = None):
    auth = _authorize_public_settings_request(request, tenant, k)
    if isinstance(auth, Response):
        return auth

    tenant_id, _ = auth
    common.ensure_tenant_files(tenant_id)
    cfg = common.read_tenant_config(tenant_id)
    persona = common.read_persona(tenant_id)
    return {"ok": True, "cfg": cfg, "persona": persona}


@router.post("/pub/settings/save")
async def settings_save(request: Request, tenant: int | str | None = None, k: str | None = None):
    auth = _authorize_public_settings_request(request, tenant, k)
    if isinstance(auth, Response):
        return auth

    tenant_id, _ = auth
    common.ensure_tenant_files(tenant_id)
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    cfg = common.read_tenant_config(tenant_id)
    if isinstance(payload.get("cfg"), dict):
        cfg = payload["cfg"]
    else:
        for section in ["passport", "behavior", "cta", "limits", "integrations", "learning"]:
            if isinstance(payload.get(section), dict):
                cfg.setdefault(section, {}).update(payload[section])
        if isinstance(payload.get("catalogs"), list):
            cfg["catalogs"] = payload["catalogs"]
    common.write_tenant_config(tenant_id, cfg)
    if isinstance(payload.get("persona"), str):
        common.write_persona(tenant_id, payload.get("persona") or "")
    return {"ok": True}


# Move public catalog upload off the client namespace to avoid route collisions
# with the client router. The tenant is accepted as a query parameter.
@router.post("/pub/catalog/upload")
async def catalog_upload(
    tenant: int,
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
):
    from . import client as client_module

    tenant_id = int(tenant)
    key = client_module._resolve_key(request, request.query_params.get("k"))
    authorized = client_module._auth(tenant_id, key)
    if not authorized:
        header_key = (request.headers.get("X-Access-Key") or "").strip()
        query_key = (request.query_params.get("k") or request.query_params.get("key") or "").strip()
        if key and key == header_key:
            authorized = True
        elif key and query_key and key == query_key:
            authorized = True
    if not authorized:
        return JSONResponse({"detail": "invalid_key"}, status_code=401)

    filename = (file.filename or "").strip()
    if not filename:
        return JSONResponse({"ok": False, "error": "empty_file"}, status_code=400)

    ext = pathlib.Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return JSONResponse({"ok": False, "error": "unsupported_type"}, status_code=400)

    raw = await file.read()
    if not raw:
        return JSONResponse({"ok": False, "error": "empty_file"}, status_code=400)
    if len(raw) > MAX_UPLOAD_SIZE_BYTES:
        return JSONResponse(
            {
                "ok": False,
                "error": "file_too_large",
                "max_size_bytes": MAX_UPLOAD_SIZE_BYTES,
            },
            status_code=400,
        )

    common.ensure_tenant_files(tenant_id)
    tenant_root = pathlib.Path(common.tenant_dir(tenant_id))
    uploads_dir = tenant_root / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)

    safe_name = _make_safe_filename(filename, ext, fallback=f"catalog_{uuid.uuid4().hex}")
    saved_upload_path = uploads_dir / safe_name
    saved_upload_path.write_bytes(raw)
    saved_upload_rel = pathlib.Path(_relative_to(saved_upload_path, tenant_root))
    relative_path = str(saved_upload_rel)

    job_id = uuid.uuid4().hex
    job_root = tenant_root / "catalog_jobs" / job_id
    job_root.mkdir(parents=True, exist_ok=True)
    status_path = job_root / "status.json"

    status_state: dict[str, Any] = {
        "job_id": job_id,
        "state": "pending",
        "error": None,
        "log": [],
        "filename": filename,
        "message": "",
    }

    def write_status(status: str | None = None, **fields: Any) -> None:
        if status is not None:
            status_state["state"] = status
        status_state["updated_at"] = int(time.time())
        for key, value in fields.items():
            status_state[key] = value
        status_path.write_text(json.dumps(status_state, ensure_ascii=False, indent=2), encoding="utf-8")

    def append_log(level: str, message: str, **extra: Any) -> None:
        entry = {"ts": int(time.time()), "level": level, "message": message}
        if extra:
            entry.update({k: v for k, v in extra.items() if v is not None})
        status_state.setdefault("log", []).append(entry)
        write_status(None, log=status_state["log"])

    def fail(error_key: str, *, http_status: int = 400, **details: Any):
        append_log("error", error_key, **details)
        write_status("failed", error=error_key, message=error_key, **details)
        return JSONResponse({"ok": False, "error": error_key, "job_id": job_id, **details}, status_code=http_status)

    mime_type, _ = mimetypes.guess_type(filename)
    write_status("received", size=len(raw), mime=mime_type, source_path=relative_path)
    append_log("info", "file_received", size=len(raw), mime=mime_type)

    # Build background job that performs heavy processing to avoid request timeouts
    def process_job() -> None:
        try:
            write_status("processing")
            append_log("info", "job_started")
            base_name = pathlib.Path(filename).stem or f"catalog_{job_id}"
            normalized_rows: list[dict[str, Any]]
            meta: dict[str, Any]
            manifest_rel: str | None = None

            # Read back from disk to keep memory footprint small
            try:
                if ext == ".csv":
                    file_bytes = saved_upload_path.read_bytes()
                    normalized_rows, meta = _read_csv_bytes(file_bytes)
                elif ext in {".xlsx", ".xls"}:
                    file_bytes = saved_upload_path.read_bytes()
                    normalized_rows, meta = _read_excel_bytes(file_bytes)
                else:
                    normalized_rows, meta, manifest_rel = _process_pdf(
                        tenant=tenant_id,
                        saved_path=saved_upload_path,
                        tenant_root=tenant_root,
                        saved_rel_path=saved_upload_rel,
                        original_name=filename,
                    )
            except CatalogIndexError as exc:
                logger.warning("PDF indexing failed", exc_info=exc)
                fail("pdf_index_failed", detail=str(exc))
                return
            except Exception as exc:  # pragma: no cover - defensive
                logger.exception("catalog processing failed", exc_info=exc)
                fail("processing_failed", detail=str(exc))
                return

            parsed_count = len(normalized_rows)
            append_log("info", "rows_parsed", items=parsed_count)

            try:
                result = write_catalog_csv(tenant_id, normalized_rows, base_name, meta)
            except Exception as exc:  # pragma: no cover - disk errors
                logger.exception("write_catalog_csv raised", exc_info=exc)
                fail("csv_write_failed", detail=str(exc))
                return

            if not isinstance(result, tuple) or len(result) != 2:
                logger.error("write_catalog_csv returned unexpected result", extra={"result": result})
                fail("csv_write_failed")
                return

            csv_rel_path, ordered_columns = result
            pipeline_info = meta.get("pipeline") if isinstance(meta, dict) else None
            items = int(meta.get("items", parsed_count)) if isinstance(meta, dict) else parsed_count
            if manifest_rel:
                meta = dict(meta)
                meta["manifest_path"] = manifest_rel

            write_status(
                "done",
                csv_path=csv_rel_path,
                items=items,
                columns=ordered_columns,
                metadata=meta,
                source_path=relative_path,
                message="completed",
            )
            if manifest_rel:
                write_status(None, manifest_path=manifest_rel)
            append_log("info", "csv_written", items=items, columns=len(ordered_columns), pipeline=pipeline_info)

            # Persist config updates
            cfg = common.read_tenant_config(tenant_id)
            if not isinstance(cfg, dict):
                cfg = {}
            catalogs = cfg.get("catalogs") if isinstance(cfg.get("catalogs"), list) else []
            catalog_type = "pdf" if ext == ".pdf" else ("excel" if ext in {".xlsx", ".xls"} else "csv")
            catalog_entry: dict[str, Any] = {
                "name": "uploaded",
                "path": relative_path,
                "type": catalog_type,
            }
            detected_encoding = _stringify(meta.get("encoding")) if isinstance(meta, dict) else ""
            if catalog_type == "csv":
                if detected_encoding:
                    catalog_entry["encoding"] = detected_encoding
                if isinstance(meta, dict) and "delimiter" in meta:
                    catalog_entry["delimiter"] = meta.get("delimiter")
            elif detected_encoding:
                catalog_entry["encoding"] = detected_encoding
            if catalog_type == "pdf":
                if isinstance(meta, dict):
                    for key in ("index_path", "indexed_at", "chunk_count", "sha1"):
                        if meta.get(key) is not None:
                            catalog_entry[key] = meta.get(key)

            if csv_rel_path:
                catalog_entry["csv_path"] = csv_rel_path

            cfg["catalogs"] = [catalog_entry] + [entry for entry in catalogs if entry.get("path") != relative_path]

            integrations = cfg.setdefault("integrations", {})
            uploaded_meta: dict[str, Any] = {
                "path": relative_path,
                "original": filename,
                "uploaded_at": int(time.time()),
                "type": catalog_type,
                "size": len(raw),
                "mime": mime_type or "application/octet-stream",
                "csv_path": csv_rel_path,
            }
            if pipeline_info:
                uploaded_meta["pipeline"] = pipeline_info
            if catalog_type == "csv":
                if detected_encoding:
                    uploaded_meta["encoding"] = detected_encoding
                if isinstance(meta, dict) and "delimiter" in meta:
                    uploaded_meta["delimiter"] = meta.get("delimiter")
            if catalog_type == "pdf" and isinstance(meta, dict):
                index_meta = {
                    "path": meta.get("index_path"),
                    "generated_at": meta.get("indexed_at"),
                    "chunks": meta.get("chunk_count"),
                    "pages": meta.get("page_count"),
                    "sha1": meta.get("sha1"),
                }
                index_meta = {k: v for k, v in index_meta.items() if v is not None}
                if index_meta:
                    uploaded_meta["index"] = index_meta
            uploaded_meta = {k: v for k, v in uploaded_meta.items() if v is not None}
            integrations["uploaded_catalog"] = uploaded_meta

            common.write_tenant_config(tenant_id, cfg)
            append_log("info", "config_updated", catalog_type=catalog_type)
        except Exception as exc:  # final safety net
            logger.exception("catalog job crashed", exc_info=exc)
            fail("job_crashed", detail=str(exc))

    # Enqueue the job and return immediately to avoid Cloudflare 524 timeouts
    if background_tasks is not None:
        background_tasks.add_task(process_job)
    else:
        # Fallback: run inline (tests) but still return fast behavior below
        try:
            process_job()
        except Exception:
            pass

    # HTML form fallback: redirect back to settings quickly
    accept_header = (request.headers.get("accept") or "").lower()
    sec_fetch_mode = (request.headers.get("sec-fetch-mode") or "").lower()
    sec_fetch_dest = (request.headers.get("sec-fetch-dest") or "").lower()
    wants_html = (
        "text/html" in accept_header
        or "application/xhtml+xml" in accept_header
        or sec_fetch_mode == "navigate"
        or sec_fetch_dest == "document"
    )
    if wants_html and (request.headers.get("x-requested-with", "").lower() == "xmlhttprequest"):
        wants_html = False
    if wants_html:
        redirect_url = request.url_for("client_settings", tenant=str(tenant_id))
        if key:
            redirect_url = f"{redirect_url}?k={quote_plus(key)}"
        return RedirectResponse(url=redirect_url, status_code=303)

    # Return job descriptor for polling client
    return JSONResponse({"ok": True, "job_id": job_id, "state": "queued"})


# Public job status endpoint aligned with the new public upload path
@router.get("/pub/catalog/upload/status/{job_id}")
def catalog_upload_status(tenant: int, job_id: str, request: Request):
    from . import client as client_module

    tenant_id = int(tenant)
    key = client_module._resolve_key(request, request.query_params.get("k"))
    authorized = client_module._auth(tenant_id, key)
    if not authorized:
        header_key = (request.headers.get("X-Access-Key") or "").strip()
        query_key = (request.query_params.get("k") or request.query_params.get("key") or "").strip()
        if key and key == header_key:
            authorized = True
        elif key and query_key and key == query_key:
            authorized = True
    if not authorized:
        return JSONResponse({"detail": "invalid_key"}, status_code=401)

    tenant_root = pathlib.Path(common.tenant_dir(tenant_id))
    status_path = tenant_root / "catalog_jobs" / job_id / "status.json"
    if not status_path.exists():
        return JSONResponse({"ok": False, "error": "not_found"}, status_code=404)
    try:
        data = json.loads(status_path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning("status read failed", exc_info=exc)
        return JSONResponse({"ok": False, "error": "status_read_failed"}, status_code=500)
    return JSONResponse({"ok": True, **data})
