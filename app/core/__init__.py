from __future__ import annotations
import os, json, re, csv, asyncio, pathlib, time, random, hashlib, logging
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field
import urllib.request, urllib.error

# Redis (асинхронный клиент можно использовать при необходимости)
import redis.asyncio as redis_async
import redis as redis_sync
from redis import exceptions as redis_ex

# OpenAI как опциональная зависимость
try:
    import openai  # type: ignore
except Exception:  # библиотека может быть не установлена
    openai = None  # type: ignore

try:
    from ..brain import planner, quality
except Exception:  # pragma: no cover
    import importlib

    planner = importlib.import_module("app.brain.planner")  # type: ignore
    quality = importlib.import_module("app.brain.quality")  # type: ignore

try:
    from ..catalog import retriever as catalog_retriever  # type: ignore
except Exception:  # pragma: no cover
    catalog_retriever = None
try:
    from ..training import retriever as training_retriever  # type: ignore
except Exception:  # pragma: no cover
    training_retriever = None

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # опциональная зависимость для Excel
    load_workbook = None  # type: ignore


BASE_DIR = pathlib.Path(__file__).resolve().parent.parent
ROOT_DIR = BASE_DIR.parent
DATA_DIR = pathlib.Path(os.getenv("APP_DATA_DIR") or (BASE_DIR / "data"))


_public_key_warning_logged = False


def _resolve_public_key(admin_token: str) -> str:
    """Return PUBLIC_KEY with ADMIN_TOKEN fallback and warn once."""

    global _public_key_warning_logged

    raw_value = os.getenv("PUBLIC_KEY")
    normalized = "" if raw_value is None else str(raw_value).strip()
    if normalized:
        return normalized

    if not _public_key_warning_logged:
        logger.warning("PUBLIC_KEY is empty; falling back to ADMIN_TOKEN")
        _public_key_warning_logged = True

    return admin_token


def _resolve_tenants_dir() -> pathlib.Path:
    env_value = os.getenv("TENANTS_DIR")
    if env_value:
        return pathlib.Path(env_value)

    app_tenants = ROOT_DIR / "app" / "tenants"
    try:
        app_tenants.mkdir(parents=True, exist_ok=True)
        return app_tenants
    except OSError:
        pass

    default_parent = ROOT_DIR / "data"
    if default_parent.exists():
        return default_parent / "tenants"

    fallback = DATA_DIR / "tenants"
    fallback.mkdir(parents=True, exist_ok=True)
    return fallback


TENANTS_DIR = _resolve_tenants_dir()

# Lightweight in-memory caches (mtime-based invalidation)
_TENANT_CONFIG_CACHE: Dict[int, Tuple[float, dict]] = {}
_TENANT_PERSONA_CACHE: Dict[int, Tuple[float, str]] = {}
# Key: (tenant or None, tuple of (path, mtime, size)) -> parsed, normalized items
_CATALOG_CACHE: Dict[Tuple[Optional[int], Tuple[Tuple[str, float, int], ...]], List[Dict[str, Any]]] = {}


def _env_bool(name: str, default: bool = False) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "yes", "on"}


# Precompiled regexes reused across hot paths
_FIELD_CLEAN_RE = re.compile(r"[^0-9a-zA-Zа-яА-ЯёЁ]+")
_PERSONA_HINTS_KEY_RE = re.compile(
    r"^(greeting|приветств(?:ие|уй)|cta|призыв|closing|завершение|tone|тон|language|язык|max(?:imum)?\s*(?:questions|вопросов|уточнений))\s*[:\-]\s*(.+)$",
    re.IGNORECASE,
)


class Settings:
    APP_VERSION   = os.getenv("APP_VERSION", "v21.0")
    SEND          = os.getenv("SEND_ENABLED", "true").lower() == "true"

    REDIS_URL     = os.getenv("REDIS_URL", "redis://redis:6379/0")
    r = redis_async.from_url(REDIS_URL, decode_responses=True)

    # Публичный URL API (для вебхука waweb)
    APP_PUBLIC_URL   = (os.getenv("APP_PUBLIC_URL") or "").rstrip("/")
    APP_INTERNAL_URL = os.getenv("APP_INTERNAL_URL", "http://app:8000").rstrip("/")

    # waweb
    WA_WEB_URL    = (os.getenv("WA_WEB_URL", "http://waweb:8088") or "http://waweb:8088").rstrip("/")
    WA_PREFETCH_START = _env_bool("WA_PREFETCH_START", True)

    # Админка
    ADMIN_TOKEN   = (os.getenv("ADMIN_TOKEN") or "sueta").strip()
    PUBLIC_KEY    = _resolve_public_key(ADMIN_TOKEN)
    WEBHOOK_SECRET = (os.getenv("WEBHOOK_SECRET", "") or "").strip()

    # LLM
    OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
    OPENAI_MODEL   = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
    try:
        OPENAI_TIMEOUT_SECONDS = float(os.getenv("OPENAI_TIMEOUT_SECONDS", "4"))
    except ValueError:
        OPENAI_TIMEOUT_SECONDS = 4.0

    # Бизнес-поля
    AGENT_NAME    = os.getenv("AGENT_NAME", "Акакий")
    BRAND_NAME    = os.getenv("BRAND_NAME", "Гермес")
    WHATSAPP_LINK = os.getenv("WHATSAPP_LINK", "https://wa.me/7XXXXXXXXXX")
    CITY          = os.getenv("CITY", "Уфа")

    # Персоны/промпты с диска
    PERSONA_MD    = os.getenv("PERSONA_MD") or str(DATA_DIR / "persona.md")


settings = Settings()


_openai_client: Any | None = None
_openai_client_key: str | None = None

_sync_redis_client: redis_sync.Redis | None = None


def _resolve_chat_completion_callable(obj: Any):
    chat = getattr(obj, "chat", None)
    if chat is None:
        return None
    completions = getattr(chat, "completions", None)
    if completions is None:
        return None
    create_fn = getattr(completions, "create", None)
    if not callable(create_fn):
        return None
    return create_fn


def _get_openai_client() -> Any | None:
    """Return an OpenAI client compatible with chat.completions.create."""

    global _openai_client, _openai_client_key

    if not (openai and settings.OPENAI_API_KEY):
        return None

    if hasattr(openai, "OpenAI"):
        if _openai_client is None or _openai_client_key != settings.OPENAI_API_KEY:
            try:
                _openai_client = openai.OpenAI(api_key=settings.OPENAI_API_KEY)  # type: ignore[attr-defined]
            except TypeError:
                _openai_client = openai.OpenAI()  # type: ignore[attr-defined]
            except Exception as exc:  # pragma: no cover - сетевые/валидационные ошибки
                logger.warning("openai client init failed: %s", exc)
                _openai_client = None
                return None
            _openai_client_key = settings.OPENAI_API_KEY

        if _openai_client is None:
            return None

        if _resolve_chat_completion_callable(_openai_client) is None:
            logger.warning("openai client missing chat.completions.create")
            return None

        return _openai_client

    if not hasattr(openai, "OpenAI"):
        _openai_client = None
        _openai_client_key = None

    try:
        setattr(openai, "api_key", settings.OPENAI_API_KEY)  # type: ignore[attr-defined]
    except Exception as exc:  # pragma: no cover - старые клиенты без api_key
        logger.warning("failed to set openai api_key: %s", exc)
        return None

    if _resolve_chat_completion_callable(openai) is None:
        logger.warning("openai module missing chat.completions.create")
        return None

    return openai


def _redis_sync_client() -> redis_sync.Redis:
    global _sync_redis_client
    if _sync_redis_client is None:
        _sync_redis_client = redis_sync.from_url(settings.REDIS_URL, decode_responses=True)
    return _sync_redis_client


def _with_sync_redis(func, default=None):
    global _sync_redis_client
    for _ in range(2):
        try:
            return func(_redis_sync_client())
        except redis_ex.ConnectionError:
            _sync_redis_client = None
        except redis_ex.RedisError:
            return default
    return default


# Куки и ключи
ADMIN_COOKIE        = "admin_token"
TENANT_PUBKEYS_HASH = "tenant_pubkeys"


# --------------------------- состояние диалогов -----------------------------
STATE_KEY_PREFIX = "sales_state"
STATE_TTL_SECONDS = 8 * 3600
_STATE_CACHE: Dict[str, "SalesState"] = {}


def _state_key(tenant: int | None, contact_id: int | None) -> str:
    tenant_id = int(tenant or 0)
    contact = int(contact_id or 0)
    return f"{STATE_KEY_PREFIX}:{tenant_id}:{contact}"


def _state_store_read(key: str) -> Optional[dict]:
    try:
        raw = _with_sync_redis(lambda client: client.get(key), None)
        if not raw:
            cached = _STATE_CACHE.get(key)
            if cached:
                return cached.to_dict()
            return None
        return json.loads(raw)
    except Exception:
        return None


def _state_store_write(key: str, payload: dict) -> None:
    _with_sync_redis(
        lambda client: client.setex(key, STATE_TTL_SECONDS, json.dumps(payload, ensure_ascii=False)),
        None,
    )


@dataclass
class SalesState:
    tenant: int
    contact_id: int
    channel: str = "whatsapp"
    needs: Dict[str, Any] = field(default_factory=dict)
    spin: Dict[str, str] = field(default_factory=lambda: {stage: "pending" for stage in ("s", "p", "i", "n")})
    bant: Dict[str, Any] = field(default_factory=dict)
    asked_questions: List[str] = field(default_factory=list)
    challenger_cursor: int = 0
    social_proof_cursor: int = 0
    scarcity_cursor: int = 0
    reciprocity_cursor: int = 0
    history: List[Dict[str, str]] = field(default_factory=list)
    last_items: List[Dict[str, Any]] = field(default_factory=list)
    last_bot_reply: str = ""
    last_user_text: str = ""
    last_updated_ts: float = field(default_factory=lambda: time.time())
    conversion_score: float = 0.0
    catalog_sent: bool = False
    catalog_sent_at: float = 0.0
    catalog_delivery_mode: str = ""
    last_plan: Dict[str, Any] = field(default_factory=dict)
    profile: Dict[str, Any] = field(default_factory=dict)
    sentiment_score: float = 0.0
    user_message_count: int = 0

    def to_dict(self) -> dict:
        return {
            "tenant": self.tenant,
            "contact_id": self.contact_id,
            "channel": self.channel,
            "needs": self.needs,
            "spin": self.spin,
            "bant": self.bant,
            "asked_questions": self.asked_questions,
            "challenger_cursor": self.challenger_cursor,
            "social_proof_cursor": self.social_proof_cursor,
            "scarcity_cursor": self.scarcity_cursor,
            "reciprocity_cursor": self.reciprocity_cursor,
            "history": self.history[-20:],
            "last_items": self.last_items[-8:],
            "last_bot_reply": self.last_bot_reply,
            "last_user_text": self.last_user_text,
            "last_updated_ts": self.last_updated_ts,
            "conversion_score": self.conversion_score,
            "catalog_sent": self.catalog_sent,
            "catalog_sent_at": self.catalog_sent_at,
            "catalog_delivery_mode": self.catalog_delivery_mode,
            "last_plan": self.last_plan,
            "profile": self.profile,
            "sentiment_score": self.sentiment_score,
            "user_message_count": self.user_message_count,
        }

    @classmethod
    def from_dict(cls, payload: dict) -> "SalesState":
        payload = payload or {}
        tenant = int(payload.get("tenant", 0))
        contact_id = int(payload.get("contact_id", 0))
        obj = cls(tenant=tenant, contact_id=contact_id)
        obj.channel = payload.get("channel", obj.channel)
        obj.needs = payload.get("needs", {}) or {}
        obj.spin = payload.get("spin", obj.spin) or {stage: "pending" for stage in ("s", "p", "i", "n")}
        obj.bant = payload.get("bant", {}) or {}
        obj.asked_questions = payload.get("asked_questions", []) or []
        obj.challenger_cursor = int(payload.get("challenger_cursor", 0))
        obj.social_proof_cursor = int(payload.get("social_proof_cursor", 0))
        obj.scarcity_cursor = int(payload.get("scarcity_cursor", 0))
        obj.reciprocity_cursor = int(payload.get("reciprocity_cursor", 0))
        obj.history = payload.get("history", []) or []
        obj.last_items = payload.get("last_items", []) or []
        obj.last_bot_reply = payload.get("last_bot_reply", "") or ""
        obj.last_user_text = payload.get("last_user_text", "") or ""
        obj.last_updated_ts = float(payload.get("last_updated_ts", time.time()))
        obj.conversion_score = float(payload.get("conversion_score", 0.0))
        obj.catalog_sent = bool(payload.get("catalog_sent", False))
        obj.catalog_sent_at = float(payload.get("catalog_sent_at", 0.0) or 0.0)
        obj.catalog_delivery_mode = payload.get("catalog_delivery_mode", "") or ""
        obj.last_plan = payload.get("last_plan", {}) or {}
        obj.profile = payload.get("profile", {}) or {}
        try:
            obj.sentiment_score = float(payload.get("sentiment_score", 0.0))
        except Exception:
            obj.sentiment_score = 0.0
        obj.user_message_count = int(payload.get("user_message_count", 0))
        return obj

    def append_history(self, role: str, content: str) -> None:
        if not content:
            return
        content = content.strip()
        if not content:
            return
        if self.history and self.history[-1].get("role") == role and self.history[-1].get("content") == content:
            return
        self.history.append({"role": role, "content": content})
        if len(self.history) > 24:
            self.history = self.history[-24:]

    def mark_spin_stage(self, stage: str, status: str) -> None:
        if stage not in self.spin:
            self.spin[stage] = status
        else:
            order = {"pending": 0, "asked": 1, "covered": 2}
            if order.get(status, 0) >= order.get(self.spin.get(stage, "pending"), 0):
                self.spin[stage] = status


@dataclass
class PersonaHints:
    greeting: str = ""
    cta: str = ""
    closing: str = ""
    tone: str = ""
    language: str = ""
    max_questions: Optional[int] = None
    style_short: bool = False
    style_friendly: bool = False

    def wants_short(self) -> bool:
        if self.style_short:
            return True
        tone = (self.tone or "").lower()
        return any(token in tone for token in ("корот", "лакон", "brief", "concise", "short"))

    def wants_friendly(self) -> bool:
        if self.style_friendly:
            return True
        tone = (self.tone or "").lower()
        return any(token in tone for token in ("дружелюб", "тепл", "friendly", "human"))


def _clean_persona_line(line: str) -> str:
    return re.sub(r"^[\-•*\s]+", "", line or "").strip()


def extract_persona_hints(persona: str) -> PersonaHints:
    hints = PersonaHints()
    if not persona:
        return hints

    lines = [_clean_persona_line(line) for line in persona.splitlines()]
    persona_lower = persona.lower()

    for raw in lines:
        if not raw:
            continue
        m = _PERSONA_HINTS_KEY_RE.match(raw)
        if not m:
            continue
        key, value = m.group(1).lower(), m.group(2).strip()
        if key.startswith("greeting") or key.startswith("приветств"):
            hints.greeting = value
        elif key == "cta" or key.startswith("призыв"):
            hints.cta = value
        elif key.startswith("closing") or key.startswith("заверш"):
            hints.closing = value
        elif key.startswith("tone") or key.startswith("тон"):
            hints.tone = value
        elif key.startswith("language") or key.startswith("язык"):
            hints.language = value
        elif "max" in key:
            digits = re.findall(r"\d+", value)
            if digits:
                try:
                    hints.max_questions = int(digits[0])
                except Exception:
                    pass

    if not hints.greeting:
        for raw in lines:
            if not raw or raw.startswith("#"):
                continue
            low = raw.lower()
            if low.startswith(("правила", "техники")):
                continue
            if any(token in low for token in ("привет", "здрав", "меня зовут")):
                hints.greeting = raw
                break
        if not hints.greeting:
            for raw in lines:
                if raw and not raw.startswith(('#', '-', '*')):
                    hints.greeting = raw
                    break

    if not hints.cta:
        m = re.search(r"cta[^\n]*?:\s*(.+)", persona, re.IGNORECASE)
        if m:
            hints.cta = m.group(1).strip()

    if hints.max_questions is None:
        m = re.search(r"≤\s*(\d+)\s*(?:уточн|вопрос)", persona_lower)
        if m:
            try:
                hints.max_questions = int(m.group(1))
            except Exception:
                pass

    if any(token in persona_lower for token in ("коротко", "кратко", "лаконич", "brief", "concise", "short")):
        hints.style_short = True
    if any(token in persona_lower for token in ("дружелюб", "тепл", "friendly", "улыб")):
        hints.style_friendly = True

    return hints


_PERSONA_HINTS_CACHE: Dict[int | None, Tuple[str, PersonaHints]] = {}


def load_persona_hints(tenant: int | None = None) -> PersonaHints:
    persona_text = load_persona(tenant)
    fingerprint = hashlib.sha1(persona_text.encode("utf-8")).hexdigest() if persona_text else ""
    key: int | None
    try:
        key = int(tenant) if tenant is not None else None
    except Exception:
        key = None
    cached = _PERSONA_HINTS_CACHE.get(key)
    if cached and cached[0] == fingerprint:
        return cached[1]
    hints = extract_persona_hints(persona_text)
    _PERSONA_HINTS_CACHE[key] = (fingerprint, hints)
    return hints


def load_sales_state(tenant: int | None, contact_id: int | None) -> SalesState:
    key = _state_key(tenant, contact_id)
    if key in _STATE_CACHE:
        return _STATE_CACHE[key]
    payload = _state_store_read(key)
    if payload:
        state = SalesState.from_dict(payload)
    else:
        state = SalesState(tenant=int(tenant or 0), contact_id=int(contact_id or 0))
    _STATE_CACHE[key] = state
    return state


def save_sales_state(state: SalesState) -> None:
    key = _state_key(state.tenant, state.contact_id)
    payload = state.to_dict()
    _STATE_CACHE[key] = state
    _state_store_write(key, payload)


def reset_sales_state(tenant: int | None, contact_id: int | None) -> None:
    key = _state_key(tenant, contact_id)
    _STATE_CACHE.pop(key, None)
    _with_sync_redis(lambda client: client.delete(key), None)


# --------------------------- хранилище ключей (Redis) ------------------------

def get_tenant_pubkey(tenant: int) -> str:
    return _with_sync_redis(
        lambda client: client.hget(TENANT_PUBKEYS_HASH, str(int(tenant))) or "",
        "",
    )


def set_tenant_pubkey(tenant: int, key: str) -> None:
    key_norm = (key or "").strip().lower()

    def _apply(client: redis_sync.Redis) -> None:
        if key_norm:
            client.hset(TENANT_PUBKEYS_HASH, str(int(tenant)), key_norm)
        else:
            client.hdel(TENANT_PUBKEYS_HASH, str(int(tenant)))

    _with_sync_redis(_apply, None)


# ----------------------------- утилиты HTTP ---------------------------------
def http_json(method: str, url: str, data: dict | None = None, timeout: float = 8.0):
    body = None
    headers = {"Accept": "application/json"}
    if data is not None:
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=body, method=method, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            ctype = resp.headers.get("Content-Type", "")
            raw = resp.read()
            return resp.status, ctype, raw
    except urllib.error.HTTPError as e:
        return e.code, e.headers.get("Content-Type", ""), e.read()
    except Exception as e:
        return 599, "text/plain", str(e).encode("utf-8")


# ----------------------- данные и файлы арендаторов -------------------------
DEFAULT_TENANT_JSON = {
    "passport": {
        "tenant_id": 0,
        "brand": "Мой Бренд",
        "agent_name": "Менеджер",
        "city": "Город",
        "currency": "₽",
        "channel": "WhatsApp",
        "whatsapp_link": "https://wa.me/7XXXXXXXXXX",
    },
    "behavior": {
        "always_full_catalog": True,
        "send_catalog_as_pages": True,
        "max_clarifying_questions": 1,
        "single_cta_per_reply": True,
        "tone": "коротко-дружелюбно",
        "anti_repeat_window": 6,
        "dedupe_catalog_titles": True,
        "allow_filter_commands": True,
        "pdf_one_item_per_page": False,
    },
    "cta": {
        "primary": "Оставьте контакт или удобный канал связи — подготовлю точный расчёт сегодня.",
        "fallback": "Поделитесь, что важно в продукте, и соберу подбор за пару минут.",
        "handoff_wa": "Готов перейти в WhatsApp. Напишите мне — отвечаю быстро.",
    },
    "catalogs": [
        {
            "name": "catalog",
            "path": str(DATA_DIR / "catalog_sample.csv"),
            "type": "csv",
            "delimiter": ",",
            "encoding": "utf-8",
            "fields": {
                "id": "id",
                "title": "name",
                "price": "price",
                "brand": "brand",
                "material": "material",
                "color": "color",
                "stock": "stock",
                "image": "image",
                "url": "url",
                "tags": "tags",
            },
            "ranking": {
                "boost_tags": ["хит", "новинка", "склад", "топ"],
                "boost_stock": 1.0,
                "boost_margin": 0.2,
                "min_stock": 0,
                "min_score": 0,
                "sort": [
                    {"by": "score", "order": "desc"},
                    {"by": "price", "order": "asc"},
                ],
                "filters_default": {"stock": [">", 0]},
            },
            "presentation": {
                "price_format": "{price} {CUR}",
                "line_format": "{title} — {price} {CUR}. Цвет: {color}. Материал: {material}. [{url}]",
                "group_by": "brand",
            },
        }
    ],
    "funnel": {
        "avito_to_wa": {
            "enabled": True,
            "trigger_phrases": ["напишите в whatsapp", "скину в ватсап"],
        }
    },
    "learning": {
        "enabled": True,
        "retriever": "tfidf",
        "top_k": 2,
        "max_tokens": 320,
        "min_chars": 15,
        "track_outcomes": True,
        "auto_vitrine": True,
        "memory_window_dialogs": 50,
        "pinned_items": [],
        "negatives": ["дорого", "долго"],
    },
    "limits": {
        "catalog_page_size": 8,
        "max_pages_per_reply": 5,
        "rate_limit_per_contact_min": 1,
        "send_throttle_ms": 250,
    },
    "integrations": {
        "pdf_catalog_url": "",
        "crm_webhook": "",
        "analytics_pixel": "",
        "ga_id": "",
        "uploaded_catalog": "",
    },
}

DEFAULT_PERSONA_MD = """{AGENT_NAME} из {BRAND}, {CITY}. Канал: {CHANNEL}. Валюта: {CURRENCY}.\nПравила:\n- Говори живо и предметно: 2–3 коротких абзаца или списки.\n- Показывай конкретные товары с их выгодами, держи фокус на продаже.\n- Максимум один уточняющий вопрос в ответе.\n- Один понятный CTA в финале, без длинных сценариев.\n- Если клиент просит каталог — предложи лучшие позиции и ссылку.\n\nТактика: активное слушание, выгоды «что получите», уместное соцдоказательство и мягкая допродажа (≤1 за ответ). Антидублирование: не повторяй вступление и одинаковые товары подряд.\n"""


def tenant_dir(tenant: int) -> pathlib.Path:
    return TENANTS_DIR / str(int(tenant))


def ensure_tenant_files(tenant: int) -> pathlib.Path:
    td = tenant_dir(tenant)
    td.mkdir(parents=True, exist_ok=True)
    tj = td / "tenant.json"
    pm = td / "persona.md"

    if not tj.exists() or tj.stat().st_size == 0:
        cfg = json.loads(json.dumps(DEFAULT_TENANT_JSON, ensure_ascii=False))
        cfg.setdefault("passport", {})["tenant_id"] = int(tenant)
        with open(tj, "w", encoding="utf-8") as fh:
            json.dump(cfg, fh, ensure_ascii=False, indent=2)

    if not pm.exists() or pm.stat().st_size == 0:
        with open(pm, "w", encoding="utf-8") as fh:
            fh.write(DEFAULT_PERSONA_MD)

    return td


def read_tenant_config(tenant: int) -> dict:
    ensure_tenant_files(tenant)
    path = tenant_dir(tenant) / "tenant.json"
    try:
        mtime = path.stat().st_mtime
        cached = _TENANT_CONFIG_CACHE.get(int(tenant))
        if cached and cached[0] == mtime:
            return cached[1]
    except Exception:
        mtime = 0.0
    with open(path, "r", encoding="utf-8") as fh:
        data = json.load(fh)
    try:
        _TENANT_CONFIG_CACHE[int(tenant)] = (mtime, data)
    except Exception:
        pass
    return data


def write_tenant_config(tenant: int, cfg: dict) -> None:
    ensure_tenant_files(tenant)
    path = tenant_dir(tenant) / "tenant.json"
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(cfg, fh, ensure_ascii=False, indent=2)
    try:
        mtime = path.stat().st_mtime
        _TENANT_CONFIG_CACHE[int(tenant)] = (mtime, cfg)
    except Exception:
        _TENANT_CONFIG_CACHE.pop(int(tenant), None)


def _persist_pdf_index_metadata(
    tenant: int,
    source_key: str,
    rel_index_path: str,
    index_meta: Dict[str, Any],
) -> None:
    try:
        cfg = read_tenant_config(tenant)
    except Exception:
        return

    catalogs = cfg.get("catalogs")
    catalogs = catalogs if isinstance(catalogs, list) else []
    source_candidates = {source_key.strip()}
    source_candidates.add(index_meta.get("source_path", ""))
    resolved: set[str] = set()
    for token in list(source_candidates):
        if not token:
            continue
        resolved.add(token)
        try:
            abs_path = str((tenant_dir(tenant) / token).resolve())
            resolved.add(abs_path)
        except Exception:
            pass
    resolved = {value for value in resolved if value}

    for entry in catalogs:
        entry_path = str(entry.get("path") or "").strip()
        if not entry_path:
            continue
        candidate_set = {entry_path}
        try:
            candidate_set.add(str((tenant_dir(tenant) / entry_path).resolve()))
        except Exception:
            pass
        if candidate_set & resolved:
            entry["index_path"] = rel_index_path
            entry["indexed_at"] = index_meta.get("generated_at")
            entry["chunk_count"] = index_meta.get("chunk_count")
            entry["sha1"] = index_meta.get("sha1")
            break

    integrations = cfg.setdefault("integrations", {})
    uploaded = integrations.get("uploaded_catalog")
    if isinstance(uploaded, dict) and (uploaded.get("path") in resolved):
        uploaded["index"] = {
            "path": rel_index_path,
            "generated_at": index_meta.get("generated_at"),
            "chunks": index_meta.get("chunk_count"),
            "pages": index_meta.get("page_count"),
            "sha1": index_meta.get("sha1"),
        }

    try:
        write_tenant_config(tenant, cfg)
    except Exception:
        pass


def read_persona(tenant: int) -> str:
    ensure_tenant_files(tenant)
    path = tenant_dir(tenant) / "persona.md"
    try:
        mtime = path.stat().st_mtime
        cached = _TENANT_PERSONA_CACHE.get(int(tenant))
        if cached and cached[0] == mtime:
            return cached[1]
    except Exception:
        mtime = 0.0
    with open(path, "r", encoding="utf-8") as fh:
        text = fh.read()
    try:
        _TENANT_PERSONA_CACHE[int(tenant)] = (mtime, text)
    except Exception:
        pass
    return text


def write_persona(tenant: int, text: str) -> None:
    ensure_tenant_files(tenant)
    path = tenant_dir(tenant) / "persona.md"
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(text or "")
    try:
        mtime = path.stat().st_mtime
        _TENANT_PERSONA_CACHE[int(tenant)] = (mtime, text or "")
    except Exception:
        _TENANT_PERSONA_CACHE.pop(int(tenant), None)


def load_tenant(tenant: int) -> dict:
    try:
        return read_tenant_config(tenant)
    except Exception:
        cfg = json.loads(json.dumps(DEFAULT_TENANT_JSON, ensure_ascii=False))
        cfg.setdefault("passport", {})["tenant_id"] = int(tenant)
        return cfg


def _branding_for_tenant(tenant: int | None = None) -> Dict[str, str]:
    passport: Dict[str, Any] = {}
    if tenant is not None:
        try:
            cfg = read_tenant_config(tenant)
            passport = cfg.get("passport", {}) if isinstance(cfg, dict) else {}
        except Exception:
            passport = {}
    return {
        "AGENT_NAME": (passport.get("agent_name") or settings.AGENT_NAME).strip() or settings.AGENT_NAME,
        "BRAND": (passport.get("brand") or settings.BRAND_NAME).strip() or settings.BRAND_NAME,
        "BRAND_NAME": (passport.get("brand") or settings.BRAND_NAME).strip() or settings.BRAND_NAME,
        "WHATSAPP_LINK": (passport.get("whatsapp_link") or settings.WHATSAPP_LINK).strip() or settings.WHATSAPP_LINK,
        "CITY": (passport.get("city") or settings.CITY).strip() or settings.CITY,
        "CHANNEL": (passport.get("channel") or "WhatsApp").strip() or "WhatsApp",
        "CURRENCY": (passport.get("currency") or "₽").strip() or "₽",
    }

# ------------------------------ промпты --------------------------------------
PERSONA_MD = """# Персона «Универсальный продавец»

- Роль: онлайн-менеджер (не раскрывайся как ИИ).
- Представление: «Здравствуйте! Меня зовут {AGENT_NAME}, {BRAND}.»
- Тон: кратко, естественно, без лишних вопросов.
- Цель: дать пользу и один понятный CTA.
- Правила:
  - Всегда прикладывай ВЕСЬ КАТАЛОГ. Если не помещается — отправляй постранично подряд, без ожидания «продолжать?».
  - ≤1 уточнение на ход, только если повышает точность следующего шага.
  - Используй техники продаж: AIDA, PAS, лёгкий SPIN (≤1 вопрос), LAER, якорение (лучше/оптимально/бюджетно), соцдоказательство, мягкий up/cross-sell (≤1/ход), микрокоммит.
- Канал: {CHANNEL}. Если Avito — в конце можно мягко предложить перейти в WhatsApp. Если WhatsApp — работай по делу.
- Локаль: {CITY}, валюта: {CURRENCY}.
"""

RULES_YAML = """
take_control: true
cta:
  - "Сформирую подбор. Назначим замер на завтра: утро/день/вечер?"
  - "Готов оформить заказ сегодня. Подходит?"
  - "Забронирую цену на сутки. Идём дальше?"
"""


# ---------------------------- персонализация ---------------------------------
def load_persona(tenant: int | None = None) -> str:
    """Возвращает persona.md с подстановкой брендинга."""
    if tenant is not None:
        try:
            persona = read_persona(tenant)
            if not persona.strip():
                persona = DEFAULT_PERSONA_MD
        except Exception:
            persona = DEFAULT_PERSONA_MD
    else:
        try:
            with open(settings.PERSONA_MD, "r", encoding="utf-8") as fh:
                persona = fh.read()
        except Exception:
            persona = PERSONA_MD

    tokens = _branding_for_tenant(tenant)
    tokens.setdefault("WHATSAPP_LINK", settings.WHATSAPP_LINK)
    for key, value in tokens.items():
        persona = persona.replace(f"{{{key}}}", value)
    return persona


# ---------------------- простая rule-based логика ----------------------------
CATALOG_CSV = DATA_DIR / "catalog_sample.csv"


def _canonicalize_field_name(name: str) -> str:
    return _FIELD_CLEAN_RE.sub("", (name or "").lower())


_FIELD_SYNONYMS: Dict[str, List[str]] = {
    "title": [
        "title",
        "name",
        "product",
        "productname",
        "item",
        "itemname",
        "goods",
        "model",
        "модель",
        "товар",
        "наименование",
        "название",
        "позиция",
        "описание",
        "характеристика",
    ],
    "price": [
        "price",
        "cost",
        "стоимость",
        "цена",
        "ценаактуальная",
        "ценапродажи",
        "ценаруб",
        "ценазасистему",
        "ценазасчет",
        "ценаскидкой",
        "ценабезскидки",
        "ценазам2",
        "ценазамкв",
        "ценазаметры",
        "ценарозничная",
        "ценазапозицию",
    ],
    "sku": [
        "sku",
        "код",
        "кодтовара",
        "артикул",
        "арт",
        "код1с",
        "идентификатор",
        "id",
        "article",
    ],
    "url": [
        "url",
        "link",
        "urlтовара",
        "ссылка",
        "hyperlink",
        "страница",
    ],
    "brand": [
        "brand",
        "бренд",
        "марка",
        "производитель",
        "manufacturer",
    ],
    "stock": [
        "stock",
        "наличие",
        "остаток",
        "остатки",
        "количество",
        "qty",
        "quantity",
        "available",
    ],
    "image": [
        "image",
        "photo",
        "img",
        "picture",
        "изображение",
        "картинка",
        "фото",
        "фотография",
    ],
    "description": [
        "description",
        "описание",
        "details",
        "характеристики",
        "features",
        "comment",
    ],
}


_FIELD_TOKEN_MAP: Dict[str, List[str]] = {
    key: sorted({_canonicalize_field_name(token) for token in tokens if token}, key=len, reverse=True)
    for key, tokens in _FIELD_SYNONYMS.items()
}


def _prepare_field_mapping(meta: Dict[str, Any], items: List[Dict[str, Any]]) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    if not items:
        return mapping

    sample_cols = list(items[0].keys())
    sample_canon = {col: _canonicalize_field_name(col) for col in sample_cols}

    meta_fields = meta.get("fields") if isinstance(meta, dict) else None
    if isinstance(meta_fields, dict):
        for canonical, source in meta_fields.items():
            if not isinstance(canonical, str) or not isinstance(source, str):
                continue
            if source in sample_cols:
                mapping[canonical.strip().lower()] = source

    used_sources = set(mapping.values())

    def _find_column(tokens: List[str], preferred: List[str] | None = None) -> str | None:
        preferred = preferred or []
        for col in sample_cols:
            if col in used_sources:
                continue
            canon = sample_canon.get(col) or ""
            for p in preferred:
                if p and (canon == p or canon.startswith(p) or p in canon):
                    used_sources.add(col)
                    return col
            for token in tokens:
                if not token:
                    continue
                if canon == token or canon.startswith(token) or token in canon:
                    used_sources.add(col)
                    return col
        return None

    for field, tokens in _FIELD_TOKEN_MAP.items():
        if field in mapping:
            continue
        column = _find_column(tokens)
        if column:
            mapping[field] = column

    # Extra heuristics if price or title still missing
    if "price" not in mapping:
        numeric_candidates: List[str] = []
        for col in sample_cols:
            if col in used_sources:
                continue
            canon = sample_canon.get(col) or ""
            if any(token in canon for token in ("цен", "price", "cost", "стоим", "руб", "uah", "usd", "eur")):
                mapping["price"] = col
                used_sources.add(col)
                break
            # Look at data if no obvious hints
            values = [str((row.get(col) or "")).strip() for row in items[:5]]
            digits = [re.sub(r"\D", "", val) for val in values if val]
            if any(len(d) >= 4 for d in digits):
                numeric_candidates.append(col)
        if "price" not in mapping and numeric_candidates:
            mapping["price"] = numeric_candidates[0]
            used_sources.add(numeric_candidates[0])

    if "title" not in mapping:
        for col in sample_cols:
            if col in used_sources:
                continue
            canon = sample_canon.get(col) or ""
            if any(token in canon for token in ("name", "товар", "пози", "model", "тип", "item", "наимен")):
                mapping["title"] = col
                used_sources.add(col)
                break

    return mapping


def _has_price_digits(value: Any) -> bool:
    text = str(value or "")
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 4:
        return True
    lowered = text.lower()
    if len(digits) >= 3 and any(tok in lowered for tok in ("руб", "uah", "eur", "usd", "$", "€", "₽")):
        return True
    try:
        # Attempt to parse decimal values like "99.5"
        normalized = text.replace(" ", "").replace(",", ".")
        float(normalized)
        return True
    except Exception:
        return False


def _normalize_catalog_item(record: Dict[str, Any], mapping: Dict[str, str]) -> Dict[str, Any]:
    normalized = dict(record)
    for target, source in mapping.items():
        if not source:
            continue
        if target in normalized and str(normalized[target]).strip():
            continue
        value = record.get(source)
        if value is None:
            continue
        normalized[target] = value

    def _ensure_title() -> None:
        title_candidates = [normalized.get("title"), normalized.get("name")]
        for candidate in title_candidates:
            if candidate and str(candidate).strip():
                normalized.setdefault("name", candidate)
                if not str(normalized.get("title") or "").strip():
                    normalized["title"] = candidate
                return

        for key, value in record.items():
            if key in {"price", mapping.get("price", "")}:  # avoid grabbing price column
                continue
            text = str(value or "").strip()
            if len(text) >= 3 and not text.isdigit():
                normalized.setdefault("title", text)
                normalized.setdefault("name", text)
                return

    def _ensure_price() -> None:
        current = normalized.get("price")
        if current and _has_price_digits(current):
            return

        if current and isinstance(current, str) and current.strip():
            digits = re.sub(r"\D", "", current)
            if digits and len(digits) >= 4:
                return

        preferred_columns = [mapping.get("price")]
        for key, value in record.items():
            if key in preferred_columns:
                preferred_columns.append(key)
        seen = set(filter(None, preferred_columns))
        for key in preferred_columns:
            if not key:
                continue
            text = str(record.get(key) or "").strip()
            if _has_price_digits(text):
                normalized["price"] = text
                return

        for key, value in record.items():
            if key in seen:
                continue
            text = str(value or "").strip()
            if _has_price_digits(text):
                normalized["price"] = text
                return

    _ensure_title()
    _ensure_price()
    return normalized


def _normalize_catalog_items(items: List[Dict[str, Any]], meta: Dict[str, Any] | Any) -> List[Dict[str, Any]]:
    if not items:
        return items
    meta_dict = meta if isinstance(meta, dict) else {}
    mapping = _prepare_field_mapping(meta_dict, items)
    if not mapping:
        # Even without explicit mapping try to enrich titles and prices
        return [_normalize_catalog_item(record, {}) for record in items]
    return [_normalize_catalog_item(record, mapping) for record in items]


def _read_catalog(tenant: int | None = None) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    candidates: List[tuple[pathlib.Path, Dict[str, Any]]] = []
    has_custom_catalogs = False

    if tenant is not None:
        try:
            cfg = load_tenant(tenant)
            catalogs = cfg.get("catalogs") or []
            if isinstance(catalogs, list):
                for entry in catalogs:
                    if not isinstance(entry, dict):
                        continue
                    raw_path = entry.get("path")
                    if not raw_path:
                        continue
                    path = pathlib.Path(str(raw_path))
                    if not path.is_absolute():
                        path = tenant_dir(tenant) / path
                    candidates.append((path, entry))
                    has_custom_catalogs = True
        except Exception:
            pass

    if not candidates:
        candidates.append((CATALOG_CSV, {"delimiter": ",", "encoding": "utf-8"}))

    # mtime/size-based cache key to avoid repeated heavy parsing
    key_fps: List[Tuple[str, float, int]] = []
    try:
        for pth, meta in candidates:
            meta = meta if isinstance(meta, dict) else {}
            meta_type = (meta.get("type") or pth.suffix.lstrip(".")).lower()
            stat_target = pth
            if pth.suffix.lower() == ".pdf" or meta_type == "pdf":
                idx_val = meta.get("index_path")
                if idx_val and tenant is not None:
                    cand = pathlib.Path(str(idx_val))
                    if not cand.is_absolute():
                        cand = tenant_dir(int(tenant)) / cand
                    if cand.exists():
                        stat_target = cand
            if stat_target.exists():
                st = stat_target.stat()
                key_fps.append((str(stat_target.resolve()), st.st_mtime, int(getattr(st, 'st_size', 0) or 0)))
    except Exception:
        key_fps = []
    cache_key: Tuple[Optional[int], Tuple[Tuple[str, float, int], ...]] = (
        (int(tenant) if tenant is not None else None), tuple(sorted(key_fps))
    )
    cached = _CATALOG_CACHE.get(cache_key)
    if cached:
        return cached

    for path, meta in candidates:
        try:
            if not path.exists():
                continue

            meta = meta if isinstance(meta, dict) else {}
            encoding = meta.get("encoding", "utf-8")
            meta_type = (meta.get("type") or path.suffix.lstrip(".")).lower()

            if path.suffix.lower() in {".xlsx", ".xls"} or meta_type == "excel":
                if load_workbook is None:
                    continue
                wb = load_workbook(filename=str(path), read_only=True, data_only=True)
                ws = wb.active
                headers = []
                for cell in next(ws.iter_rows(min_row=1, max_row=1), []):
                    headers.append(str(cell.value or "").strip())
                if not headers:
                    wb.close()
                    continue
                collected: List[Dict[str, Any]] = []
                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                    record = {}
                    for h, val in zip(headers, row):
                        record[str(h or "").strip()] = str(val).strip() if val is not None else ""
                    if any(record.values()):
                        collected.append(record)
                    if idx >= 500:
                        break
                wb.close()
                if collected:
                    items = _normalize_catalog_items(collected, meta)
                    if items:
                        break
                continue

            if path.suffix.lower() == ".pdf" or meta_type == "pdf":
                if not isinstance(meta, dict):
                    continue
                raw_source_key = str(meta.get("path") or path)
                index_path_value = meta.get("index_path")
                index_path_obj: pathlib.Path | None = None

                if index_path_value:
                    candidate = pathlib.Path(str(index_path_value))
                    if not candidate.is_absolute() and tenant is not None:
                        candidate = tenant_dir(int(tenant)) / candidate
                    if candidate.exists():
                        index_path_obj = candidate

                if index_path_obj is None and tenant is not None:
                    try:
                        from catalog_index import build_pdf_index

                        try:
                            rel_source = str(path.relative_to(tenant_dir(int(tenant))))
                        except Exception:
                            rel_source = str(meta.get("path") or path.name)

                        index_dir = tenant_dir(int(tenant)) / "indexes"
                        built_index = build_pdf_index(
                            path,
                            output_dir=index_dir,
                            source_relpath=rel_source,
                            original_name=path.name,
                        )

                        index_path_obj = built_index.index_path
                        try:
                            rel_index_path = str(index_path_obj.relative_to(tenant_dir(int(tenant))))
                        except Exception:
                            rel_index_path = str(index_path_obj)

                        meta["index_path"] = rel_index_path
                        meta["indexed_at"] = built_index.generated_at
                        meta["chunk_count"] = built_index.chunk_count
                        meta["sha1"] = built_index.sha1

                        _persist_pdf_index_metadata(
                            int(tenant),
                            raw_source_key,
                            rel_index_path,
                            {
                                "generated_at": built_index.generated_at,
                                "chunk_count": built_index.chunk_count,
                                "sha1": built_index.sha1,
                                "page_count": built_index.page_count,
                                "source_path": built_index.source_path,
                            },
                        )
                    except Exception:
                        continue

                if not index_path_obj:
                    continue

                try:
                    from catalog_index import load_index, index_to_catalog_items

                    index = load_index(index_path_obj)
                    indexed_items = index_to_catalog_items(index)
                    if indexed_items:
                        items = indexed_items
                        break
                except Exception:
                    continue
                continue

            # CSV и подобные
            delimiter = meta.get("delimiter")
            enc_candidates: List[str] = []
            if isinstance(encoding, str) and encoding:
                enc_candidates.append(encoding)
            # Try declared encoding first, then common Russian CSV fallbacks.
            for fallback in ("utf-8", "utf-8-sig", "cp1251", "windows-1251", "koi8-r"):
                if fallback not in enc_candidates:
                    enc_candidates.append(fallback)

            used_items: List[Dict[str, Any]] = []
            for enc in enc_candidates or ["utf-8"]:
                try:
                    with open(path, "r", encoding=enc, newline="") as fh:
                        local_delimiter = delimiter
                        if not local_delimiter:
                            sample = fh.read(2048)
                            fh.seek(0)
                            try:
                                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                                local_delimiter = dialect.delimiter
                            except Exception:
                                local_delimiter = ","

                        reader = csv.reader(fh, delimiter=local_delimiter or ",")
                        header: List[str] = []
                        for raw_header in reader:
                            if not raw_header or not any((cell or "").strip() for cell in raw_header):
                                continue
                            header = raw_header
                            break
                        if not header:
                            continue

                        normalized: List[str] = []
                        seen_headers: Dict[str, int] = {}
                        for idx_h, cell in enumerate(header):
                            name = (cell or "").strip().lstrip("\ufeff")
                            if not name:
                                name = f"column_{idx_h + 1}"
                            if name in seen_headers:
                                seen_headers[name] += 1
                                name = f"{name}_{seen_headers[name]}"
                            else:
                                seen_headers[name] = 0
                            normalized.append(name)

                        columns = normalized[:]
                        local_items: List[Dict[str, Any]] = []
                        for row in reader:
                            if not row or not any((val.strip() if isinstance(val, str) else str(val or "").strip()) for val in row):
                                continue
                            while len(columns) < len(row):
                                columns.append(f"column_{len(columns) + 1}")
                            record: Dict[str, Any] = {}
                            for idx_col, value in enumerate(row):
                                key = columns[idx_col]
                                if isinstance(value, str):
                                    clean = value.strip()
                                else:
                                    clean = str(value or "").strip()
                                record[key] = clean
                            if any(record.values()):
                                local_items.append(record)
                            if len(local_items) >= 500:
                                break
                    if local_items:
                        used_items = local_items
                        break
                except UnicodeDecodeError:
                    continue
            if used_items:
                items = _normalize_catalog_items(used_items, meta)
                break
        except Exception:
            continue

    if not items:
        if has_custom_catalogs:
            return []
        items = [
            {
                "sku": "SKU-101",
                "type": "гаджет",
                "title": "Умная колонка Echo Mini",
                "price": "5990",
                "color": "графит",
                "brand": "Soundify",
            },
            {
                "sku": "SKU-204",
                "type": "освещение",
                "title": "Лампа Loft Aura",
                "price": "8900",
                "color": "латунь",
                "brand": "Loftly",
            },
            {
                "sku": "SKU-350",
                "type": "офис",
                "title": "Кресло Support Pro",
                "price": "21900",
                "color": "чёрный",
                "brand": "Ergo",
            },
            {
                "sku": "SKU-480",
                "type": "кухня",
                "title": "Набор ножей ChefLine",
                "price": "12900",
                "color": "стальной",
                "brand": "ChefLine",
            },
        ]

    # Precompute normalized search text for matching speed
    try:
        for it in items:
            if isinstance(it, dict):
                parts: List[str] = []
                for k in (
                    "title","name","sku","id","brand","collection","category","series","model","color","material","decor","finish","tags","description","notes","features",
                ):
                    v = it.get(k)
                    if isinstance(v, (list, tuple, set)):
                        parts.extend(str(x) for x in v if x)
                    elif v:
                        parts.append(str(v))
                it["_search_text"] = (" ".join(parts)).casefold().replace("ё", "е")
    except Exception:
        pass

    # Store in cache
    try:
        _CATALOG_CACHE[cache_key] = items
    except Exception:
        pass

    return items


def read_all_catalog(cfg: Optional[Dict[str, Any]] = None, tenant: int | None = None) -> List[Dict[str, Any]]:
    """Возвращает список позиций каталога для арендатора."""
    tenant_id: Optional[int] = None
    if tenant is not None:
        try:
            tenant_id = int(tenant)
        except Exception:
            tenant_id = None
    elif isinstance(cfg, dict):
        passport = cfg.get("passport") if isinstance(cfg.get("passport"), dict) else {}
        raw_id = passport.get("tenant_id")
        try:
            tenant_id = int(raw_id) if raw_id is not None else None
        except Exception:
            tenant_id = None
    return _read_catalog(tenant_id)


def paginate_catalog_text(
    items: List[Dict[str, Any]],
    cfg: Optional[Dict[str, Any]] = None,
    page_size: int = 10,
) -> List[str]:
    """Формирует текстовые страницы каталога."""
    if not items:
        return []

    try:
        page_size = int(page_size)
    except Exception:
        page_size = 10
    if page_size <= 0:
        page_size = 10

    currency = "₽"
    if isinstance(cfg, dict):
        passport = cfg.get("passport") if isinstance(cfg.get("passport"), dict) else {}
        cur = passport.get("currency")
        if cur:
            currency = str(cur)

    formatted_lines = format_items_for_prompt(items, currency).splitlines()
    pages: List[str] = []
    for idx in range(0, len(formatted_lines), page_size):
        chunk = formatted_lines[idx : idx + page_size]
        if not chunk:
            continue
        page_no = idx // page_size + 1
        header = f"Каталог, страница {page_no}:"
        pages.append("\n".join([header, *chunk]))
    return pages

NEEDS_STOPWORDS = {
    "нужно",
    "нужна",
    "нужен",
    "нужны",
    "ищу",
    "ищем",
    "ищет",
    "ищете",
    "ищите",
    "хочу",
    "интересует",
    "интересуют",
    "каталог",
    "про",
    "для",
    "подбор",
    "бюджет",
    "стоимость",
    "цена",
    "цену",
    "ценник",
    "до",
    "подберите",
    "подбер",
    "посоветуйте",
    "подскажите",
    "расскажите",
    "рассмотрите",
    "еще",
    "ещё",
    "можно",
    "пожалуйста",
    "дайте",
    "заказ",
    "добрый",
    "вечер",
    "день",
    "привет",
    "меня",
    "интересуют",
    "надо",
    "пока",
    "под",
    "есть",
    "или",
    "и",
    "в",
    "на",
    "с",
    "как",
    "что",
    "так",
    "же",
    "подбор",
}

COLOR_STEMS = {
    "бел": "белый",
    "черн": "чёрный",
    "чёр": "чёрный",
    "чер": "чёрный",
    "венг": "венге",
    "дуб": "дуб",
    "сер": "серый",
    "корич": "коричневый",
    "красн": "красный",
    "син": "синий",
    "голуб": "голубой",
    "зел": "зелёный",
    "зол": "золотой",
    "сталь": "стальной",
    "беж": "бежевый",
}

SIZE_PATTERN = re.compile(r"(?P<value>\d{2,4})(?:\s|\-)?(?P<unit>см|mm|мм|cm|м|kg|кг|g|гр|ml|мл|l|л)", re.IGNORECASE)


def _extract_budget(text: str) -> Optional[int]:
    if not text:
        return None
    lowered = text.lower()
    candidates: List[int] = []
    for match in re.finditer(r"\d+[\d\s]*", lowered):
        raw_number = match.group(0)
        digits = re.sub(r"\D", "", raw_number)
        if not digits:
            continue
        try:
            value = int(digits)
        except Exception:
            continue

        suffix = lowered[match.end() : match.end() + 4]
        prefix = lowered[max(0, match.start() - 12) : match.start()]

        def _has_token(container: str, tokens: Tuple[str, ...]) -> bool:
            return any(token in container for token in tokens)

        thousand_tokens = ("k", "к", "тыс", "т.", "т ", "тысяч")
        million_tokens = ("млн", "mln")
        currency_tokens = ("₽", "р", "rub", "руб", "eur", "€", "usd", "$")
        context_tokens = ("цен", "стоим", "бюдж", "до", "≈", "~", "max", "за ", "по ")

        if _has_token(suffix, thousand_tokens):
            value *= 1000
        elif _has_token(suffix, million_tokens):
            value *= 1_000_000

        has_currency = _has_token(suffix, currency_tokens) or _has_token(prefix, currency_tokens)
        has_context = _has_token(prefix, context_tokens)

        if value < 100:
            continue
        if not has_currency and not has_context and value < 1000:
            continue

        candidates.append(value)

    if not candidates:
        return None
    return max(candidates)


def infer_user_needs(text: str) -> Dict[str, Any]:
    raw = text or ""
    lowered = raw.lower()
    needs: Dict[str, Any] = {}

    tokens = _tokenize_query(raw)
    keywords = [tok for tok in tokens if tok and tok not in NEEDS_STOPWORDS and not tok.isdigit()]
    if keywords:
        needs["keywords"] = keywords[:6]
        needs["type"] = keywords[0]
        needs["focus"] = " ".join(keywords[:3])

    size_match = SIZE_PATTERN.search(lowered)
    if size_match:
        value = size_match.group("value")
        unit = size_match.group("unit").lower()
        normalized_unit = {
            "mm": "мм",
            "cm": "см",
            "m": "м",
            "kg": "кг",
            "g": "г",
            "gr": "г",
            "l": "л",
        }.get(unit, unit)
        needs["size"] = f"{value} {normalized_unit}"
        if normalized_unit in {"см", "mm", "мм"}:
            needs["width"] = value

    budget = _extract_budget(lowered)
    if budget:
        needs["budget_max"] = budget

    for stem, title in COLOR_STEMS.items():
        if stem in lowered:
            needs["color"] = title
            break

    return needs


def _value_matches(item: Dict[str, Any], fields: Tuple[str, ...], needle: str) -> bool:
    for field in fields:
        val = item.get(field)
        if not val:
            continue
        if isinstance(val, (list, tuple, set)):
            texts = [str(v) for v in val if v]
        else:
            texts = [str(val)]
        for text in texts:
            if needle in _normalize_text(text):
                return True
    return False


def _score(item: Dict[str, Any], needs: Dict[str, Any]) -> float:
    s = 0.0
    haystack_text = _normalize_text(_collect_item_text(item))

    primary = needs.get("type")
    if primary:
        needle = _normalize_text(primary)
        if needle and (
            _value_matches(item, ("type", "category", "segment", "group"), needle)
            or needle in haystack_text
        ):
            s += 3.0

    keywords = needs.get("keywords") or []
    if keywords:
        for kw in keywords[:3]:
            needle = _normalize_text(kw)
            if needle and needle in haystack_text:
                s += 1.0

    size = needs.get("size") or needs.get("width")
    if size:
        size_str = _normalize_text(str(size))
        if _value_matches(item, ("size", "width", "dimensions", "length", "height", "depth"), size_str):
            s += 1.5

    color = needs.get("color")
    if color:
        color_token = _normalize_text(color)
        if _value_matches(item, ("color", "finish", "shade", "title", "name", "tags"), color_token):
            s += 0.8

    budget = needs.get("budget_max")
    if budget:
        try:
            price = int(re.sub(r"\D", "", str(item.get("price") or "0")))
            if price and price <= int(budget):
                s += 1.5
        except Exception:
            pass

    return s

_WORD_TOKEN_RE = re.compile(r"[0-9a-zа-яё]+", re.IGNORECASE)


def _normalize_text(value: Any) -> str:
    text = str(value or "")
    return text.casefold().replace("ё", "е")


def _collect_item_text(item: Dict[str, Any]) -> str:
    cached = item.get("_search_text")
    if isinstance(cached, str) and cached:
        return cached
    parts: List[str] = []
    for key in (
        "title",
        "name",
        "sku",
        "id",
        "brand",
        "collection",
        "category",
        "series",
        "model",
        "color",
        "material",
        "decor",
        "finish",
        "tags",
        "description",
        "notes",
        "features",
    ):
        if key not in item:
            continue
        value = item.get(key)
        if isinstance(value, (list, tuple, set)):
            parts.extend(str(v) for v in value if v)
        elif value:
            parts.append(str(value))
    return " ".join(parts)


def _tokenize_query(text: str | None) -> List[str]:
    if not text:
        return []
    cleaned = _normalize_text(text)
    tokens: List[str] = []
    for raw in _WORD_TOKEN_RE.findall(cleaned):
        token = raw.strip()
        if not token:
            continue
        if token.isdigit():
            tokens.append(token)
            continue
        if len(token) >= 3:
            tokens.append(token)
    return tokens[:12]


def _tag_boost(item: Dict[str, Any]) -> float:
    tags = item.get("tags")
    if not tags:
        return 0.0
    if isinstance(tags, str):
        normalized = _normalize_text(tags)
        tags_iterable = [normalized]
    else:
        tags_iterable = [_normalize_text(tag) for tag in tags if tag]
    bonus = 0.0
    for tag in tags_iterable:
        if "хит" in tag:
            bonus += 0.4
        if "новин" in tag:
            bonus += 0.2
        if "склад" in tag:
            bonus += 0.1
    return bonus


def _text_match_score(item: Dict[str, Any], tokens: List[str]) -> float:
    if not tokens:
        return 0.0
    haystack = _normalize_text(_collect_item_text(item))
    if not haystack:
        return 0.0
    hay_tokens = set(_WORD_TOKEN_RE.findall(haystack))
    score = 0.0
    for token in tokens:
        if not token:
            continue
        if token in hay_tokens:
            score += 2.5
            continue
        if token.isdigit() and token in haystack:
            score += 1.5
            continue
        if len(token) >= 4:
            prefix = token[:4]
            if prefix in haystack:
                score += 0.75
                continue
    return score


def _legacy_rank_catalog(
    items: List[Dict[str, Any]],
    needs: Dict[str, Any],
    limit: int,
    query: str | None,
) -> List[Dict[str, Any]]:
    query_tokens = _tokenize_query(query)

    def _total_score(item: Dict[str, Any]) -> float:
        base = _score(item, needs)
        matched = _text_match_score(item, query_tokens)
        tag_bonus = _tag_boost(item)
        return base + matched + tag_bonus

    scored = sorted(items, key=_total_score, reverse=True)
    if limit <= 0:
        return scored
    return scored[:limit]


def search_catalog(
    needs: Dict[str, Any],
    limit: int = 5,
    tenant: int | None = None,
    query: str | None = None,
) -> List[Dict[str, Any]]:
    needs = needs or {}
    items = _read_catalog(tenant)
    if not items:
        items = _read_catalog(None)

    advanced: List[Dict[str, Any]] = []
    if catalog_retriever and items:
        try:
            advanced = catalog_retriever.retrieve_context(
                items=items,
                needs=needs,
                query=query or "",
                tenant=tenant,
                limit=limit,
            )
        except Exception as exc:
            logger.exception("catalog retriever failed", exc_info=exc)

    if advanced:
        if limit <= 0:
            return advanced
        return advanced[:limit]

    return _legacy_rank_catalog(items, needs, limit, query)

def format_items_for_prompt(items: List[Dict[str, Any]], currency: str = "₽") -> str:
    if not items:
        return "— подходящих позиций не найдено."
    out = []
    for idx, it in enumerate(items, start=1):
        title = (
            it.get("title")
            or it.get("name")
            or it.get("sku")
            or it.get("id")
            or f"Позиция {idx}"
        )
        raw_price = str(it.get("price") or "").strip()
        digits = re.sub(r"\D", "", raw_price)
        if digits:
            try:
                price_fmt = f"{int(digits):,}".replace(",", " ")
            except Exception:
                price_fmt = raw_price
        else:
            price_fmt = raw_price or "цена по запросу"

        details: List[str] = []
        if it.get("brand"):
            details.append(str(it.get("brand")).strip())
        if it.get("width"):
            details.append(f"{it['width']} см")
        if it.get("color"):
            details.append(str(it.get("color")).strip())
        stock = it.get("stock")
        if stock is not None and str(stock).strip():
            try:
                stock_val = int(str(stock).strip())
                if stock_val > 0:
                    details.append("в наличии")
            except Exception:
                details.append(str(stock))
        url = (it.get("url") or "").strip()
        meta = f" ({', '.join(details)})" if details else ""
        line = f"{idx}. {title} — {price_fmt} {currency}{meta}"
        rag_score = it.get("_rag_score")
        if isinstance(rag_score, (int, float)) and rag_score > 0:
            line += f" (релевантность {rag_score:.2f})"
        if url:
            line += f" · {url}"
        excerpt = str(it.get("_match_excerpt") or "").strip()
        if excerpt:
            line = f"{line}\n   ↳ {excerpt}"
        out.append(line)
    return "\n".join(out)

def format_needs_for_prompt(needs: Dict[str, Any]) -> str:
    if not needs:
        return "не распознано"
    parts = []
    for k in ["type", "width", "color", "budget_max"]:
        if k in needs:
            parts.append(f"{k}={needs[k]}")
    return ", ".join(parts) if parts else "не распознано"

def pick_cta(contact_id: int, channel: str | None, stage: str = "intro") -> Dict[str, str]:
    opts = [
        "Зафиксирую лучшие условия и пришлю подбор сегодня. Подходит?",
        "Готов обсудить детали и оформить заказ без задержек. Продолжаем?",
        "Забронирую цену на сутки и подготовлю договор. Двигаемся?",
    ]
    return {"text": opts[hash(contact_id) % len(opts)]}


SPIN_TEMPLATES = {
    "s": [
        "Чтобы точно попасть в цель, подскажите, где будете использовать {focus} и на что делаете упор?",
        "Расскажите, какие модели нравились раньше и что хотите сохранить в {focus}?",
    ],
    "p": [
        "Что хотелось бы улучшить по сравнению с тем, что есть сейчас?",
        "Каких возможностей не хватает текущему решению — удобства, дизайна или сервиса?",
    ],
    "i": [
        "Если подобрать подходящую модель, какой результат почувствуете первым делом?",
        "Чем быстрее закроем вопрос с {focus}, что это даст вашей команде или дому?",
    ],
    "n": [
        "По каким двум критериям поймёте, что решение идеально подошло?",
        "Что должно случиться, чтобы вы сказали: «берём»?",
    ],
}

BANT_TEMPLATES = {
    "budget": [
        "В какой диапазон по {currency} хотите уложиться, чтобы я показал точные варианты?",
        "Подскажите предел по стоимости, чтобы держать баланс цена/качество?",
    ],
    "authority": [
        "Кого ещё стоит подключить к обсуждению, чтобы быстро согласовать заказ?",
        "Принимаете решение сами или подключим коллегу для финального слова?",
    ],
    "need": [
        "Что должно быть обязательно — тишина, дизайн, дополнительные опции?",
        "Какой ключевой результат хотите получить после установки?",
    ],
    "timeline": [
        "К какому сроку хотелось бы запустить поставку — на этой неделе, в течение месяца?",
        "Когда удобно получить решение, чтобы вписаться в ваши планы?",
    ],
}

PROBLEM_KEYWORDS = ("проблем", "сложн", "не устраивает", "жалоб", "минус", "трудно", "болит")
IMPLICATION_KEYWORDS = ("теря", "штраф", "простой", "срыв", "дороже", "рискуем", "потер")
NEED_PAYOFF_KEYWORDS = ("хочу", "нужно", "важно", "интересует", "ищу", "готов")
POSITIVE_KEYWORDS = ("давайте", "беру", "подходит", "соглас", "старт", "нравится")
NEGATIVE_KEYWORDS = ("дорого", "позже", "не сейчас", "сомнева", "подум", "не готов")
AUTHORITY_KEYWORDS = ("я решаю", "сам", "директор", "руковод", "владел", "согласую")

SENTIMENT_POSITIVE_HINTS = (
    "спасибо",
    "класс",
    "отлично",
    "супер",
    "идеально",
    "🔥",
    "😍",
)
SENTIMENT_NEGATIVE_HINTS = (
    "не нравится",
    "разочаров",
    "расстро",
    "недоволен",
    "плохо",
    "ужас",
    "проблема",
    "печаль",
    "😔",
    "😢",
)

EMPATHY_NEGATIVE_TEMPLATES = (
    "Понимаю, что ситуация неприятная — сосредотачиваюсь на надежных решениях для {focus}.",
    "Сожалею, что предыдущий опыт подвёл — подберу спокойные варианты по {focus}.",
)
EMPATHY_POSITIVE_TEMPLATES = (
    "Здорово, что вам откликается идея с {focus} — ускорю подбор.",
    "Классно слышать ваш энтузиазм по {focus}, покажу топовые позиции сразу.",
)


def analyze_sentiment_delta(text: str) -> float:
    raw = text or ""
    if not raw.strip():
        return 0.0

    lowered = raw.lower()
    score = 0.0

    for word in POSITIVE_KEYWORDS:
        if word in lowered:
            score += 0.7
    for hint in SENTIMENT_POSITIVE_HINTS:
        if hint in lowered or hint in raw:
            score += 0.6

    for word in NEGATIVE_KEYWORDS:
        if word in lowered:
            score -= 0.8
    for hint in SENTIMENT_NEGATIVE_HINTS:
        if hint in lowered or hint in raw:
            score -= 0.7

    exclamation_bonus = raw.count("!") * 0.1
    score += min(exclamation_bonus, 0.3)

    question_penalty = raw.count("?") * 0.05
    score -= min(question_penalty, 0.2)

    return max(-3.0, min(3.0, score))

TIMELINE_PATTERNS: List[Tuple[re.Pattern[str], str]] = [
    (re.compile(r"сегодня|сейчас|в ближайшие сутки", re.IGNORECASE), "сегодня"),
    (re.compile(r"завтра", re.IGNORECASE), "завтра"),
    (re.compile(r"(на|в течение) этой недели", re.IGNORECASE), "на этой неделе"),
    (re.compile(r"следующ(ий|ая) (недел|месяц)", re.IGNORECASE), "в следующем периоде"),
    (re.compile(r"до\s+(конца|\d{1,2})", re.IGNORECASE), "до указанного срока"),
]

SOCIAL_PROOF_TEMPLATES = [
    "{brand} оформил десятки заказов по этой категории в {city} — клиенты отмечают стабильные сроки и качество.",
    "9 из 10 покупателей в {city} возвращаются за повторными заказами — поделюсь отзывами по запросу.",
    "Эти модели чаще берут команды из {city}, когда нужна надёжность без переплаты.",
]

SCARCITY_TEMPLATES = [
    "На складе сейчас {stock} комплектов в нужной комплектации. Держу резерв на сутки, если подтверждаем.",
    "Ближайшая поставка — {slot}. Зафиксирую слот, чтобы не потерять очередь.",
    "Популярные модели быстро разбирают. Могу забрать под вас остаток на 24 часа.",
]

RECIPROCITY_TEMPLATES = [
    "Вышлю чек-лист по установке и подготовлю бонус на дополнительные аксессуары.",
    "Поделюсь шаблоном коммерческого предложения и памяткой по монтажу, чтобы пройти путь без лишних шагов.",
]

UPSELL_TEMPLATES = [
    "При желании добавим комплект фурнитуры и сервис — получите решение под ключ.",
    "Могу предложить расширенную гарантию и послепродажную поддержку, чтобы не думать о сервисе.",
]

CHALLENGER_PLAYBOOK = {
    "default": [
        {
            "teach": "Собрал короткий шорт-лист по {focus}: только позиции с лучшими отзывами и наличием.",
            "tailor": "Смотрю на реальные сроки для {city}, чтобы запуститься без задержек.",
            "control": "Готов отправить финальные цены и бонусы. Подойдёт, если сразу перейдём к оформлению?",
        },
        {
            "teach": "По этой категории клиенты в {city} чаще выбирают модели, где сочетаются дизайн и шумоизоляция.",
            "tailor": "Я оставил варианты, которые точно впишутся в ваш запрос и бюджет.",
            "control": "Если такой формат ок, резервирую условия и собираю документы. Продолжаем?",
        },
        {
            "teach": "Отслеживаю наличие по {focus} ежедневно — сейчас есть позиции, которые можно отгрузить сразу.",
            "tailor": "В подборку попали решения с проверенной логистикой и поддержкой.",
            "control": "Готов закрепить цену и отправить договор. Подтвердите — сделаю резерв.",
        },
    ],
}


class SalesConversationEngine:
    def __init__(
        self,
        state: SalesState,
        branding: Dict[str, str],
        tenant_cfg: Dict[str, Any],
        channel_name: str,
        persona_hints: Optional[PersonaHints] = None,
    ) -> None:
        self.state = state
        self.branding = branding
        self.cfg = tenant_cfg if isinstance(tenant_cfg, dict) else {}
        self.channel_name = channel_name.strip() or branding.get("CHANNEL", "WhatsApp")
        self.persona_hints = persona_hints or PersonaHints()

    # ------------------------ анализ входящего текста --------------------
    def observe_user(self, text: str) -> None:
        incoming = (text or "").strip()
        if not incoming:
            return
        if incoming == self.state.last_user_text:
            return
        self.state.last_user_text = incoming
        self.state.append_history("user", incoming)
        self.state.last_updated_ts = time.time()
        self.state.user_message_count += 1

        self._touch_profile()

        needs = infer_user_needs(incoming)
        if needs:
            for key, value in needs.items():
                if value:
                    self.state.needs[key] = value
        self._update_spin(incoming)
        self._update_bant(incoming)
        self._update_conversion_score(incoming)
        self._update_sentiment(incoming)
        self._update_profile_preferences()

    def _update_conversion_score(self, text: str) -> None:
        score = self.state.conversion_score * 0.9  # лёгкое затухание — RL-подход
        low = text.lower()
        if any(word in low for word in POSITIVE_KEYWORDS):
            score += 0.8
        if any(word in low for word in NEGATIVE_KEYWORDS):
            score -= 0.9
        score = max(-3.0, min(5.0, score))
        self.state.conversion_score = score

    def _update_spin(self, text: str) -> None:
        low = text.lower()
        if self.state.needs and self.state.spin.get("s") != "covered":
            self.state.mark_spin_stage("s", "covered")
        if any(word in low for word in PROBLEM_KEYWORDS):
            self.state.mark_spin_stage("p", "covered")
        if any(word in low for word in IMPLICATION_KEYWORDS):
            self.state.mark_spin_stage("i", "covered")
        if any(word in low for word in NEED_PAYOFF_KEYWORDS):
            self.state.mark_spin_stage("n", "covered")

    def _update_bant(self, text: str) -> None:
        budget = self._extract_budget(text)
        if budget:
            self.state.bant["budget"] = budget
            self.state.bant.pop("_asked_budget", None)
            self.state.conversion_score += 0.2

        if any(key in text.lower() for key in AUTHORITY_KEYWORDS):
            self.state.bant["authority"] = "decision_maker"
            self.state.bant.pop("_asked_authority", None)

        if any(word in text.lower() for word in NEED_PAYOFF_KEYWORDS):
            self.state.bant["need"] = True
            self.state.bant.pop("_asked_need", None)

        timeline = self._extract_timeline(text)
        if timeline:
            self.state.bant["timeline"] = timeline
            self.state.bant.pop("_asked_timeline", None)

    def _touch_profile(self) -> None:
        if not isinstance(self.state.profile, dict):
            self.state.profile = {}
        profile = self.state.profile
        now_ts = time.time()
        profile["last_seen_ts"] = now_ts
        day_bucket = int(now_ts // 86400)
        last_bucket = int(profile.get("last_visit_day", -1) or -1)
        if last_bucket != day_bucket:
            profile["visits"] = int(profile.get("visits", 0)) + 1
            profile["last_visit_day"] = day_bucket

        channels = profile.setdefault("channels", [])
        if self.channel_name and self.channel_name not in channels:
            channels.append(self.channel_name)
            if len(channels) > 6:
                profile["channels"] = channels[-6:]

    def _update_sentiment(self, text: str) -> None:
        delta = analyze_sentiment_delta(text)
        if delta == 0.0:
            self.state.sentiment_score *= 0.85
            return
        blended = self.state.sentiment_score * 0.6 + delta
        self.state.sentiment_score = max(-3.0, min(3.0, blended))

    def _update_profile_preferences(self) -> None:
        profile = self.state.profile
        if not isinstance(profile, dict):
            profile = {}
            self.state.profile = profile
        prefs = profile.setdefault("preferences", {})

        color = self.state.needs.get("color")
        if color:
            colors = prefs.setdefault("colors", [])
            if color not in colors:
                colors.append(color)
                if len(colors) > 5:
                    prefs["colors"] = colors[-5:]

        keywords = self.state.needs.get("keywords") or []
        if keywords:
            pref_keywords = prefs.setdefault("keywords", [])
            for kw in keywords[:4]:
                if kw not in pref_keywords:
                    pref_keywords.append(kw)
            if len(pref_keywords) > 10:
                prefs["keywords"] = pref_keywords[-10:]

        budget = self.state.needs.get("budget_max")
        if budget:
            try:
                budget_val = int(budget)
            except Exception:
                budget_val = None
            if budget_val:
                prev = prefs.get("budget_max")
                if not prev:
                    prefs["budget_max"] = budget_val
                else:
                    try:
                        prev_val = int(prev)
                    except Exception:
                        prefs["budget_max"] = budget_val
                    else:
                        prefs["budget_max"] = min(prev_val, budget_val)

    @staticmethod
    def _extract_budget(text: str) -> Optional[int]:
        raw = text.lower()
        matches = re.finditer(r"(\d+[\s\d]*)\s*(k|тыс|тысяч)?", raw)
        best: Optional[int] = None
        for m in matches:
            digits = m.group(1).replace(" ", "")
            try:
                val = int(digits)
            except Exception:
                continue
            if m.group(2):
                val *= 1000
            if val < 100:  # вероятно указали тысячи
                continue
            if not best or val > best:
                best = val
        return best

    @staticmethod
    def _extract_timeline(text: str) -> Optional[str]:
        for pattern, label in TIMELINE_PATTERNS:
            if pattern.search(text):
                return label
        return None

    # ------------------------ формирование предложения -------------------
    def register_recommendations(self, items: List[Dict[str, Any]]) -> None:
        if items:
            self.state.last_items = items[:]

    def _focus_phrase(self) -> str:
        focus = str(self.state.needs.get("focus") or "").strip()
        if focus:
            return focus
        tokens: List[str] = []
        if self.state.needs.get("type"):
            tokens.append(str(self.state.needs["type"]).strip())
        if self.state.needs.get("width"):
            tokens.append(f"{self.state.needs['width']} см")
        if self.state.needs.get("color"):
            tokens.append(f"цвет {self.state.needs['color']}")
        if self.state.needs.get("keywords"):
            tokens.extend(str(x) for x in self.state.needs["keywords"][:2])
        focus_line = " ".join(token for token in tokens if token).strip()
        return focus_line or "вашей задаче"

    def _active_listening_line(self, text: str) -> str:
        cleaned = re.sub(r"\s+", " ", (text or "").strip())
        if len(cleaned) > 120:
            cleaned = cleaned[:117] + "..."
        focus = self._focus_phrase()
        empathy = self._empathy_prefix()
        if cleaned:
            base = f"Понял запрос: {cleaned}. Держу в фокусе {focus}."
        else:
            base = f"Учитываю частые запросы по {focus} и сразу показываю сильные позиции."
        if empathy:
            return f"{empathy} {base}".strip()
        return base

    def _empathy_prefix(self) -> str:
        score = self.state.sentiment_score
        focus = self._focus_phrase()
        if score <= -0.5 and EMPATHY_NEGATIVE_TEMPLATES:
            idx = max(0, (self.state.user_message_count - 1)) % len(EMPATHY_NEGATIVE_TEMPLATES)
            return EMPATHY_NEGATIVE_TEMPLATES[idx].format(focus=focus)
        if score >= 1.2 and EMPATHY_POSITIVE_TEMPLATES:
            idx = max(0, (self.state.user_message_count - 1)) % len(EMPATHY_POSITIVE_TEMPLATES)
            return EMPATHY_POSITIVE_TEMPLATES[idx].format(focus=focus)
        visits = int((self.state.profile or {}).get("visits", 0))
        if visits > 1:
            return "Рада снова помочь — учитываю прошлые пожелания."
        return ""

    def _format_price(self, price: Optional[int], currency: str) -> str:
        if price is None:
            return "цена по запросу"
        return f"{price:,}".replace(",", " ") + f" {currency}"

    def _price_from_item(self, item: Dict[str, Any]) -> Optional[int]:
        raw = str(item.get("price") or "").strip()
        digits = re.sub(r"\D", "", raw)
        if not digits:
            return None
        try:
            return int(digits)
        except Exception:
            return None

    def _fab_line(self, item: Dict[str, Any], idx: int, currency: str) -> str:
        title = (
            item.get("title")
            or item.get("name")
            or item.get("sku")
            or item.get("id")
            or f"Позиция {idx}"
        )
        price_val = self._price_from_item(item)
        price_text = self._format_price(price_val, currency)

        highlight_bits = []
        if item.get("brand"):
            highlight_bits.append(f"бренд {item['brand']}")
        if item.get("material"):
            highlight_bits.append(f"материал {item['material']}")
        if item.get("color"):
            highlight_bits.append(f"цвет {item['color']}")
        if item.get("width"):
            highlight_bits.append(f"ширина {item['width']} см")

        stock_note: List[str] = []
        stock = item.get("stock")
        try:
            stock_val = int(str(stock)) if stock is not None and str(stock).strip() else None
        except Exception:
            stock_val = None
        if stock_val is not None and stock_val > 0:
            stock_note.append("в наличии, отправим без ожидания")
        if item.get("tags") and "хит" in str(item.get("tags")).lower():
            stock_note.append("хит продаж")

        benefit = self._benefit_hint(item)

        line = f"{idx}. {title} — {price_text}"
        if highlight_bits:
            line += f"; {', '.join(highlight_bits)}"
        if stock_note:
            line += f"; {', '.join(stock_note)}"
        line += f". {benefit}"

        url = (item.get("url") or "").strip()
        if url:
            line += f" [{url}]"
        return line

    def _benefit_hint(self, item: Dict[str, Any]) -> str:
        needs = self.state.needs
        if needs.get("budget_max"):
            return "Укладывается в ваш бюджет и закрывает задачу без скрытых доплат."
        keywords = needs.get("keywords") or []
        if keywords:
            return f"Помогает с {keywords[0]} и экономит время на выборе." if keywords else "Подходит под ваш запрос."
        focus = needs.get("focus") or needs.get("type")
        if focus:
            return f"Даёт готовое решение по направлению «{focus}»."
        return "Экономит время и даёт предсказуемый результат."

    def _fab_block(self, items: List[Dict[str, Any]], currency: str) -> str:
        if not items:
            return "Пока без точных позиций — готов подобрать после пары уточнений."
        lines = [self._fab_line(item, idx, currency) for idx, item in enumerate(items, start=1)]
        return "\n".join(lines)

    def _next_spin_question(self) -> Optional[str]:
        focus = self._focus_phrase()
        for stage in ("s", "p", "i", "n"):
            status = self.state.spin.get(stage, "pending")
            if status == "pending":
                template = random.choice(SPIN_TEMPLATES[stage])
                question = template.format(focus=focus)
                self.state.mark_spin_stage(stage, "asked")
                return question
        return None

    def _next_bant_question(self, currency: str) -> Optional[str]:
        order = ["budget", "need", "timeline", "authority"]
        focus = self._focus_phrase()
        for key in order:
            value = self.state.bant.get(key)
            asked_flag = self.state.bant.get(f"_asked_{key}")
            if value or asked_flag:
                continue
            template = random.choice(BANT_TEMPLATES[key])
            question = template.format(currency=currency, focus=focus, city=self.branding.get("CITY", ""))
            self.state.bant[f"_asked_{key}"] = True
            return question
        return None

    def _choose_question(self, currency: str, max_per_turn: int) -> Optional[str]:
        if max_per_turn <= 0:
            return None
        question = self._next_spin_question()
        if not question:
            question = self._next_bant_question(currency)
        if question and question not in self.state.asked_questions:
            self.state.asked_questions.append(question)
        return question

    def pending_question(self) -> Optional[str]:
        focus = self._focus_phrase()
        for stage in ("s", "p", "i", "n"):
            if self.state.spin.get(stage, "pending") == "pending":
                return SPIN_TEMPLATES[stage][0].format(focus=focus)
        for key in ("budget", "need", "timeline", "authority"):
            if not self.state.bant.get(key):
                template = BANT_TEMPLATES[key][0]
                return template.format(currency=self.branding.get("CURRENCY", "₽"), focus=focus, city=self.branding.get("CITY", ""))
        return None

    def _challenger_block(self) -> Tuple[str, str, str]:
        key = str(self.state.needs.get("type") or "default").lower()
        options = CHALLENGER_PLAYBOOK.get(key) or CHALLENGER_PLAYBOOK["default"]
        idx = self.state.challenger_cursor % len(options)
        play = options[idx]
        self.state.challenger_cursor += 1
        focus = self._focus_phrase()
        teach = play["teach"].format(city=self.branding.get("CITY", ""), focus=focus)
        tailor = play["tailor"].format(city=self.branding.get("CITY", ""), focus=focus)
        control = play["control"].format(city=self.branding.get("CITY", ""), focus=focus)
        return teach, tailor, control

    def _choose_social_proof(self, items: List[Dict[str, Any]]) -> Optional[str]:
        template = SOCIAL_PROOF_TEMPLATES[self.state.social_proof_cursor % len(SOCIAL_PROOF_TEMPLATES)]
        self.state.social_proof_cursor += 1
        return template.format(brand=self.branding.get("BRAND", "Бренд"), city=self.branding.get("CITY", ""))

    def _choose_scarcity(self, items: List[Dict[str, Any]]) -> Optional[str]:
        stock_values = []
        for it in items:
            stock = it.get("stock")
            try:
                val = int(str(stock))
            except Exception:
                continue
            if val >= 0:
                stock_values.append(val)
        template = SCARCITY_TEMPLATES[self.state.scarcity_cursor % len(SCARCITY_TEMPLATES)]
        self.state.scarcity_cursor += 1
        slot = "завтра"
        if self.state.bant.get("timeline"):
            slot = self.state.bant["timeline"]
        stock_text = "несколько" if not stock_values else (str(min(stock_values)) if min(stock_values) > 0 else "последние")
        return template.format(stock=stock_text, city=self.branding.get("CITY", ""), slot=slot)

    def _choose_reciprocity(self) -> Optional[str]:
        template = RECIPROCITY_TEMPLATES[self.state.reciprocity_cursor % len(RECIPROCITY_TEMPLATES)]
        self.state.reciprocity_cursor += 1
        return template

    def _choose_upsell(self) -> str:
        template = random.choice(UPSELL_TEMPLATES)
        integrations = self.cfg.get("integrations", {}) if isinstance(self.cfg, dict) else {}
        if integrations.get("pdf_catalog_url"):
            template += " Могу отправить PDF/Excel с полным каталогом — скажите формат."
        return template

    def _choose_cta(self, cta_primary: str, cta_fallback: str) -> str:
        if self.persona_hints.cta:
            return self.persona_hints.cta
        score = self.state.conversion_score
        timeline = self.state.bant.get("timeline")
        handoff = (self.cfg.get("cta") or {}).get("handoff_wa") if isinstance(self.cfg, dict) else ""
        sentiment = self.state.sentiment_score
        if sentiment <= -1.2:
            soothing = "Могу предложить более мягкий вариант или подключить эксперта по нюансам. Продолжим подбор?"
            return soothing
        if score >= 1.5 and timeline:
            return f"Зафиксирую запуск на {timeline}. Подтверждаем — оформляю?"
        if sentiment >= 1.2 and score >= 1.0:
            return "Готов сразу оформить заказ и зафиксировать цену. Даем старт?"
        if score <= -1:
            return "Могу предложить более бюджетный пакет или рассрочку. Продолжаем подбор?"
        if self.channel_name.lower() == "avito" and handoff:
            return handoff
        candidate = cta_primary or cta_fallback or "Готов подключиться и довести до заказа — двигаемся?"
        if self.persona_hints.wants_friendly() and candidate and not re.search(r"[)☺😊🙂]$", candidate):
            candidate = candidate.rstrip(".") + ". 🙂"
        return candidate

    def _personalized_greeting(self) -> str:
        default_greeting = f"Здравствуйте! Меня зовут {self.branding.get('AGENT_NAME', 'Менеджер')}, {self.branding.get('BRAND', '')}."
        greeting = (self.persona_hints.greeting or default_greeting).strip()
        if not greeting:
            greeting = default_greeting
        friendly = self.persona_hints.wants_friendly()
        if friendly and greeting and not re.search(r"[)☺😊🙂]$", greeting):
            greeting = greeting.rstrip(".") + " 🙂"
        visits = int((self.state.profile or {}).get("visits", 0))
        if visits > 1:
            addon = "Рады снова вас видеть и продолжить подбор."
            if friendly:
                addon = addon.rstrip(".") + " 🙂"
            if addon not in greeting:
                if greeting.endswith(('.', '!', '…')):
                    greeting = f"{greeting.rstrip('.')}. {addon}"
                else:
                    greeting = f"{greeting}. {addon}"
        return greeting

    def _loyalty_line(self) -> Optional[str]:
        profile = self.state.profile if isinstance(self.state.profile, dict) else {}
        prefs = profile.get("preferences") or {}
        pieces: List[str] = []
        colors = prefs.get("colors") or []
        if colors:
            pieces.append(f"цвету {colors[-1]}")
        keywords = prefs.get("keywords") or []
        if keywords:
            pieces.append(f"темам «{keywords[-1]}»")
        budget = prefs.get("budget_max")
        if budget:
            pieces.append(
                f"бюджету до {self._format_price(int(budget), self.branding.get('CURRENCY', '₽'))}"
            )

        if pieces:
            joined = ", ".join(pieces)
            return f"Помню ваши предпочтения по {joined} — покажу то, что действительно откликается."

        visits = int(profile.get("visits", 0) or 0)
        if visits > 1:
            return "Учитываю прошлые диалоги и подберу обновлённые варианты."
        return None

    def build_reply(
        self,
        items: List[Dict[str, Any]],
        cta_primary: str,
        cta_fallback: str,
        currency: str,
        last_user_text: str,
    ) -> str:
        self.register_recommendations(items)
        max_questions_cfg = int((self.cfg.get("behavior", {}) or {}).get("max_clarifying_questions", 1))
        if self.persona_hints.max_questions is not None:
            try:
                max_questions_cfg = max(0, int(self.persona_hints.max_questions))
            except Exception:
                pass
        question_line = self._choose_question(currency, max_questions_cfg)
        teach, tailor, control = self._challenger_block()
        listening_line = self._active_listening_line(last_user_text)
        fab_block = self._fab_block(items, currency)
        social_proof = self._choose_social_proof(items)
        scarcity = self._choose_scarcity(items)
        reciprocity = self._choose_reciprocity()
        upsell = self._choose_upsell()
        cta_line = self._choose_cta(cta_primary, cta_fallback)

        greeting = self._personalized_greeting()
        loyalty_line = self._loyalty_line()

        action_lines = [control]
        if question_line:
            action_lines.append(question_line)
        if cta_line:
            action_lines.append(cta_line)

        actions_block = "\n".join(line for line in action_lines if line)

        message_parts = {
            "greeting": greeting,
            "teach": teach,
            "listening": listening_line,
            "tailor": tailor,
            "loyalty": loyalty_line or "",
            "fab": fab_block,
            "social": social_proof,
            "scarcity": scarcity,
            "upsell": upsell,
            "reciprocity": reciprocity,
            "actions": actions_block,
            "closing": self.persona_hints.closing or "",
        }

        ordered_keys = [
            "greeting",
            "teach",
            "listening",
            "loyalty",
            "tailor",
            "fab",
            "social",
            "scarcity",
            "upsell",
            "reciprocity",
            "actions",
            "closing",
        ]

        cleaned = [message_parts[key].strip() for key in ordered_keys if message_parts[key] and message_parts[key].strip()]
        if self.persona_hints.wants_short():
            prioritized = [
                message_parts.get("greeting", ""),
                message_parts.get("listening", ""),
                message_parts.get("loyalty", ""),
                message_parts.get("fab", ""),
                message_parts.get("actions", ""),
                message_parts.get("closing", ""),
            ]
            cleaned = [part.strip() for part in prioritized if part and part.strip()]

        reply = "\n\n".join(cleaned)
        self.state.last_bot_reply = reply
        self.state.append_history("assistant", reply)
        self.state.last_updated_ts = time.time()
        return reply

    def summary_for_llm(self) -> str:
        needs_summary = format_needs_for_prompt(self.state.needs)
        bant_parts = []
        for key in ("budget", "need", "timeline", "authority"):
            if key in self.state.bant and not str(key).startswith("_"):
                bant_parts.append(f"{key}={self.state.bant[key]}")
        if not bant_parts:
            bant_parts.append("недостаточно данных")
        spin_parts = [f"{stage.upper()}={self.state.spin.get(stage, 'pending')}" for stage in ("s", "p", "i", "n")]
        pending_question = self.pending_question()
        summary = [
            f"Needs: {needs_summary}",
            f"BANT: {', '.join(bant_parts)}",
            f"SPIN: {', '.join(spin_parts)}",
            f"Score={round(self.state.conversion_score, 2)}",
        ]
        if pending_question:
            summary.append(f"Следующий вопрос: {pending_question}")
        return "; ".join(summary)


def observe_user_message(
    contact_id: int,
    tenant: int | None,
    channel: str | None,
    text: str,
    tenant_cfg: Optional[dict] = None,
    branding: Optional[Dict[str, str]] = None,
    persona_hints: Optional[PersonaHints] = None,
) -> SalesState:
    cfg = tenant_cfg
    if cfg is None:
        cfg = load_tenant(tenant or 0)
    brand = branding or _branding_for_tenant(tenant)
    state = load_sales_state(tenant, contact_id)
    hints = persona_hints or load_persona_hints(tenant)
    engine = SalesConversationEngine(state, brand, cfg, channel or brand["CHANNEL"], persona_hints=hints)
    engine.observe_user(text or "")
    save_sales_state(state)
    return state


def summarize_sales_state(
    contact_id: int,
    tenant: int | None,
    channel: str | None,
    tenant_cfg: Optional[dict] = None,
    branding: Optional[Dict[str, str]] = None,
) -> str:
    cfg = tenant_cfg if tenant_cfg is not None else load_tenant(tenant or 0)
    brand = branding or _branding_for_tenant(tenant)
    state = load_sales_state(tenant, contact_id)
    hints = load_persona_hints(tenant)
    engine = SalesConversationEngine(state, brand, cfg, channel or brand["CHANNEL"], persona_hints=hints)
    return engine.summary_for_llm()


def record_bot_reply(
    contact_id: int,
    tenant: int | None,
    channel: str | None,
    reply: str,
    tenant_cfg: Optional[dict] = None,
    branding: Optional[Dict[str, str]] = None,
) -> None:
    cfg = tenant_cfg if tenant_cfg is not None else load_tenant(tenant or 0)
    brand = branding or _branding_for_tenant(tenant)
    state = load_sales_state(tenant, contact_id)
    hints = load_persona_hints(tenant)
    engine = SalesConversationEngine(state, brand, cfg, channel or brand["CHANNEL"], persona_hints=hints)
    if reply:
        state.last_bot_reply = reply.strip()
        state.append_history("assistant", reply.strip())
        state.last_updated_ts = time.time()
    save_sales_state(state)
def make_rule_based_reply(
    last_user_text: str,
    channel: str | None,
    contact_id: int,
    tenant: int | None = None,
) -> str:
    branding = _branding_for_tenant(tenant)
    channel_name = (channel or branding["CHANNEL"]).strip() or "WhatsApp"

    cfg = json.loads(json.dumps(DEFAULT_TENANT_JSON, ensure_ascii=False))
    if tenant is not None:
        try:
            cfg = load_tenant(tenant)
        except Exception:
            cfg = json.loads(json.dumps(DEFAULT_TENANT_JSON, ensure_ascii=False))

    persona_hints = load_persona_hints(tenant)
    state = load_sales_state(tenant, contact_id)
    engine = SalesConversationEngine(state, branding, cfg, channel_name, persona_hints=persona_hints)
    engine.observe_user(last_user_text or "")

    needs = state.needs if state.needs else infer_user_needs(last_user_text or "")
    currency = branding["CURRENCY"]
    items = search_catalog(needs, limit=4, tenant=tenant, query=last_user_text)

    cta_cfg = cfg.get("cta", {}) if isinstance(cfg, dict) else {}
    cta_primary = (cta_cfg.get("primary") or pick_cta(contact_id, channel_name).get("text") or "").strip()
    cta_fallback = (cta_cfg.get("fallback") or '').strip()

    reply = engine.build_reply(items, cta_primary, cta_fallback, currency, last_user_text or "")
    save_sales_state(state)
    return reply


# ----------------------- интерфейс для main.py -------------------------------
async def build_llm_messages(
    contact_id: int,
    last_user_text: str,
    channel: str | None = None,
    tenant: int | None = None,
):
    """Собираем системный промпт с учётом брендинга арендатора."""
    persona = load_persona(tenant)
    persona_hints = extract_persona_hints(persona)
    cache_key: int | None
    try:
        cache_key = int(tenant) if tenant is not None else None
    except Exception:
        cache_key = None
    fingerprint = hashlib.sha1(persona.encode("utf-8")).hexdigest() if persona else ""
    _PERSONA_HINTS_CACHE[cache_key] = (fingerprint, persona_hints)
    branding = _branding_for_tenant(tenant)

    cfg = json.loads(json.dumps(DEFAULT_TENANT_JSON, ensure_ascii=False))
    if tenant is not None:
        try:
            cfg = load_tenant(tenant)
        except Exception:
            pass

    state = observe_user_message(
        contact_id,
        tenant,
        channel or branding["CHANNEL"],
        last_user_text or "",
        tenant_cfg=cfg,
        branding=branding,
        persona_hints=persona_hints,
    )
    engine = SalesConversationEngine(state, branding, cfg, channel or branding["CHANNEL"], persona_hints=persona_hints)
    summary = engine.summary_for_llm()

    cta_cfg = cfg.get("cta", {}) if isinstance(cfg, dict) else {}
    limits_cfg = cfg.get("limits", {}) if isinstance(cfg, dict) else {}

    try:
        catalog_window = int(limits_cfg.get("catalog_page_size", 8))
    except Exception:
        catalog_window = 8
    preview_limit = min(12, max(4, catalog_window))
    needs_snapshot: Dict[str, Any] = dict(state.needs) if state.needs else {}
    if not needs_snapshot and last_user_text:
        needs_snapshot = infer_user_needs(last_user_text)
    context_items = search_catalog(
        needs_snapshot,
        limit=preview_limit,
        tenant=tenant,
        query=last_user_text,
    )
    if context_items:
        engine.register_recommendations(context_items)

    system_blocks = [persona.strip()]
    system_blocks.append(
        " | ".join(
            filter(
                None,
                [
                    f"Бренд: {branding['BRAND']} ({branding['CITY']})",
                    f"Канал: {channel or branding['CHANNEL']}",
                    f"CTA: {cta_cfg.get('primary') or 'держи жёсткий CTA в конце'}",
                    f"Каталог на ответ: {limits_cfg.get('catalog_page_size', 8)} позиций",
                ],
            )
        )
    )
    system_blocks.append(summary)

    if context_items:
        catalog_block = format_items_for_prompt(context_items, branding["CURRENCY"])
        system_blocks.append(f"Релевантные позиции каталога:\n{catalog_block}")

    # Добавим обучающие примеры диалогов (1–2) из базы арендатора
    if training_retriever and tenant is not None and (last_user_text or "").strip():
        try:
            block = training_retriever.build_examples_block(int(tenant), last_user_text)
        except Exception:
            block = ""
        if block.strip():
            system_blocks.append(block)

    history_tail = [item for item in (state.history[-6:] if state.history else []) if item.get("role") in {"user", "assistant"}]
    if history_tail:
        trimmed = history_tail[:-1] if history_tail and history_tail[-1].get("role") == "user" else history_tail
        if trimmed:
            transcript = "\n".join(f"{msg['role']}: {msg['content']}" for msg in trimmed)
            if transcript.strip():
                system_blocks.append(f"Недавний диалог:\n{transcript}")

    system_blocks.append(f"Идентификатор контакта: {contact_id}")

    sys = "\n\n".join(block for block in system_blocks if block)
    messages: List[Dict[str, str]] = [{"role": "system", "content": sys}]

    if history_tail:
        trimmed = history_tail[:-1] if history_tail and history_tail[-1].get("role") == "user" else history_tail
        for msg in trimmed:
            messages.append({"role": msg["role"], "content": msg["content"]})

    messages.append({"role": "user", "content": (last_user_text or "")})
    return messages


async def ask_llm(
    messages: List[Dict[str, str]],
    tenant: int | None = None,
    contact_id: int | None = None,
    channel: str | None = None,
) -> str:
    """
    Если задан OPENAI_API_KEY — спросим модель. Если нет — сгенерируем быстрый rule-based ответ.
    """
    # Попытка понять последний запрос и канал
    last = ""
    for m in reversed(messages):
        if m.get("role") == "user":
            last = m.get("content") or ""
            break

    channel_name = (channel or "whatsapp")
    contact_ref = int(contact_id or 0)

    # Без ключа — быстрый локальный ответ
    client = _get_openai_client()
    if client is None:
        return make_rule_based_reply(last, channel_name, contact_ref, tenant=tenant)

    try:


        openai.api_key = settings.OPENAI_API_KEY  # type: ignore

        persona_hints = load_persona_hints(tenant)
        try:
            plan, answer = await planner.generate_sales_reply(
                messages,
                openai_module=client,
                model=settings.OPENAI_MODEL,
                timeout=settings.OPENAI_TIMEOUT_SECONDS,
                persona_language=persona_hints.language if persona_hints and persona_hints.language else None,
            )
            state = load_sales_state(tenant, contact_ref)
            state.last_plan = plan.to_dict()
            save_sales_state(state)
            refined = quality.enforce_plan_alignment(answer, plan, persona_hints)
            record_bot_reply(contact_ref, tenant, channel_name, refined)
            return refined
        except planner.PlannerError as exc:  # type: ignore[attr-defined]
            logger.warning("planner failed: %s", exc)
        except Exception as exc:
            logger.exception("llm planner error", exc_info=exc)

        # Попробуем прямой ответ модели, затем фоллбек на правила
        try:

            create_fn = _resolve_chat_completion_callable(client)
            if not create_fn:
                raise RuntimeError("openai client missing chat.completions.create")

            resp = await asyncio.to_thread(
                create_fn,
                model=settings.OPENAI_MODEL,
                messages=messages,
                max_tokens=220,
                temperature=0.6,
                timeout=settings.OPENAI_TIMEOUT_SECONDS,
            )
            answer = resp.choices[0].message.content.strip()  # type: ignore
            record_bot_reply(contact_ref, tenant, channel_name, answer)
            return answer
        except Exception as exc:
            logger.exception("direct llm call failed", exc_info=exc)
            return make_rule_based_reply(last, channel_name, contact_ref, tenant=tenant)

    except Exception as exc:
        logger.exception("ask_llm unexpected error", exc_info=exc)
        return make_rule_based_reply(last, channel_name, contact_ref, tenant=tenant)


__all__ = [
    "Settings", "settings",
    "ADMIN_COOKIE",
    "get_tenant_pubkey", "set_tenant_pubkey",
    "http_json",
    "tenant_dir", "ensure_tenant_files",
    "read_tenant_config", "write_tenant_config",
    "read_persona", "write_persona",
    "load_tenant", "load_persona", "PersonaHints", "extract_persona_hints", "load_persona_hints",
    "build_llm_messages", "ask_llm",
    # helpers ниже могут понадобиться в других частях
    "infer_user_needs", "search_catalog", "format_needs_for_prompt",
    "format_items_for_prompt", "pick_cta",
    "load_sales_state", "save_sales_state", "observe_user_message",
    "record_bot_reply", "summarize_sales_state",
    "read_all_catalog", "paginate_catalog_text",
]
